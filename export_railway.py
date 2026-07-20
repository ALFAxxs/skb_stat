"""
Temir yo'lchi bemorlar va ularga biriktirilgan ma'lumotlarni Excel ga eksport qilish.

Ishlatish:
    python export_railway.py
"""
import os, sys, django

sys.path.insert(0, os.path.dirname(__file__))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'conf.settings')
django.setup()

from django.utils import timezone
from django.db.models import Prefetch
import datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from apps.patients.models import PatientCard, TreatmentProcedure, Prescription
from apps.services.models import PatientService

# ─── Sana oralig'i ───────────────────────────────────────────────────────────
DATE_FROM = timezone.make_aware(datetime.datetime(2026, 6, 25))
DATE_TO   = timezone.make_aware(datetime.datetime(2026, 7, 15, 23, 59, 59))

# ─── Bemorlar ────────────────────────────────────────────────────────────────
patients = PatientCard.objects.filter(
    patient_category='railway',
    created_at__gte=DATE_FROM,
    created_at__lte=DATE_TO,
).order_by('created_at')

print(f"Topildi: {patients.count()} ta temir yo'lchi bemor")

# ─── Excel ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 18


# ══════════════════════════════════════════════════════════════
# 1-varaq: BEMORLAR
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Bemorlar"
headers1 = [
    "ID", "Tibbiy karta №", "FIO", "Tug'ilgan sana", "Jins",
    "JSHSHIR", "Telefon", "Kategoriya", "Tashrif turi", "Statusi",
    "Bo'lim", "Manzil", "Ro'yxatga olingan",
]
write_header(ws1, headers1)

for i, p in enumerate(patients, 2):
    ws1.append([
        p.pk,
        getattr(p, 'medical_record_number', ''),
        p.full_name,
        p.birth_date.strftime('%d.%m.%Y') if p.birth_date else '',
        p.get_gender_display() if p.gender else '',
        p.JSHSHIR,
        p.phone,
        p.get_patient_category_display(),
        p.get_visit_type_display() if p.visit_type else '',
        p.get_status_display(),
        str(p.department) if p.department_id else '',
        p.street_address,
        timezone.localtime(p.created_at).strftime('%d.%m.%Y %H:%M'),
    ])

for col in ws1.columns:
    ws1.column_dimensions[col[0].column_letter].width = 18


# ══════════════════════════════════════════════════════════════
# 2-varaq: XIZMATLAR
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Xizmatlar")
headers2 = [
    "Bemor ID", "FIO", "JSHSHIR",
    "Xizmat kodi", "Xizmat nomi", "Miqdori", "Narxi", "Jami",
    "Statusi", "Qo'shilgan sana",
]
write_header(ws2, headers2)

services = PatientService.objects.filter(
    patient_card__patient_category='railway',
    patient_card__created_at__gte=DATE_FROM,
    patient_card__created_at__lte=DATE_TO,
).select_related('patient_card', 'service').order_by('patient_card__full_name')

for ps in services:
    p = ps.patient_card
    price = getattr(ps, 'price', None) or 0
    qty   = ps.quantity or 1
    ws2.append([
        p.pk,
        p.full_name,
        p.JSHSHIR,
        getattr(ps.service, 'code', ''),
        ps.service.name if ps.service_id else '',
        qty,
        float(price),
        float(price) * qty,
        ps.get_status_display(),
        timezone.localtime(ps.created_at).strftime('%d.%m.%Y %H:%M') if ps.created_at else '',
    ])

for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 18


# ══════════════════════════════════════════════════════════════
# 3-varaq: DORILAR (TreatmentProcedure)
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Dorilar")
headers3 = [
    "Bemor ID", "FIO", "JSHSHIR",
    "Dori nomi", "Dozasi", "Miqdor", "Manba",
    "Statusi", "Tayinlagan shifokor", "Tayinlangan sana",
]
write_header(ws3, headers3)

procedures = TreatmentProcedure.objects.filter(
    patient_card__patient_category='railway',
    patient_card__created_at__gte=DATE_FROM,
    patient_card__created_at__lte=DATE_TO,
).select_related('patient_card', 'assigned_by').order_by('patient_card__full_name')

for tp in procedures:
    p = tp.patient_card
    ws3.append([
        p.pk,
        p.full_name,
        p.JSHSHIR,
        tp.medicine_name,
        tp.dosage,
        tp.quantity,
        tp.get_source_display() if tp.source else '',
        tp.get_status_display(),
        str(tp.assigned_by) if tp.assigned_by_id else '',
        timezone.localtime(tp.created_at).strftime('%d.%m.%Y %H:%M') if tp.created_at else '',
    ])

for col in ws3.columns:
    ws3.column_dimensions[col[0].column_letter].width = 18


# ══════════════════════════════════════════════════════════════
# 4-varaq: RETSEPTLAR (Prescription)
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Retseptlar")
headers4 = [
    "Bemor ID", "FIO", "JSHSHIR",
    "Dori nomi", "Chiqarish shakli", "Doza", "Qabul qilish",
    "Davomiyligi (kun)", "Boshlanish sanasi", "Yozgan shifokor",
]
write_header(ws4, headers4)

prescriptions = Prescription.objects.filter(
    patient_card__patient_category='railway',
    patient_card__created_at__gte=DATE_FROM,
    patient_card__created_at__lte=DATE_TO,
).select_related('patient_card', 'doctor').order_by('patient_card__full_name')

for rx in prescriptions:
    p = rx.patient_card
    ws4.append([
        p.pk,
        p.full_name,
        p.JSHSHIR,
        rx.drug_name,
        rx.dosage_form,
        rx.dose,
        f"{rx.frequency_num} {rx.frequency_unit}" if rx.frequency_num else '',
        rx.duration_days,
        rx.date_start.strftime('%d.%m.%Y') if rx.date_start else '',
        str(rx.doctor) if rx.doctor_id else '',
    ])

for col in ws4.columns:
    ws4.column_dimensions[col[0].column_letter].width = 18


# ─── Saqlash ─────────────────────────────────────────────────────────────────
filename = f"temir_yolchi_bemorlar_25.06-15.07.2026.xlsx"
wb.save(filename)
print(f"✅ Fayl saqlandi: {filename}")
print(f"   Bemorlar:   {patients.count()} ta")
print(f"   Xizmatlar:  {services.count()} ta")
print(f"   Dorilar:    {procedures.count()} ta")
print(f"   Retseptlar: {prescriptions.count()} ta")
