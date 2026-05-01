# apps/statistic/report_export.py
# To'liq hisobotlar — bitta Excel faylda 8 ta sheet

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from datetime import date
from dateutil.relativedelta import relativedelta

from apps.patients.models import PatientCard
from .exports import get_filtered_queryset


# ==================== STILLAR ====================

def _styles():
    brd = Side(style='thin', color='CCCCCC')
    return {
        'BRD': Border(left=brd, right=brd, top=brd, bottom=brd),
        'BLUE':  PatternFill('solid', fgColor='1F4E79'),
        'LBLUE': PatternFill('solid', fgColor='D6E4F0'),
        'GREEN': PatternFill('solid', fgColor='145A32'),
        'LGREEN':PatternFill('solid', fgColor='E9F7EF'),
        'AMBER': PatternFill('solid', fgColor='7D6608'),
        'LAMBER':PatternFill('solid', fgColor='FEF9E7'),
        'GRAY':  PatternFill('solid', fgColor='F2F3F4'),
        'WHITE': PatternFill('solid', fgColor='FFFFFF'),
        'WF':    Font(color='FFFFFF', bold=True, size=10),
        'BOLD':  Font(bold=True, size=10),
        'NORM':  Font(size=9),
        'C': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'L': Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'R': Alignment(horizontal='right',  vertical='center'),
    }


def _set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _hdr(ws, row, title, ncols, color='1F4E79'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = PatternFill('solid', fgColor=color)
    c.font = Font(color='FFFFFF', bold=True, size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28
    return row + 1


def _info(ws, row, label, value, ncols):
    ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    ws.cell(row=row, column=2, value=str(value) if value else '—').font = Font(size=9)
    ws.row_dimensions[row].height = 16
    return row + 1


def _col_hdrs(ws, row, headers, widths, S, color='1F4E79'):
    _set_cols(ws, widths)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = PatternFill('solid', fgColor=color)
        c.font = S['WF']; c.alignment = S['C']; c.border = S['BRD']
    ws.row_dimensions[row].height = 30
    return row + 1


def _row(ws, r, vals, S, aligns=None, nums=None, even=True):
    fill = S['GRAY'] if even else S['WHITE']
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.font = S['NORM']; c.border = S['BRD']; c.fill = fill
        al = aligns[ci-1] if aligns and ci <= len(aligns) else 'left'
        c.alignment = Alignment(horizontal=al, vertical='center', wrap_text=True)
        if nums and ci in nums:
            c.number_format = '#,##0'
    ws.row_dimensions[r].height = 17
    return r + 1


def _total(ws, r, vals, S, color='1F4E79'):
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=ci, value=val)
        c.fill = PatternFill('solid', fgColor=color)
        c.font = S['WF']; c.border = S['BRD']
        c.alignment = Alignment(
            horizontal='right' if isinstance(val,(int,float)) and val else 'left',
            vertical='center')
        if isinstance(val,(int,float)) and val > 100:
            c.number_format = '#,##0'
    ws.row_dimensions[r].height = 22


def _filter_text(request):
    MONTHS = {str(i): m for i, m in enumerate(
        ['Yanvar','Fevral','Mart','Aprel','May','Iyun',
         'Iyul','Avgust','Sentabr','Oktabr','Noyabr','Dekabr'], 1)}
    CATS = {'railway':"Temir yo'lchi", 'paid':'Pullik', 'non_resident':'Norezident'}
    parts = []
    if request.GET.get('year'):   parts.append(f"Yil: {request.GET['year']}")
    if request.GET.get('month'):  parts.append(f"Oy: {MONTHS.get(request.GET['month'], request.GET['month'])}")
    if request.GET.get('date_from'): parts.append(f"Dan: {request.GET['date_from']}")
    if request.GET.get('date_to'):   parts.append(f"Gacha: {request.GET['date_to']}")
    if request.GET.get('patient_category'):
        parts.append(f"Kategoriya: {CATS.get(request.GET['patient_category'], '')}")
    return ' | '.join(parts) if parts else 'Barcha davr'


def _age(patient):
    if not patient.birth_date: return None
    today = date.today()
    return relativedelta(today, patient.birth_date).years


def _workplace(patient):
    if patient.workplace_org: return str(patient.workplace_org)
    return patient.workplace or '—'


# ==================== SHEETLAR ====================

def _sheet_statsionar(wb, qs, S, filter_text):
    """1-Sheet: Statsionar bemorlar ro'yxati"""
    from django.db.models import Count
    ws = wb.create_sheet("1. Statsionar bemorlar")
    ncols = 16
    r = _hdr(ws, 1, "STATSIONAR BEMORLAR RO'YXATI", ncols)
    r = _info(ws, r, "Davr:", filter_text, ncols)
    r = _info(ws, r, "Jami:", qs.filter(visit_type='inpatient').count(), ncols)

    headers = [
        '№', 'F.I.Sh', "Tug'ilgan sana", 'Yosh', 'Jinsi',
        'Bayonnoma', 'JSHSHIR', 'Yashash manzili',
        'Qabul sanasi', 'Chiqish sanasi', "Bo'lim", 'Shifokor',
        'Tashxis', 'Bemor turi', 'Yotoq kun', 'Natija'
    ]
    widths = [5, 28, 13, 6, 7, 14, 16, 30, 13, 13, 18, 20, 30, 13, 9, 14]
    aligns = ['center','left','center','center','center',
              'center','center','left','center','center',
              'left','left','left','center','center','left']
    r = _col_hdrs(ws, r, headers, widths, S)

    patients = qs.filter(visit_type='inpatient').select_related(
        'department','attending_doctor','region','district'
    ).order_by('admission_date')

    for i, p in enumerate(patients, 1):
        age = _age(p)
        vals = [
            i, p.full_name,
            p.birth_date.strftime('%d.%m.%Y') if p.birth_date else '—',
            age or '—', p.get_gender_display() if p.gender else '—',
            p.medical_record_number or '—',
            p.JSHSHIR or '—',
            ', '.join(filter(None, [str(p.region) if p.region else '', str(p.district) if p.district else '', p.street_address or ''])) or '—',
            p.admission_date.strftime('%d.%m.%Y') if p.admission_date else '—',
            p.discharge_date.strftime('%d.%m.%Y') if p.discharge_date else '—',
            str(p.department) if p.department else '—',
            str(p.attending_doctor) if p.attending_doctor else '—',
            p.admission_diagnosis or '—',
            p.get_patient_category_display(),
            p.days_in_hospital or 0,
            p.get_outcome_display() if p.outcome else '—',
        ]
        r = _row(ws, r, vals, S, aligns, nums={15}, even=(i%2==0))

    _total(ws, r, [f'JAMI: {patients.count()} ta bemor'] + ['']*(ncols-1), S)


def _sheet_ambulatory(wb, qs, S, filter_text):
    """2-Sheet: Ambulator bemorlar ro'yxati"""
    ws = wb.create_sheet("2. Ambulator bemorlar")
    ncols = 12
    r = _hdr(ws, 1, "AMBULATOR BEMORLAR RO'YXATI", ncols, '145A32')
    r = _info(ws, r, "Davr:", filter_text, ncols)
    r = _info(ws, r, "Jami:", qs.filter(visit_type='ambulatory').count(), ncols)

    headers = ['№', 'F.I.Sh', "Tug'ilgan sana", 'Yosh', 'Jinsi',
               'Bayonnoma', 'JSHSHIR', 'Telefon',
               'Qabul sanasi', 'Bemor turi', 'Tashxis', 'Natija']
    widths = [5, 28, 13, 6, 7, 14, 16, 14, 13, 13, 30, 14]
    aligns = ['center','left','center','center','center',
              'center','center','center','center','center','left','left']
    r = _col_hdrs(ws, r, headers, widths, S, '145A32')

    patients = qs.filter(visit_type='ambulatory').order_by('admission_date')
    for i, p in enumerate(patients, 1):
        vals = [
            i, p.full_name,
            p.birth_date.strftime('%d.%m.%Y') if p.birth_date else '—',
            _age(p) or '—', p.get_gender_display() if p.gender else '—',
            p.medical_record_number or '—',
            p.JSHSHIR or '—', p.phone or '—',
            p.admission_date.strftime('%d.%m.%Y') if p.admission_date else '—',
            p.get_patient_category_display(),
            p.admission_diagnosis or '—',
            p.get_outcome_display() if p.outcome else '—',
        ]
        r = _row(ws, r, vals, S, aligns, even=(i%2==0))

    _total(ws, r, [f'JAMI: {patients.count()} ta bemor'] + ['']*(ncols-1), S, '145A32')


def _sheet_categories(wb, qs, S, filter_text):
    """3-Sheet: Kategoriya bo'yicha (bitta sheet, jami qatorlar bilan)"""
    from django.db.models import Count, Sum
    ws = wb.create_sheet("3. Kategoriyalar")
    ncols = 10
    r = _hdr(ws, 1, "BEMORLAR KATEGORIYA BO'YICHA", ncols, '7D6608')
    r = _info(ws, r, "Davr:", filter_text, ncols)

    headers = [
        "Bo'lim", 'Jami bemor',
        "TY soni", "TY %",
        "Pullik soni", "Pullik %",
        "Norezident soni", "Norezident %",
        "Statsionar", "Ambulator"
    ]
    widths = [22, 11, 10, 8, 11, 8, 13, 8, 11, 10]
    r = _col_hdrs(ws, r, headers, widths, S, '7D6608')

    from apps.patients.models import Department
    from django.db.models import Q
    departments = Department.objects.filter(is_active=True).order_by('name')

    grand = {'total':0,'ry':0,'pd':0,'nr':0,'inp':0,'amb':0}
    for i, dept in enumerate(departments, 1):
        dqs = qs.filter(department=dept)
        total = dqs.count()
        if total == 0: continue
        ry  = dqs.filter(patient_category='railway').count()
        pd  = dqs.filter(patient_category='paid').count()
        nr  = dqs.filter(patient_category='non_resident').count()
        inp = dqs.filter(visit_type='inpatient').count()
        amb = dqs.filter(visit_type='ambulatory').count()
        pct = lambda n: round(n/total*100, 1) if total else 0
        vals = [str(dept), total, ry, pct(ry), pd, pct(pd), nr, pct(nr), inp, amb]
        al = ['left','center','center','center','center','center','center','center','center','center']
        r = _row(ws, r, vals, S, al, nums={2,3,5,7,9,10}, even=(i%2==0))
        for k, v in zip(['total','ry','pd','nr','inp','amb'],[total,ry,pd,nr,inp,amb]):
            grand[k] += v

    pct = lambda n: round(n/grand['total']*100,1) if grand['total'] else 0
    _total(ws, r, ['JAMI:', grand['total'],
        grand['ry'], pct(grand['ry']),
        grand['pd'], pct(grand['pd']),
        grand['nr'], pct(grand['nr']),
        grand['inp'], grand['amb']], S, '7D6608')


def _sheet_services(wb, qs, S, filter_text):
    """4-Sheet: Xizmatlar va daromad"""
    from apps.services.models import PatientService
    from django.db.models import Sum, Count, Q
    ws = wb.create_sheet("4. Xizmatlar daromadi")
    ncols = 8
    r = _hdr(ws, 1, "XIZMATLAR VA DAROMAD HISOBOTI", ncols)
    r = _info(ws, r, "Davr:", filter_text, ncols)

    headers = ['№', 'Xizmat kategoriyasi', 'Jami soni',
               'Bemorlar', "TY summa", "Pullik summa", "Nores. summa", "Umumiy summa"]
    widths = [5, 28, 11, 10, 16, 16, 16, 18]
    r = _col_hdrs(ws, r, headers, widths, S)

    from apps.services.models import ServiceCategory
    from django.db.models import ExpressionWrapper, DecimalField, F
    cats = ServiceCategory.objects.filter(is_active=True).order_by('name')
    grand_sum = 0
    for i, cat in enumerate(cats, 1):
        sqs = PatientService.objects.filter(
            service__category=cat, patient_card__in=qs
        )
        total_cnt = sqs.count()
        if total_cnt == 0: continue
        # Jami xizmat miqdori (quantity larni yig'ish)
        total_qty = sqs.aggregate(t=Sum('quantity'))['t'] or 0
        patients_cnt = sqs.values('patient_card').distinct().count()
        # Summa = price * quantity (har bir xizmat miqdori bilan)
        price_x_qty = ExpressionWrapper(
            F('price') * F('quantity'), output_field=DecimalField()
        )
        ry_sum = float(sqs.filter(patient_card__patient_category='railway').annotate(
            pxq=price_x_qty).aggregate(t=Sum('pxq'))['t'] or 0)
        pd_sum = float(sqs.filter(patient_card__patient_category='paid').annotate(
            pxq=price_x_qty).aggregate(t=Sum('pxq'))['t'] or 0)
        nr_sum = float(sqs.filter(patient_card__patient_category='non_resident').annotate(
            pxq=price_x_qty).aggregate(t=Sum('pxq'))['t'] or 0)
        tot_sum = ry_sum + pd_sum + nr_sum
        grand_sum += tot_sum
        vals = [i, f"{cat.icon or ''} {cat.name}", total_qty, patients_cnt,
                ry_sum, pd_sum, nr_sum, tot_sum]
        al = ['center','left','center','center','right','right','right','right']
        r = _row(ws, r, vals, S, al, nums={3,4,5,6,7,8}, even=(i%2==0))

    _total(ws, r, ['', 'UMUMIY DAROMAD:', '', '', '', '', '', grand_sum], S)


def _sheet_medicines(wb, qs, S, filter_text):
    """5-Sheet: Dori-darmonlar"""
    from apps.services.models import PatientMedicine
    from django.db.models import Sum, Count
    ws = wb.create_sheet("5. Dori-darmonlar")
    ncols = 7
    r = _hdr(ws, 1, "DORI-DARMONLAR HISOBOTI", ncols, '7D6608')
    r = _info(ws, r, "Davr:", filter_text, ncols)

    headers = ['№', 'Dori nomi', 'Birlik', 'Jami miqdor', 'Bemorlar', 'Amallar soni', "Jami summa"]
    widths = [5, 35, 10, 12, 10, 12, 16]
    r = _col_hdrs(ws, r, headers, widths, S, '7D6608')

    from django.db.models import ExpressionWrapper, DecimalField, F
    price_x_qty_med = ExpressionWrapper(
        F('price') * F('quantity'), output_field=DecimalField()
    )
    meds = (
        PatientMedicine.objects.filter(patient_card__in=qs)
        .annotate(pxq=price_x_qty_med)
        .values('medicine__name', 'medicine__unit')
        .annotate(
            total_qty=Sum('quantity'),
            patients=Count('patient_card', distinct=True),
            cnt=Count('id'),
            total_sum=Sum('pxq'),        # price * quantity
        ).order_by('-total_sum')
    )

    grand = 0
    for i, m in enumerate(meds, 1):
        ts = float(m['total_sum'] or 0)
        grand += ts
        vals = [i, m['medicine__name'], m['medicine__unit'],
                float(m['total_qty'] or 0), m['patients'], m['cnt'], ts]
        al = ['center','left','center','center','center','center','right']
        r = _row(ws, r, vals, S, al, nums={4,5,6,7}, even=(i%2==0))

    _total(ws, r, ['', 'JAMI:', '', '', '', '', grand], S, '7D6608')


def _sheet_operations(wb, qs, S, filter_text):
    """6-Sheet: Operatsiyalar"""
    from apps.patients.models import SurgicalOperation
    from django.db.models import Count, Q
    ws = wb.create_sheet("6. Operatsiyalar")
    ncols = 10
    r = _hdr(ws, 1, "OPERATSIYALAR HISOBOTI", ncols, '512E5F')
    r = _info(ws, r, "Davr:", filter_text, ncols)

    from apps.services.models import PatientService
    from django.db.models import Sum as OSum, ExpressionWrapper, DecimalField, F

    headers = ['№', 'Operatsiya nomi', 'Jami soni',
               "TY soni", "Pullik soni", "Nores. soni",
               'Narkoz bilan', 'Mahalliy narkoz', 'Narkozsiz', 'Asorat']
    widths = [5, 32, 10, 9, 10, 10, 12, 13, 10, 9]
    r = _col_hdrs(ws, r, headers, widths, S, '512E5F')

    ops = (
        SurgicalOperation.objects.filter(
            patient_card__in=qs, operation_type__isnull=False
        )
        .values('operation_type__name')
        .annotate(
            total=Count('id'),
            ry=Count('id', filter=Q(patient_card__patient_category='railway')),
            pd=Count('id', filter=Q(patient_card__patient_category='paid')),
            nr=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            anes_yes=Count('id', filter=Q(anesthesia='yes')),
            anes_loc=Count('id', filter=Q(anesthesia='local')),
            anes_no=Count('id', filter=Q(anesthesia='no')),
            comp=Count('id', filter=Q(complication__isnull=False) & ~Q(complication='')),
        ).order_by('-total')
    )

    grand_total = 0
    for i, op in enumerate(ops, 1):
        vals = [i, op['operation_type__name'], op['total'],
                op['ry'], op['pd'], op['nr'],
                op['anes_yes'], op['anes_loc'], op['anes_no'], op['comp']]
        al = ['center','left'] + ['center']*8
        r = _row(ws, r, vals, S, al, nums={3,4,5,6,7,8,9,10}, even=(i%2==0))
        grand_total += op['total']

    _total(ws, r, ['', 'JAMI:', grand_total] + ['']*(ncols-3), S, '512E5F')


def _sheet_age_groups(wb, qs, S, filter_text):
    """7-Sheet: Yosh bo'yicha taqsimot"""
    ws = wb.create_sheet("7. Yosh guruhlari")
    ncols = 9
    r = _hdr(ws, 1, "YOSH BO'YICHA TAQSIMOT", ncols, '1A5276')
    r = _info(ws, r, "Davr:", filter_text, ncols)

    headers = ['Yosh guruhi', 'Jami', 'Erkak', 'Ayol',
               "TY", "Pullik", "Norezident", "Statsionar", "Ambulator"]
    widths = [20, 9, 9, 9, 9, 9, 11, 11, 10]
    r = _col_hdrs(ws, r, headers, widths, S, '1A5276')

    today = date.today()
    age_groups = [
        ("0–15 yosh (bolalar)",  None, 16),
        ("16–17 yosh",           16, 18),
        ("18–29 yosh",           18, 30),
        ("30–39 yosh",           30, 40),
        ("40–49 yosh",           40, 50),
        ("50–59 yosh",           50, 60),
        ("60–69 yosh",           60, 70),
        ("70+ yosh",             70, None),
    ]

    grand = {'total':0,'m':0,'f':0,'ry':0,'pd':0,'nr':0,'inp':0,'amb':0}

    for i, (label, age_from, age_to) in enumerate(age_groups, 1):
        aqs = qs.filter(birth_date__isnull=False)
        if age_from is not None:
            cutoff_from = today - relativedelta(years=age_from)
            aqs = aqs.filter(birth_date__lte=cutoff_from)
        if age_to is not None:
            cutoff_to = today - relativedelta(years=age_to)
            aqs = aqs.filter(birth_date__gt=cutoff_to)

        total = aqs.count()
        m   = aqs.filter(gender='M').count()
        f   = aqs.filter(gender='F').count()
        ry  = aqs.filter(patient_category='railway').count()
        pd  = aqs.filter(patient_category='paid').count()
        nr  = aqs.filter(patient_category='non_resident').count()
        inp = aqs.filter(visit_type='inpatient').count()
        amb = aqs.filter(visit_type='ambulatory').count()

        al = ['left'] + ['center']*8
        r = _row(ws, r, [label,total,m,f,ry,pd,nr,inp,amb], S, al,
                 nums={2,3,4,5,6,7,8,9}, even=(i%2==0))

        for k,v in zip(['total','m','f','ry','pd','nr','inp','amb'],
                       [total,m,f,ry,pd,nr,inp,amb]):
            grand[k] += v

    _total(ws, r, ['JAMI:', grand['total'], grand['m'], grand['f'],
                   grand['ry'], grand['pd'], grand['nr'],
                   grand['inp'], grand['amb']], S, '1A5276')


def _sheet_organizations(wb, qs, S, filter_text):
    """8-Sheet: Tashkilotlar bo'yicha (TY korxonalari)"""
    from django.db.models import Count, Sum
    from apps.services.models import PatientService
    ws = wb.create_sheet("8. TY tashkilotlar")
    ncols = 9
    r = _hdr(ws, 1, "TEMIR YO'L TASHKILOTLARI BO'YICHA", ncols)
    r = _info(ws, r, "Davr:", filter_text, ncols)

    headers = ['№', 'Tashkilot nomi', 'Filial',
               'Bemorlar soni', 'Statsionar', 'Ambulator',
               "Xizmatlar summasi", "Dorilar summasi", "Jami summa"]
    widths = [5, 30, 24, 12, 11, 10, 16, 15, 16]
    r = _col_hdrs(ws, r, headers, widths, S)

    from apps.patients.models import Organization
    orgs = (
        qs.filter(workplace_org__isnull=False)
        .values('workplace_org__enterprise_name', 'workplace_org__branch_name', 'workplace_org_id')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )

    from apps.services.models import PatientMedicine
    grand_pts = grand_svc = grand_med = 0

    for i, org in enumerate(orgs, 1):
        org_id = org['workplace_org_id']
        oqs = qs.filter(workplace_org_id=org_id)
        cnt = org['cnt']
        inp = oqs.filter(visit_type='inpatient').count()
        amb = oqs.filter(visit_type='ambulatory').count()
        from django.db.models import ExpressionWrapper, DecimalField, F
        px_svc = ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())
        px_med = ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())
        svc = float(PatientService.objects.filter(patient_card__in=oqs).annotate(pxq=px_svc).aggregate(t=Sum('pxq'))['t'] or 0)
        med = float(PatientMedicine.objects.filter(patient_card__in=oqs).annotate(pxq=px_med).aggregate(t=Sum('pxq'))['t'] or 0)
        tot = svc + med

        vals = [i, org['workplace_org__enterprise_name'] or '—',
                org['workplace_org__branch_name'] or '—',
                cnt, inp, amb, svc, med, tot]
        al = ['center','left','left','center','center','center','right','right','right']
        r = _row(ws, r, vals, S, al, nums={4,5,6,7,8,9}, even=(i%2==0))
        grand_pts += cnt; grand_svc += svc; grand_med += med

    _total(ws, r, ['', 'JAMI:', '', grand_pts, '', '', grand_svc, grand_med, grand_svc+grand_med], S)


# ==================== ASOSIY FUNKSIYA ====================

@login_required
def export_full_report(request):
    """Barcha hisobotlar — bitta Excel faylda"""
    qs = get_filtered_queryset(request)
    qs = qs.select_related(
        'department', 'attending_doctor', 'workplace_org',
        'region', 'district', 'parent_workplace_org'
    )

    filter_text = ''
    MONTHS = {str(i): m for i, m in enumerate(
        ['Yanvar','Fevral','Mart','Aprel','May','Iyun',
         'Iyul','Avgust','Sentabr','Oktabr','Noyabr','Dekabr'], 1)}
    CATS = {'railway':"Temir yo'lchi", 'paid':'Pullik', 'non_resident':'Norezident'}
    parts = []
    if request.GET.get('year'):  parts.append(f"Yil: {request.GET['year']}")
    if request.GET.get('month'): parts.append(f"Oy: {MONTHS.get(request.GET['month'],'')}")
    if request.GET.get('date_from'): parts.append(f"Dan: {request.GET['date_from']}")
    if request.GET.get('date_to'):   parts.append(f"Gacha: {request.GET['date_to']}")
    if request.GET.get('patient_category'):
        parts.append(f"Kategoriya: {CATS.get(request.GET['patient_category'],'')}")
    filter_text = ' | '.join(parts) if parts else 'Barcha davr'

    S = _styles()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # bo'sh sheet'ni o'chirish

    # Mundarija sheet
    ws0 = wb.create_sheet("Mundarija")
    _hdr(ws0, 1, "SKB SHIFOXONASI — TO'LIQ HISOBOTLAR", 3)
    _info(ws0, 2, "Davr:", filter_text, 3)
    _info(ws0, 3, "Jami bemorlar:", qs.count(), 3)
    ws0.cell(row=5, column=1, value="Sheet").font = Font(bold=True)
    ws0.cell(row=5, column=2, value="Hisobot nomi").font = Font(bold=True)
    ws0.cell(row=5, column=3, value="Bemorlar soni").font = Font(bold=True)
    sheets_info = [
        ("1. Statsionar", "Statsionar bemorlar ro'yxati", qs.filter(visit_type='inpatient').count()),
        ("2. Ambulator", "Ambulator bemorlar ro'yxati", qs.filter(visit_type='ambulatory').count()),
        ("3. Kategoriyalar", "Bemorlar kategoriya bo'yicha", qs.count()),
        ("4. Xizmatlar", "Xizmatlar va daromad hisoboti", '—'),
        ("5. Dorilar", "Dori-darmonlar hisoboti", '—'),
        ("6. Operatsiyalar", "Operatsiyalar hisoboti", '—'),
        ("7. Yosh guruhlari", "Yosh bo'yicha taqsimot", qs.count()),
        ("8. TY tashkilotlar", "Temir yo'l tashkilotlari", qs.filter(workplace_org__isnull=False).count()),
    ]
    for i, (code, name, cnt) in enumerate(sheets_info, 6):
        ws0.cell(row=i, column=1, value=code)
        ws0.cell(row=i, column=2, value=name)
        ws0.cell(row=i, column=3, value=cnt)
    ws0.column_dimensions['A'].width = 18
    ws0.column_dimensions['B'].width = 35
    ws0.column_dimensions['C'].width = 15

    # Sheetlarni yaratish
    _sheet_statsionar(wb, qs, S, filter_text)
    _sheet_ambulatory(wb, qs, S, filter_text)
    _sheet_categories(wb, qs, S, filter_text)
    _sheet_services(wb, qs, S, filter_text)
    _sheet_medicines(wb, qs, S, filter_text)
    _sheet_operations(wb, qs, S, filter_text)
    _sheet_age_groups(wb, qs, S, filter_text)
    _sheet_organizations(wb, qs, S, filter_text)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="skb_hisobot.xlsx"'
    wb.save(response)
    return response