# apps/statistic/monthly_report.py
# Oylik rasmiy hisobotlar — asl Excel shakliga mos

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from datetime import date
from dateutil.relativedelta import relativedelta

from apps.patients.models import PatientCard, Department, SurgicalOperation
from apps.services.models import PatientService, PatientMedicine
from .exports import get_filtered_queryset

MONTHS_UZ = {
    1:'Yanvar', 2:'Fevral', 3:'Mart', 4:'Aprel',
    5:'May', 6:'Iyun', 7:'Iyul', 8:'Avgust',
    9:'Sentabr', 10:'Oktabr', 11:'Noyabr', 12:'Dekabr'
}

# ==================== STILLAR ====================
def S():
    t = Side(style='thin', color='000000')
    m = Side(style='medium', color='000000')
    return {
        'thin': Border(left=t, right=t, top=t, bottom=t),
        'med':  Border(left=m, right=m, top=m, bottom=m),
        'C': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'L': Alignment(horizontal='left',   vertical='center', wrap_text=True),
        'R': Alignment(horizontal='right',  vertical='center'),
        'BLUE':  PatternFill('solid', fgColor='BDD7EE'),
        'BLUE2': PatternFill('solid', fgColor='9DC3E6'),
        'GREEN': PatternFill('solid', fgColor='C6EFCE'),
        'YELL':  PatternFill('solid', fgColor='FFEB9C'),
        'GRAY':  PatternFill('solid', fgColor='D9D9D9'),
        'WHITE': PatternFill('solid', fgColor='FFFFFF'),
        'BOLD':  Font(bold=True, size=9),
        'NORM':  Font(size=9),
        'SM':    Font(size=8),
    }


def _safe_cell(ws, row, col, value=None):
    """MergedCell xatosini oldini oluvchi ws.cell o'rini bosuvchi"""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell
    if value is not None:
        cell.value = value
    return cell


def c(ws, row, col, value, style, font=None, fill=None, align=None, border=None, numfmt=None):
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return cell  # Merge qilingan katakka yozib bo'lmaydi
    cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if numfmt: cell.number_format = numfmt
    return cell


def merge(ws, r1, c1, r2, c2, value, style, **kw):
    # Avvalgi merge ni olib tashlash (agar bo'lsa)
    try:
        ws.unmerge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    except Exception:
        pass
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return c(ws, r1, c1, value, style, **kw)


def pct(a, b):
    """Foiz hisoblash"""
    if not b: return 0
    return round(a / b * 100, 1)


def avg_days(total_days, discharged):
    """O'rtacha yotish kuni"""
    if not discharged: return 0
    return round(total_days / discharged, 1)


def bed_turnover(discharged, beds):
    """O'rin aylanishi"""
    if not beds: return 0
    return round(discharged / beds, 1)


def bed_work(total_days, beds, days_in_month=30):
    """O'rin ishi"""
    if not beds: return 0
    return round(total_days / beds, 1)


# ==================== MA'LUMOT OLISH ====================

def get_dept_data(dept_name, qs_ty, qs_pullik, days_in_month=30):
    """Bo'lim bo'yicha barcha ko'rsatkichlar"""
    dept = Department.objects.filter(name__icontains=dept_name, is_active=True).first()
    if not dept:
        # Partial match
        depts = Department.objects.filter(is_active=True)
        for d in depts:
            if dept_name.lower() in d.name.lower():
                dept = d
                break

    def stats(qs):
        if dept:
            dqs = qs.filter(department=dept)
        else:
            dqs = qs.none()

        beds = dept.beds if dept and hasattr(dept, 'beds') else 0
        admitted  = dqs.count()
        discharged = dqs.filter(status='completed').count()
        deceased  = dqs.filter(outcome='deceased').count()
        # O'rin-kun: discharge_date - admission_date yig'indisi
        total_days = 0
        for p in dqs.filter(status='completed', discharge_date__isnull=False, admission_date__isnull=False):
            delta = (p.discharge_date - p.admission_date.date()).days if hasattr(p.admission_date, 'date') else 0
            total_days += delta

        return {
            'beds': beds,
            'admitted': admitted,
            'discharged': discharged,
            'deceased': deceased,
            'total_days': total_days,
            'avg_days': avg_days(total_days, discharged),
            'turnover': bed_turnover(discharged, beds),
            'bed_work': bed_work(total_days, beds, days_in_month),
        }

    return {'ty': stats(qs_ty), 'pullik': stats(qs_pullik)}


# ==================== HISOBOT 1: O'RINLAR FONDI ====================

def sheet_orinlar_fondi(wb, qs, year, month, S_):
    """Asl formaga mos O'rinlar fondi hisoboti"""
    ws = wb.create_sheet("O'rinlar fondi")

    month_name = MONTHS_UZ.get(month, str(month))
    days_in_month = (date(year, month % 12 + 1, 1) - date(year, month, 1)).days if month < 12 else 31

    # Bo'limlar ro'yxati
    BOLIMLAR = [
        ("Jarrohlik",            ['jarroh']),
        ("LOR (Quloq,tomoq,burun)", ['lor', 'quloq', 'otorinolaringolog']),
        ("Ko'z kasalliklari",    ["ko'z", 'oftalm']),
        ("Yurak kasalliklari",   ['yurak', 'kardio']),
        ("Ichki kasalliklar",    ['ichki', 'terapev']),
        ("Me'da-ichak",          ["me'da", 'gastro']),
        ("Asab kasalliklari",    ['asab', 'nevro']),
        ("Reanimatsiya",         ['reanim', 'intensiv']),
    ]

    qs_ty     = qs.filter(patient_category='railway')
    qs_pullik = qs.filter(patient_category='paid')

    # Ustun kengliklari
    col_w = [4, 28, 6, 8, 10, 10, 8, 8, 8, 7, 8, 8, 7, 8, 8, 8, 8]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def write_section(start_row, qs_sect, title):
        """Bitta bo'lim (TY yoki Pullik) uchun jadval"""
        r = start_row
        ws.row_dimensions[r].height = 40

        # Sarlavha
        try:
            ws.unmerge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
        except Exception:
            pass
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=17)
        cell = _safe_cell(ws, r, 1, title)
        cell.font = Font(bold=True, size=10)
        cell.alignment = S_['C']
        r += 1

        # Ustun sarlavhalari — 3 qator
        hdr_data = [
            # (row_offset, col, rowspan, colspan, text)
            (0, 1, 2, 1, "T/R"),
            (0, 2, 2, 1, "Bo'lim nomi"),
            (0, 3, 1, 2, "O'rin soni"),
            (0, 5, 1, 3, "Bemor harakati"),
            (0, 8, 1, 3, "Davolangan bemor"),
            (0, 11, 1, 3, "O'tkazilgan o'rin-kun"),
            (0, 14, 2, 1, "O'lim %"),
            (0, 15, 1, 3, "Ko'rsatkichlar"),
        ]
        sub_hdrs = [
            (1, 3, "Fakt"),
            (1, 4, "O'rtacha"),
            (1, 5, "Yotqizilgan"),
            (1, 6, "Chiqarilgan"),
            (1, 7, "Vafot"),
            (1, 8, "Reja"),
            (1, 9, "Fakt"),
            (1, 10, "%"),
            (1, 11, "Reja"),
            (1, 12, "Fakt"),
            (1, 13, "%"),
            (1, 15, "O'rin ishi"),
            (1, 16, "O'rt. yotish"),
            (1, 17, "O'rin ayl."),
        ]

        for rd, col, rs, cs, text in hdr_data:
            if rs > 1 or cs > 1:
                try:
                    ws.unmerge_cells(start_row=r+rd, start_column=col, end_row=r+rd+rs-1, end_column=col+cs-1)
                except Exception:
                    pass
                ws.merge_cells(start_row=r+rd, start_column=col, end_row=r+rd+rs-1, end_column=col+cs-1)
            from openpyxl.cell.cell import MergedCell
            raw = ws.cell(row=r+rd, column=col)
            if not isinstance(raw, MergedCell):
                raw.value = text
                raw.font = S_['BOLD']; raw.alignment = S_['C']
                raw.border = S_['thin']; raw.fill = S_['BLUE']

        for rd, col, text in sub_hdrs:
            cell = _safe_cell(ws, r+rd, col, text)
            cell.font = S_['BOLD']; cell.alignment = S_['C']
            cell.border = S_['thin']; cell.fill = S_['BLUE']

        r += 2

        # Raqamlar qatori
        for col, num in enumerate(range(1, 16), 3):
            c(ws, r, col, num, S_, font=S_['SM'], align=S_['C'], border=S_['thin'], fill=S_['GRAY'])
        r += 1

        data_rows = []
        jarrohlik_rows = []
        ichki_rows = []

        for num, (name, keys) in enumerate(BOLIMLAR, 1):
            # Bo'limni topish
            dept_obj = None
            for d in Department.objects.filter(is_active=True):
                if any(k in d.name.lower() for k in keys):
                    dept_obj = d
                    break

            dqs = qs_sect.filter(department=dept_obj) if dept_obj else qs_sect.none()

            admitted   = dqs.count()
            discharged = dqs.filter(Q(status='completed') | Q(outcome='discharged')).count()
            deceased   = dqs.filter(outcome='deceased').count()

            total_days = 0
            for p in dqs.filter(discharge_date__isnull=False):
                if p.admission_date and p.discharge_date:
                    ad = p.admission_date.date() if hasattr(p.admission_date, 'date') else p.admission_date
                    dd = p.discharge_date.date() if hasattr(p.discharge_date, 'date') else p.discharge_date
                    total_days += max(0, (dd - ad).days)

            beds = 0  # Modelda beds maydoni bo'lsa ishlatiladi

            bed_w = round(total_days / beds, 1) if beds else 0
            avg_d = round(total_days / discharged, 1) if discharged else 0
            turn  = round(discharged / beds, 1) if beds else 0
            dead_pct = round(deceased / (discharged + deceased) * 100, 2) if (discharged + deceased) else 0

            row_data = {
                'num': num, 'name': name,
                'admitted': admitted, 'discharged': discharged, 'deceased': deceased,
                'total_days': total_days, 'beds': beds,
                'bed_work': bed_w, 'avg_days': avg_d, 'turnover': turn,
                'dead_pct': dead_pct,
            }
            data_rows.append(row_data)

            # Raqam ustun
            num_val = num if name != 'Reanimatsiya' else None
            vals = [
                num_val, name, beds, beds,
                admitted, discharged, deceased,
                0, discharged, pct(discharged, discharged) if discharged else 0,
                0, total_days, 0,
                dead_pct, bed_w, avg_d, turn
            ]

            fill = S_['WHITE'] if num % 2 == 0 else S_['WHITE']
            if 'profil' in name.lower() or name == 'Reanimatsiya':
                fill = S_['YELL']

            for ci, val in enumerate(vals, 1):
                cell = c(ws, r, ci, val, S_,
                         font=S_['BOLD'] if 'profil' in name.lower() else S_['NORM'],
                         align=S_['C'] if ci not in (2,) else S_['L'],
                         border=S_['thin'], fill=fill)
                if ci in (10, 13, 14) and isinstance(val, float):
                    cell.number_format = '0.0'

            ws.row_dimensions[r].height = 16
            r += 1

            # Profil qatorlari
            if num == 3:  # Jarrohlik profili
                jarrohlik = [d for d in data_rows[-3:]]
                prof_vals = [
                    None, "Jarrohlik profili",
                    sum(d['beds'] for d in jarrohlik),
                    sum(d['beds'] for d in jarrohlik),
                    sum(d['admitted'] for d in jarrohlik),
                    sum(d['discharged'] for d in jarrohlik),
                    sum(d['deceased'] for d in jarrohlik),
                    0,
                    sum(d['discharged'] for d in jarrohlik), 0,
                    0, sum(d['total_days'] for d in jarrohlik), 0,
                    0, 0, 0, 0
                ]
                for ci, val in enumerate(prof_vals, 1):
                    c(ws, r, ci, val, S_, font=S_['BOLD'],
                      align=S_['C'] if ci != 2 else S_['L'],
                      border=S_['thin'], fill=S_['BLUE2'])
                ws.row_dimensions[r].height = 16
                r += 1

            if num == 7:  # Ichki profili
                ichki = [d for d in data_rows[3:7]]
                prof_vals = [
                    None, "Ichki kasalliklar profili",
                    sum(d['beds'] for d in ichki),
                    sum(d['beds'] for d in ichki),
                    sum(d['admitted'] for d in ichki),
                    sum(d['discharged'] for d in ichki),
                    sum(d['deceased'] for d in ichki),
                    0,
                    sum(d['discharged'] for d in ichki), 0,
                    0, sum(d['total_days'] for d in ichki), 0,
                    0, 0, 0, 0
                ]
                for ci, val in enumerate(prof_vals, 1):
                    c(ws, r, ci, val, S_, font=S_['BOLD'],
                      align=S_['C'] if ci != 2 else S_['L'],
                      border=S_['thin'], fill=S_['BLUE2'])
                ws.row_dimensions[r].height = 16
                r += 1

                # Jami
                all_d = data_rows[:7]
                jami_vals = [
                    None, "JAMI:",
                    sum(d['beds'] for d in all_d),
                    sum(d['beds'] for d in all_d),
                    sum(d['admitted'] for d in all_d),
                    sum(d['discharged'] for d in all_d),
                    sum(d['deceased'] for d in all_d),
                    0,
                    sum(d['discharged'] for d in all_d), 0,
                    0, sum(d['total_days'] for d in all_d), 0,
                    0, 0, 0, 0
                ]
                for ci, val in enumerate(jami_vals, 1):
                    c(ws, r, ci, val, S_, font=S_['BOLD'],
                      align=S_['C'] if ci != 2 else S_['L'],
                      border=S_['thin'], fill=S_['GREEN'])
                ws.row_dimensions[r].height = 18
                r += 1

        # Imzo qismi
        r += 1
        for label, name_val in [
            ("Direktor", ""), ("Bosh shifokor", ""), ("Bosh hisobchi", ""), ("Iqtisodchi", "")
        ]:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            _safe_cell(ws, r, 2, label).font = S_['NORM']
            ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=13)
            _safe_cell(ws, r, 10, name_val).font = S_['NORM']
            ws.row_dimensions[r].height = 14
            r += 1

        return r + 2

    # Uch bo'lim: TY, Pullik, Jami
    r = 1
    title_ty = (f'"O\'TY" AJ "Temir yo\'l ijtimoiy xizmatlar" MCHJ Markaziy klinik kasalxonasi filialida '
                f'TEMIRYO\'LCHI O\'RINLARI FONDI ISHLATILISHI bo\'yicha '
                f'{year} yil {month_name} oyi holatiga ma\'lumot')
    r = write_section(r, qs_ty, f'I. {title_ty}')

    title_pullik = (f'"O\'TY" AJ "Temir yo\'l ijtimoiy xizmatlar" MCHJ Markaziy klinik kasalxonasi filialida '
                   f'PULLIK O\'RINLAR FONDI ISHLATILISHI bo\'yicha '
                   f'{year} yil {month_name} oyi holatiga ma\'lumot')
    r = write_section(r, qs_pullik, f'II. {title_pullik}')

    return ws


# ==================== HISOBOT 2: OPERATSIYALAR ====================

def sheet_operatsiyalar(wb, qs, year, month, S_):
    """Operatsiyalar hisoboti — asl formaga mos"""
    ws = wb.create_sheet("Operatsiyalar")
    month_name = MONTHS_UZ.get(month, str(month))

    col_w = [4, 20, 4, 10, 12, 8, 8, 8, 10, 8, 8, 8, 10, 12]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OP_BOLIMLAR = [
        ("Jarrohlik",    ['jarroh']),
        ("LOR",          ['lor', 'quloq']),
        ("Ko'z",         ["ko'z", 'oftalm']),
    ]

    HEADERS = [
        "Chiqarilgan", "Operatsiya qilingan",
        "Shoshilinch", "18 yoshgacha", "Ektremist",
        "Operatsiyalar soni", "Shoshilinch", "18 yoshgacha", "Ekstr.",
        "Vafot (op. keyin)", "Jarrohlik faolligi %"
    ]

    def write_op_section(start_row, qs_sect, category_label):
        r = start_row

        # Sarlavha
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        cell = _safe_cell(ws, r, 1, category_label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = S_['C']
        cell.fill = S_['BLUE']
        ws.row_dimensions[r].height = 22
        r += 1

        # Ustun sarlavhalari
        hdrs = ['', "Bo'lim", '',
                'Chiqarilgan', "Op. qilingan bemor", "Shoshilinch", "18 yoshgacha", "Ekstr.",
                "Operatsiyalar soni", "Shoshilinch", "18 yosh.", "Ekstr.",
                "Vafot\n(op. keyin)", "Jarroh.\nfaolligi %"]
        for ci, h in enumerate(hdrs, 1):
            cell = _safe_cell(ws, r, ci, h)
            cell.font = S_['BOLD']; cell.alignment = S_['C']
            cell.border = S_['thin']; cell.fill = S_['BLUE']
        ws.row_dimensions[r].height = 30
        r += 1

        dept_rows = []
        for num, (name, keys) in enumerate(OP_BOLIMLAR, 1):
            dept_obj = None
            for d in Department.objects.filter(is_active=True):
                if any(k in d.name.lower() for k in keys):
                    dept_obj = d
                    break

            dqs = qs_sect.filter(department=dept_obj) if dept_obj else qs_sect.none()

            discharged = dqs.filter(Q(status='completed') | Q(outcome='discharged')).count()
            ops = SurgicalOperation.objects.filter(patient_card__in=dqs)
            op_patients = ops.values('patient_card').distinct().count()
            op_count    = ops.count()

            # Yoshlar
            today = date.today()
            cutoff_18 = today - relativedelta(years=18)
            op_under18 = ops.filter(patient_card__birth_date__gt=cutoff_18).count()
            discharged_under18 = dqs.filter(birth_date__gt=cutoff_18).count()

            # Shoshilinch
            shoshilinch = dqs.filter(is_emergency=True).count()
            op_shosh = ops.filter(patient_card__is_emergency=True).count()

            # Vafot (op. keyin)
            deceased_after_op = ops.filter(patient_card__outcome='deceased').values('patient_card').distinct().count()

            # Jarrohlik faolligi
            jar_faol = pct(op_patients, discharged)

            dept_rows.append(discharged)
            vals = ['', name, '',
                    discharged, op_patients, shoshilinch, discharged_under18, 0,
                    op_count, op_shosh, op_under18, 0,
                    deceased_after_op, jar_faol]

            for ci, val in enumerate(vals, 1):
                cell = _safe_cell(ws, r, ci, val)
                cell.font = S_['NORM']; cell.border = S_['thin']
                cell.alignment = S_['C'] if ci != 2 else S_['L']
                if ci == 14 and isinstance(val, float):
                    cell.number_format = '0.0'
            ws.row_dimensions[r].height = 16
            r += 1

        # Jami
        jami_discharged = sum(dept_rows)
        _safe_cell(ws, r, 2, "JAMI:").font = S_['BOLD']
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        jami_ops = SurgicalOperation.objects.filter(patient_card__in=qs_sect)
        jami_op_pts = jami_ops.values('patient_card').distinct().count()
        jami_op_cnt = jami_ops.count()
        jami_vals = ['', 'JAMI:', '',
                     jami_discharged, jami_op_pts, 0, 0, 0,
                     jami_op_cnt, 0, 0, 0, 0,
                     pct(jami_op_pts, jami_discharged)]
        for ci, val in enumerate(jami_vals, 1):
            cell = _safe_cell(ws, r, ci, val)
            cell.font = S_['BOLD']; cell.border = S_['thin']
            cell.fill = S_['GREEN']; cell.alignment = S_['C']
        ws.row_dimensions[r].height = 18
        return r + 2

    r = 1
    qs_ty     = qs.filter(patient_category='railway')
    qs_pullik = qs.filter(patient_category='paid')

    r = write_op_section(r, qs_ty,
        f"TEMIRYO'LCHILAR — {year} yil {month_name} oyi")
    r = write_op_section(r, qs_pullik,
        f"PULLIK BEMORLAR — {year} yil {month_name} oyi")

    # Jami (TY + Pullik)
    r = write_op_section(r, qs,
        f"JAMI (TY + Pullik) — {year} yil {month_name} oyi")

    return ws


# ==================== HISOBOT 3: XIZMATLAR DAROMADI ====================

def sheet_xizmatlar(wb, qs, year, month, S_):
    """Xizmatlar va dorilar daromadi"""
    ws = wb.create_sheet("Xizmatlar daromadi")
    month_name = MONTHS_UZ.get(month, str(month))

    col_w = [5, 30, 15, 15, 15, 18]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:F1')
    cell = ws.cell(row=1, column=1,
        value=f"XIZMATLAR VA DORILAR DAROMADI — {year} yil {month_name} oyi")
    cell.font = Font(bold=True, size=11)
    cell.alignment = S_['C']; cell.fill = S_['BLUE']
    ws.row_dimensions[1].height = 24

    hdrs = ['№', 'Kategoriya', "TY (so'm)", "Pullik (so'm)", "Norezident (so'm)", "JAMI (so'm)"]
    for ci, h in enumerate(hdrs, 1):
        cell = _safe_cell(ws, 2, ci, h)
        cell.font = S_['BOLD']; cell.alignment = S_['C']
        cell.border = S_['thin']; cell.fill = S_['BLUE']
    ws.row_dimensions[2].height = 22

    from apps.services.models import ServiceCategory
    cats = ServiceCategory.objects.filter(is_active=True).order_by('name')

    r = 3
    grand = [0, 0, 0, 0]
    for num, cat in enumerate(cats, 1):
        sqs = PatientService.objects.filter(service__category=cat, patient_card__in=qs)
        ry = float(sqs.filter(patient_card__patient_category='railway').aggregate(t=Sum('price'))['t'] or 0)
        pd = float(sqs.filter(patient_card__patient_category='paid').aggregate(t=Sum('price'))['t'] or 0)
        nr = float(sqs.filter(patient_card__patient_category='non_resident').aggregate(t=Sum('price'))['t'] or 0)
        tot = ry + pd + nr
        if tot == 0: continue
        vals = [num, f"{cat.icon or ''} {cat.name}", ry, pd, nr, tot]
        for ci, val in enumerate(vals, 1):
            cell = _safe_cell(ws, r, ci, val)
            cell.font = S_['NORM']; cell.border = S_['thin']
            cell.alignment = S_['C'] if ci == 1 else (S_['L'] if ci == 2 else S_['R'])
            if ci > 2: cell.number_format = '#,##0'
            if r % 2 == 0: cell.fill = S_['BLUE']
        ws.row_dimensions[r].height = 16
        for i, v in enumerate([ry, pd, nr, tot]):
            grand[i] += v
        r += 1

    # Dorilar
    mqs = PatientMedicine.objects.filter(patient_card__in=qs)
    m_ry = float(mqs.filter(patient_card__patient_category='railway').aggregate(t=Sum('price'))['t'] or 0)
    m_pd = float(mqs.filter(patient_card__patient_category='paid').aggregate(t=Sum('price'))['t'] or 0)
    m_nr = float(mqs.filter(patient_card__patient_category='non_resident').aggregate(t=Sum('price'))['t'] or 0)
    m_tot = m_ry + m_pd + m_nr
    if m_tot:
        vals = [r-2, '💊 Dori-darmonlar', m_ry, m_pd, m_nr, m_tot]
        for ci, val in enumerate(vals, 1):
            cell = _safe_cell(ws, r, ci, val)
            cell.font = S_['BOLD']; cell.border = S_['thin']
            cell.fill = S_['YELL']
            cell.alignment = S_['C'] if ci == 1 else (S_['L'] if ci == 2 else S_['R'])
            if ci > 2: cell.number_format = '#,##0'
        ws.row_dimensions[r].height = 16
        for i, v in enumerate([m_ry, m_pd, m_nr, m_tot]):
            grand[i] += v
        r += 1

    # Jami
    jami_vals = ['', 'UMUMIY JAMI:'] + grand
    for ci, val in enumerate(jami_vals, 1):
        cell = _safe_cell(ws, r, ci, val)
        cell.font = S_['BOLD']; cell.border = S_['thin']; cell.fill = S_['GREEN']
        cell.alignment = S_['C'] if ci == 1 else (S_['L'] if ci == 2 else S_['R'])
        if ci > 2: cell.number_format = '#,##0'
    ws.row_dimensions[r].height = 22


# ==================== ASOSIY FUNKSIYA ====================

@login_required
def export_monthly_report(request):
    """Oylik rasmiy hisobotlar"""
    year_raw  = request.GET.get('year', '').strip()
    month_raw = request.GET.get('month', '').strip()
    year  = int(year_raw)  if year_raw  else date.today().year
    month = int(month_raw) if month_raw else date.today().month

    qs = get_filtered_queryset(request)
    if not request.GET.get('year'):
        qs = qs.filter(admission_date__year=year)
    if not request.GET.get('month'):
        qs = qs.filter(admission_date__month=month)

    qs = qs.select_related(
        'department', 'attending_doctor', 'workplace_org',
        'region', 'district', 'discharge_conclusion'
    )

    S_ = S()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    month_name = MONTHS_UZ.get(month, str(month))

    sheet_orinlar_fondi(wb, qs, year, month, S_)
    sheet_operatsiyalar(wb, qs, year, month, S_)
    sheet_xizmatlar(wb, qs, year, month, S_)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="oylik_hisobot_{year}_{month:02d}.xlsx"'
    )
    wb.save(response)
    return response