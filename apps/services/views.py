# apps/services/views.py

import io
import json
from datetime import date, timedelta

import openpyxl
from django.views.decorators.http import require_POST
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, ExpressionWrapper, DecimalField, F
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.utils import timezone
from django.utils.translation import gettext as _

from apps.users.decorators import role_required
from apps.users.models import CustomUser
from apps.patients.models import PatientCard
from .models import ServiceCategory, Service, PatientService
from .forms import PatientServiceForm, ServiceResultForm


# ==================== AJAX ====================

@login_required
def service_search(request):
    """AJAX — xizmat qidirish"""
    q = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    category_type = request.GET.get('category_type', '')
    patient_id = request.GET.get('patient_id', '')

    qs = Service.objects.filter(is_active=True).select_related('category').order_by('name')
    if category_type:
        qs = qs.filter(category__category_type__in=category_type.split(','))
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(name_ru__icontains=q) | Q(code__icontains=q))
    if category_id and category_id != '__my_packages__':
        qs = qs.filter(category_id=category_id)
    # Bo'sh qidiruv — faqat 20 ta ko'rsatish
    # Kategoriya tanlangan bo'lsa — hammasi
    limit = 50 if (category_id or category_type) else 20

    # Bemor kategoriyasiga qarab narx hisoblash
    patient_category = 'normal'
    if patient_id:
        try:
            patient = PatientCard.objects.get(pk=patient_id)
            patient_category = patient.patient_category or 'normal'
        except PatientCard.DoesNotExist:
            pass

    data = []
    for s in qs[:limit]:
        price = s.price_for_patient(patient_category)
        data.append({
            'id': s.id,
            'name': str(s),
            'name_ru': s.name_ru or '',
            'category': s.category.name,
            'category_id': s.category_id,
            'price': float(price),
            'price_normal': float(s.price_normal),
            'price_railway': float(s.price_railway),
        })

    return JsonResponse(data, safe=False)


@login_required
def service_doctors(request, pk):
    """AJAX — ushbu xizmatga (masalan, konsultatsiya turiga) biriktirilgan faol shifokorlar."""
    service = get_object_or_404(Service, pk=pk)
    doctors = service.assigned_doctors.filter(is_active=True).select_related('department').order_by('department__name', 'first_name')
    data = [{
        'id':         d.pk,
        'full_name':  d.full_name,
        'department': d.department.name if d.department_id else '',
    } for d in doctors]
    return JsonResponse(data, safe=False)


# ==================== BEMOR XIZMATLARI ====================

@login_required
def patient_check(request, patient_pk):
    """80mm check chop etish"""
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    services = PatientService.objects.filter(
        patient_card=patient,
    ).exclude(status='cancelled').select_related('service').order_by('ordered_at')

    total = sum(s.total_price for s in services)

    return render(request, 'services/patient_check.html', {
        'patient':  patient,
        'services': services,
        'total':    total,
        'now':      timezone.now(),
    })


@login_required
def patient_services(request, patient_pk):
    """Bemorning barcha xizmatlari"""
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    services = PatientService.objects.filter(
        patient_card=patient
    ).select_related(
        'service__category', 'ordered_by', 'performed_by'
    ).order_by('-ordered_at')

    categories = ServiceCategory.objects.filter(is_active=True)

    # Moliyaviy umumlama
    from django.db.models import ExpressionWrapper, DecimalField, F as F_
    _pxq0 = ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())
    totals = services.annotate(_pxq=_pxq0).aggregate(
        total_sum=Sum('_pxq'),
        count=Sum('quantity'),
    )
    total_price = (totals['total_sum'] or 0)

    # Kategoriya bo'yicha umumlama
    cat_stats = services.values(
        'service__category__name',
        'service__category__icon',
    ).annotate(
        count=Sum('quantity'),
        total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
    ).order_by('-total')

    # Barcha aktiv shifokorlar (bo'lim bo'yicha guruhlab)
    doctors = CustomUser.objects.filter(
        role__in=('doctor', 'old'), is_active=True
    ).select_related('department').order_by('department__name', '-is_head', 'first_name')

    from .models import PatientMedicine
    medicines = PatientMedicine.objects.filter(
        patient_card=patient
    ).select_related('medicine', 'ordered_by').order_by('-ordered_at')
    medicines_total = sum(m.total_price for m in medicines)

    return render(request, 'services/patient_services.html', {
        'patient': patient,
        'services': services,
        'categories': categories,
        'total_price': total_price,
        'total_count': totals['count'] or 0,
        'cat_stats': cat_stats,
        'doctors': doctors,
        'medicines': medicines,
        'medicines_total': medicines_total,
    })


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception', 'head_nurse', 'old')
def add_service(request, patient_pk):
    """Bemorga xizmat qo'shish — AJAX"""
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            service_id = data.get('service_id')
            quantity = int(data.get('quantity', 1))
            ordered_by_id = data.get('ordered_by_id')
            notes = data.get('notes', '')

            service = Service.objects.get(pk=service_id, is_active=True)
            price = service.price_for_patient(patient.patient_category or 'railway')

            ps = PatientService.objects.create(
                patient_card=patient,
                service=service,
                quantity=quantity,
                price=price,
                patient_category_at_order=patient.patient_category or 'railway',
                ordered_by_id=ordered_by_id if ordered_by_id else None,
                notes=notes,
            )

            # Ambulator bemor uchun — xizmatni buyurtma qilgan shifokor avtomatik
            # davolovchi shifokor sifatida biriktiriladi, shunda bemor uning
            # kabinetida (Mening kabinetim) ko'rinadi.
            if ordered_by_id and patient.visit_type == 'ambulatory' and not patient.attending_doctor_id:
                ordering_doctor = CustomUser.objects.filter(pk=ordered_by_id, is_active=True).first()
                if ordering_doctor:
                    patient.attending_doctor = ordering_doctor
                    patient.attending_doctor_confirmed = True
                    if not patient.department_head_id and ordering_doctor.department_id:
                        patient.department_head = CustomUser.objects.filter(
                            department_id=ordering_doctor.department_id, role__in=('doctor', 'old'), is_head=True, is_active=True
                        ).first()
                    patient.save(update_fields=['attending_doctor', 'attending_doctor_confirmed', 'department_head'])

                    from apps.patients.models import DoctorNotification
                    DoctorNotification.objects.create(
                        recipient=ordering_doctor,
                        patient_card=patient,
                        message=f"Sizga yangi ambulator bemor biriktirildi: {patient.full_name} ({patient.medical_record_number})"
                    )

            # MRT xizmati bo'lsa navbat yaratish/olish
            queue_info = {}
            try:
                from apps.queue_app.views import create_queue_ticket
                ticket = create_queue_ticket(ps)
                if ticket:
                    queue_info = {
                        'queue_ticket': ticket.pk,
                        'queue_number': ticket.ticket_number,
                        'queue_is_new': ticket.service_id == ps.id,
                        'patient_name': ps.patient_card.full_name,
                    }
            except Exception as e:
                import logging
                logging.getLogger(__name__).error(f"Queue ticket error: {e}", exc_info=True)

            return JsonResponse({
                'success': True,
                'id': ps.id,
                'service_name': service.name,
                'category': service.category.name,
                'quantity': quantity,
                'price': float(price),
                'total': float(ps.total_price),
                'status': ps.get_status_display(),
                'ordered_at': ps.ordered_at.strftime('%d.%m.%Y %H:%M'),
                **queue_info,
            })

        except Service.DoesNotExist:
            return JsonResponse({'success': False, 'error': _('Xizmat topilmadi')})
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"add_service error: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': _('Xizmat qo\'shishda xato yuz berdi')})

    return JsonResponse({'success': False, 'error': _('POST required')})


@login_required
@role_required('admin', 'doctor', 'statistician', 'old')
def update_service(request, pk):
    """Xizmat holatini yangilash — AJAX"""
    ps = get_object_or_404(PatientService, pk=pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            status = data.get('status')
            result = data.get('result', '')
            is_paid = data.get('is_paid', False)
            performed_by_id = data.get('performed_by_id')
            ordered_by_id = data.get('ordered_by_id')

            if status in dict(PatientService.STATUS_CHOICES):
                ps.status = status
            ps.result = result
            ps.is_paid = is_paid
            ps.performed_by_id = performed_by_id if performed_by_id else None
            if ordered_by_id is not None:
                ps.ordered_by_id = ordered_by_id if ordered_by_id else None

            if status == 'completed' and not ps.performed_at:
                ps.performed_at = timezone.now()

            ps.save()
            return JsonResponse({'success': True})

        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"update_service error: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': _('Xizmatni yangilashda xato yuz berdi')})

    return JsonResponse({'success': False})


@login_required
@role_required('admin', 'doctor', 'reception', 'old')
def delete_service(request, pk):
    """Xizmatni o'chirish"""
    ps = get_object_or_404(PatientService, pk=pk)
    patient_pk = ps.patient_card.pk

    if request.method == 'POST':
        if ps.status == 'ordered':
            ps.delete()
            return JsonResponse({'success': True})
        return JsonResponse({
            'success': False,
            'error': _("Bajarilgan xizmatni o'chirib bo'lmaydi")
        })

    return JsonResponse({'success': False})


# ==================== STATISTIKA ====================

@login_required
@role_required('admin', 'statistician')
def service_statistics(request):
    """Xizmatlar statistikasi dashboard"""
    # Filterlar
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    category_id = request.GET.get('category', '')
    period      = request.GET.get('period', 'month')
    patient_cat = request.GET.get('patient_category', '')
    visit_type  = request.GET.get('visit_type', '')

    qs = PatientService.objects.exclude(status='cancelled').select_related(
        'service__category', 'patient_card'
    )

    if date_from:    qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:      qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id:  qs = qs.filter(service__category_id=category_id)
    if patient_cat:  qs = qs.filter(patient_category_at_order=patient_cat)
    if visit_type:   qs = qs.filter(patient_card__visit_type=visit_type)

    # Umumiy ko'rsatkichlar
    totals = qs.aggregate(
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        count=Sum('quantity'),
        railway_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()), filter=Q(patient_category_at_order='railway')),
        nonresident_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()), filter=Q(
            patient_category_at_order='non_resident'
        )),
    )

    # Kategoriya bo'yicha
    from django.db.models import ExpressionWrapper, DecimalField, F as F_
    cat_stats = qs.values(
        'service__category__name',
        'service__category__icon',
        'service__category__id',
    ).annotate(
        count=Sum('quantity'),
        total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
    ).order_by('-total')

    # Eng ko'p ishlatiladigan xizmatlar (Top 10)
    from django.db.models import ExpressionWrapper, DecimalField, F as F_
    top_services = qs.values(
        'service__name',
        'service__category__name',
    ).annotate(
        count=Sum('quantity'),
        total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
    ).order_by('-count')[:10]

    # Vaqt bo'yicha dinamika
    if period == 'day':
        time_stats = qs.annotate(
            period=TruncDay('ordered_at')
        ).values('period').annotate(
            count=Sum('quantity'),
            total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
        ).order_by('period')
    elif period == 'year':
        time_stats = qs.annotate(
            period=TruncYear('ordered_at')
        ).values('period').annotate(
            count=Sum('quantity'),
            total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
        ).order_by('period')
    else:
        time_stats = qs.annotate(
            period=TruncMonth('ordered_at')
        ).values('period').annotate(
            count=Sum('quantity'),
            total=Sum(ExpressionWrapper(F_('price') * F_('quantity'), output_field=DecimalField())),
        ).order_by('period')

    time_labels = [
        item['period'].strftime('%Y-%m-%d' if period == 'day' else '%Y-%m' if period == 'month' else '%Y')
        for item in time_stats if item['period']
    ]
    time_counts = [item['count'] for item in time_stats if item['period']]
    time_totals = [float(item['total'] or 0) for item in time_stats if item['period']]

    categories = ServiceCategory.objects.filter(is_active=True)

    return render(request, 'services/statistics.html', {
        'totals': totals,
        'cat_stats': cat_stats,
        'top_services': top_services,
        'time_labels': json.dumps(time_labels),
        'time_counts': json.dumps(time_counts),
        'time_totals': json.dumps(time_totals),
        'cat_labels': json.dumps([c['service__category__name'] for c in cat_stats]),
        'cat_values': json.dumps([float(c['total'] or 0) for c in cat_stats]),
        'categories': categories,
        'date_from': date_from,
        'date_to': date_to,
        'selected_category': category_id,
        'selected_period': period,
        'selected_patient_cat': patient_cat,
        'current_filters': request.GET.urlencode(),
    })


# ==================== EXPORT ====================

@login_required
@role_required('admin', 'statistician')
def export_services_excel(request):
    """Eski to'g'ridan-to'g'ri export URL → statistika sahifasiga yo'naltiradi."""
    params = request.GET.urlencode()
    messages.info(request, _("Excel yuklab olish uchun '📥 Excel' tugmasini bosing."))
    return redirect(f'/services/statistics/?{params}' if params else '/services/statistics/')


# ==================== BACKGROUND EXCEL EXPORT ====================

@login_required
@role_required('admin', 'statistician')
def export_services_start(request):
    """Xizmatlar Excel task'ini Celery'ga yuboradi → {task_id}"""
    from .export_tasks import generate_services_excel
    filters = {k: v for k, v in {
        'date_from':        request.GET.get('date_from'),
        'date_to':          request.GET.get('date_to'),
        'category':         request.GET.get('svc_category') or request.GET.get('category'),
        'patient_category': request.GET.get('patient_category'),
        'visit_type':       request.GET.get('visit_type'),
    }.items() if v}
    task = generate_services_excel.delay(filters)
    return JsonResponse({'task_id': task.id})


@login_required
@role_required('admin', 'statistician')
def export_medicine_start(request):
    """Dori Excel task'ini Celery'ga yuboradi → {task_id}"""
    from .export_tasks import generate_medicine_excel
    filters = {k: v for k, v in {
        'date_from':        request.GET.get('date_from'),
        'date_to':          request.GET.get('date_to'),
        'medicine':         request.GET.get('medicine'),
        'patient_category': request.GET.get('patient_category'),
        'visit_type':       request.GET.get('visit_type'),
    }.items() if v}
    task = generate_medicine_excel.delay(filters)
    return JsonResponse({'task_id': task.id})


@login_required
def export_task_status(request, task_id):
    """Celery task holati → {status:'pending'|'done'|'error', filename?, error?}"""
    from celery.result import AsyncResult
    result = AsyncResult(task_id)
    if result.state == 'SUCCESS':
        return JsonResponse({'status': 'done', 'filename': result.result})
    if result.state == 'FAILURE':
        return JsonResponse({'status': 'error', 'error': str(result.info)})
    return JsonResponse({'status': 'pending'})


@login_required
def export_file_download(request, filename):
    """temp_exports papkasidagi faylni yuklaydi."""
    import os, re
    from django.conf import settings
    from django.http import Http404
    if not re.match(r'^[a-z_]+[0-9a-f]+\.xlsx$', filename):
        raise Http404
    filepath = os.path.join(settings.MEDIA_ROOT, 'temp_exports', filename)
    if not os.path.exists(filepath):
        raise Http404
    if filename.startswith('services'):
        display = 'xizmatlar_hisoboti.xlsx'
    elif filename.startswith('patients'):
        display = 'bemorlar_royxati.xlsx'
    else:
        display = 'dori_statistika.xlsx'
    with open(filepath, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    response['Content-Disposition'] = f'attachment; filename="{display}"'
    return response


# ==================== DORI-DARMON ====================

@login_required
def medicine_search(request):
    """AJAX — dori qidirish (savdo nomi, МНН, doza bo'yicha)"""
    from .models import Medicine
    from django.db.models import Q
    q        = request.GET.get('q', '').strip()
    category = request.GET.get('category', '')
    field    = request.GET.get('field', '')    # 'mnn' | 'form' | 'strength' — distinct list uchun

    qs = Medicine.objects.filter(is_active=True)
    if category:
        qs = qs.filter(category=category)

    if field == 'mnn':
        # МНН autocomplete — unique МНН ro'yxati
        if q:
            qs = qs.filter(mnn__icontains=q).exclude(mnn='')
        else:
            qs = qs.exclude(mnn='')
        values = list(qs.values_list('mnn', flat=True).distinct().order_by('mnn')[:30])
        return JsonResponse(values, safe=False)

    if field == 'form':
        # Форма выпуска autocomplete — unique dosage_form display ro'yxati
        if q:
            qs = qs.filter(dosage_form__icontains=q).exclude(dosage_form='')
        else:
            qs = qs.exclude(dosage_form='')
        values = list(
            qs.values_list('dosage_form', flat=True).distinct().order_by('dosage_form')[:20]
        )
        choices = dict(Medicine.DOSAGE_FORM_CHOICES)
        labels  = list({choices.get(v, v) for v in values})
        return JsonResponse(sorted(labels), safe=False)

    if field == 'strength':
        # Дозировка autocomplete
        if q:
            qs = qs.filter(strength__icontains=q).exclude(strength='')
        else:
            qs = qs.exclude(strength='')
        values = list(qs.values_list('strength', flat=True).distinct().order_by('strength')[:30])
        return JsonResponse(values, safe=False)

    # Default: savdo nomi + МНН bo'yicha to'liq qidiruv
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(mnn__icontains=q) | Q(strength__icontains=q))
    qs = qs.order_by('name')[:30]
    data = [
        {
            'id':          m.id,
            'name':        m.name,
            'mnn':         m.mnn,
            'dosage_form': m.get_dosage_form_display() if m.dosage_form else '',
            'strength':    m.strength,
            'unit':        m.unit,
            'category':    m.category,
        }
        for m in qs
    ]
    return JsonResponse(data, safe=False)


def _bounded_decimal(value, max_digits, decimal_places):
    """Decimal'ga o'giradi va maydon max_digits/decimal_places chegarasini tekshiradi.
    Chegaradan oshsa yoki noto'g'ri qiymat bo'lsa — None qaytaradi."""
    from decimal import Decimal, InvalidOperation
    try:
        d = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    limit = Decimal(10) ** (max_digits - decimal_places)
    if abs(d) >= limit:
        return None
    return d


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception', 'old')
def add_medicine(request, patient_pk):
    """Bemorga dori qo'shish — AJAX POST"""
    from .models import Medicine, PatientMedicine
    patient = get_object_or_404(PatientCard, pk=patient_pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            medicine_id  = data.get('medicine_id')
            quantity     = data.get('quantity', 1)
            price        = data.get('price', 0)
            ordered_by_id = data.get('ordered_by_id')
            notes        = data.get('notes', '')

            quantity_dec = _bounded_decimal(quantity, 10, 2)
            price_dec    = _bounded_decimal(price, 12, 2)
            if quantity_dec is None:
                return JsonResponse({'success': False, 'error': _("Miqdori noto'g'ri yoki juda katta")})
            if price_dec is None:
                return JsonResponse({'success': False, 'error': _("Narx noto'g'ri yoki juda katta (maksimal 9999999999.99 so'm)")})

            medicine = Medicine.objects.get(pk=medicine_id, is_active=True)
            pm = PatientMedicine.objects.create(
                patient_card  = patient,
                medicine      = medicine,
                quantity      = quantity_dec,
                price         = price_dec,
                ordered_by_id = ordered_by_id or None,
                notes         = notes,
            )
            return JsonResponse({
                'success': True,
                'id': pm.id,
                'medicine_name': medicine.name,
                'unit': medicine.unit,
                'quantity': str(pm.quantity),
                'price': str(pm.price),
                'total': str(pm.total_price),
                'ordered_by': pm.ordered_by.full_name if pm.ordered_by else '—',
                'ordered_at': pm.ordered_at.strftime('%d.%m.%Y %H:%M'),
            })
        except Medicine.DoesNotExist:
            return JsonResponse({'success': False, 'error': _('Dori topilmadi')})
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"add_medicine error: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': _('Dori qo\'shishda xato yuz berdi')})

    return JsonResponse({'success': False})


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception', 'old')
def update_medicine(request, pk):
    """Dori yangilash — AJAX POST"""
    from .models import PatientMedicine
    pm = get_object_or_404(PatientMedicine, pk=pk)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            quantity_dec = _bounded_decimal(data.get('quantity', pm.quantity), 10, 2)
            price_dec    = _bounded_decimal(data.get('price', pm.price), 12, 2)
            if quantity_dec is None:
                return JsonResponse({'success': False, 'error': _("Miqdori noto'g'ri yoki juda katta")})
            if price_dec is None:
                return JsonResponse({'success': False, 'error': _("Narx noto'g'ri yoki juda katta (maksimal 9999999999.99 so'm)")})

            pm.quantity      = quantity_dec
            pm.price         = price_dec
            pm.ordered_by_id = data.get('ordered_by_id') or None
            pm.notes         = data.get('notes', pm.notes)
            pm.save()
            return JsonResponse({
                'success': True,
                'total': str(pm.total_price),
                'ordered_by': pm.ordered_by.full_name if pm.ordered_by else '—',
            })
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"update_medicine error: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': _('Dorini yangilashda xato yuz berdi')})
    return JsonResponse({'success': False})


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception', 'old')
def delete_medicine(request, pk):
    """Dori o'chirish — AJAX DELETE"""
    from .models import PatientMedicine
    pm = get_object_or_404(PatientMedicine, pk=pk)
    if request.method == 'POST':
        pm.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


# ==================== DORI STATISTIKASI ====================

@login_required
def medicine_statistics(request):
    """Dori-darmonlar statistikasi"""
    from .models import Medicine, PatientMedicine
    from django.db.models import Sum, Count, Q
    from django.db.models.functions import TruncMonth, TruncDay, TruncYear

    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    medicine_id = request.GET.get('medicine', '')
    period      = request.GET.get('period', 'month')

    qs = PatientMedicine.objects.select_related(
        'medicine', 'patient_card', 'ordered_by'
    )

    if date_from:
        qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(ordered_at__date__lte=date_to)
    if medicine_id:
        qs = qs.filter(medicine_id=medicine_id)

    # Umumiy ko'rsatkichlar
    totals = qs.aggregate(
        total_records=Sum('quantity'),
        total_qty=Sum('quantity'),
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        patients=Count('patient_card', distinct=True),
    )

    # Eng ko'p ishlatiladigan dorilar TOP-20
    top_medicines = (
        qs.values('medicine__name', 'medicine__unit', 'medicine_id')
        .annotate(
            rec_count=Sum('quantity'),
            total_qty=Sum('quantity'),
            total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            patient_count=Count('patient_card', distinct=True),
        )
        .order_by('-total_sum')[:20]
    )

    # Dinamika (trend)
    if period == 'day':
        trunc_fn = TruncDay
        fmt = '%d.%m.%Y'
    elif period == 'year':
        trunc_fn = TruncYear
        fmt = '%Y'
    else:
        trunc_fn = TruncMonth
        fmt = '%Y-%m'

    trend = (
        qs.annotate(period=trunc_fn('ordered_at'))
        .values('period')
        .annotate(total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())), qty=Sum('quantity'), cnt=Sum('quantity'))
        .order_by('period')
    )
    trend_labels = [t['period'].strftime(fmt) for t in trend if t['period']]
    trend_values = [float(t['total'] or 0) for t in trend if t['period']]

    # Dorilar ro'yxati (filter uchun)
    medicines_list = Medicine.objects.filter(is_active=True).order_by('name')

    current_filters = request.GET.urlencode()

    import json
    return render(request, 'services/medicine_statistics.html', {
        'totals': totals,
        'top_medicines': top_medicines,
        'trend_labels': json.dumps(trend_labels),
        'trend_values': json.dumps(trend_values),
        'medicines_list': medicines_list,
        'date_from': date_from,
        'date_to': date_to,
        'selected_medicine': medicine_id,
        'selected_period': period,
        'current_filters': current_filters,
    })


@login_required
def export_medicine_excel(request):
    """Dori statistikasi Excel (write_only + iterator — katta ma'lumot uchun)"""
    from .models import PatientMedicine
    from django.db.models import Sum, Count
    from openpyxl.cell.cell import WriteOnlyCell

    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    medicine_id = request.GET.get('medicine', '')
    patient_cat = request.GET.get('patient_category', '')
    visit_type  = request.GET.get('visit_type', '')

    qs = PatientMedicine.objects.select_related('medicine', 'patient_card', 'ordered_by')
    if date_from:   qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:     qs = qs.filter(ordered_at__date__lte=date_to)
    if medicine_id: qs = qs.filter(medicine_id=medicine_id)
    if patient_cat: qs = qs.filter(patient_card__patient_category=patient_cat)
    if visit_type:  qs = qs.filter(patient_card__visit_type=visit_type)

    # Aggregate querylarni oldindan bajarib olamiz (iterator ishlatilmaydi)
    top = list(
        qs.values('medicine__name', 'medicine__unit')
        .annotate(
            tq=Sum('quantity'),
            tp=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            pc=Count('patient_card', distinct=True),
        )
        .order_by('-tp')
    )
    grand = sum(float(r['tp'] or 0) for r in top)

    # Stillar
    GOLD  = PatternFill('solid', fgColor='856404')
    LGOLD = PatternFill('solid', fgColor='FFF8E1')
    WHITE = PatternFill('solid', fgColor='FFFFFF')
    WFONT = Font(color='FFFFFF', bold=True, size=10)
    NORM  = Font(size=10)
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT   = Alignment(horizontal='left',   vertical='center')
    RIGHT  = Alignment(horizontal='right',  vertical='center')
    brd_s  = Side(style='thin', color='CCCCCC')
    BRD    = Border(left=brd_s, right=brd_s, top=brd_s, bottom=brd_s)

    def _h(ws, val):
        c = WriteOnlyCell(ws, value=val)
        c.fill = GOLD; c.font = WFONT; c.alignment = CENTER; c.border = BRD
        return c

    def _c(ws, val, aln=None, fill=None, fmt=None):
        c = WriteOnlyCell(ws, value=val)
        c.font = NORM; c.border = BRD
        c.alignment = aln or LEFT
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    wb = openpyxl.Workbook(write_only=True)

    # ===== Sheet 1: Dorilar bo'yicha umumiy =====
    ws1 = wb.create_sheet("Hisobot")
    for col_l, w in zip('ABCDEF', [5, 35, 12, 14, 14, 18]):
        ws1.column_dimensions[col_l].width = w

    title_val = "DORI-DARMONLAR STATISTIKASI"
    if date_from or date_to:
        title_val += f"  |  Davr: {date_from or '—'} — {date_to or '—'}"
    title_c = WriteOnlyCell(ws1, value=title_val)
    title_c.fill = GOLD; title_c.font = Font(color='FFFFFF', bold=True, size=13)
    title_c.alignment = CENTER
    ws1.append([title_c, None, None, None, None, None])

    ws1.append([_h(ws1, h) for h in
                ['№', 'Dori nomi', 'Birlik', 'Jami miqdor', 'Bemorlar', "Jami summa (so'm)"]])

    for ri, row in enumerate(top, 1):
        tp = float(row['tp'] or 0)
        fill = LGOLD if ri % 2 == 0 else WHITE
        ws1.append([
            _c(ws1, ri, CENTER, fill),
            _c(ws1, row['medicine__name'], LEFT, fill),
            _c(ws1, row['medicine__unit'], CENTER, fill),
            _c(ws1, float(row['tq'] or 0), CENTER, fill),
            _c(ws1, row['pc'], CENTER, fill),
            _c(ws1, tp, RIGHT, fill, '#,##0'),
        ])

    jami_c = WriteOnlyCell(ws1, value='JAMI:')
    jami_c.fill = GOLD; jami_c.font = WFONT; jami_c.alignment = LEFT; jami_c.border = BRD
    grand_c = WriteOnlyCell(ws1, value=grand)
    grand_c.fill = GOLD; grand_c.font = WFONT; grand_c.alignment = RIGHT
    grand_c.border = BRD; grand_c.number_format = '#,##0'
    ws1.append([jami_c, None, None, None, None, grand_c])

    # ===== Sheet 2: Batafsil ro'yxat =====
    ws2 = wb.create_sheet("Batafsil")
    for col_l, w in zip('ABCDEFGH', [5, 30, 30, 12, 14, 14, 20, 20]):
        ws2.column_dimensions[col_l].width = w

    ws2.append([_h(ws2, h) for h in
                ['№', 'Bemor', 'Dori nomi', 'Birlik', 'Miqdori', 'Narxi', 'Jami', 'Sana']])

    for ri, pm in enumerate(qs.order_by('-ordered_at').iterator(chunk_size=500), 1):
        fill = LGOLD if ri % 2 == 0 else WHITE
        ws2.append([
            _c(ws2, ri,                               CENTER, fill),
            _c(ws2, pm.patient_card.full_name,        LEFT,   fill),
            _c(ws2, pm.medicine.name,                 LEFT,   fill),
            _c(ws2, pm.medicine.unit,                 CENTER, fill),
            _c(ws2, float(pm.quantity),               CENTER, fill),
            _c(ws2, float(pm.price),                  RIGHT,  fill, '#,##0'),
            _c(ws2, float(pm.total_price),            RIGHT,  fill, '#,##0'),
            _c(ws2, pm.ordered_at.strftime('%d.%m.%Y'), CENTER, fill),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="dori_statistika.xlsx"'
    return response


# ==================== OPERATSIYA STATISTIKASI ====================

@login_required
def operation_statistics(request):
    """Operatsiya turlari bo'yicha statistika"""
    from apps.patients.models import SurgicalOperation, OperationType
    from django.db.models import Count, Q
    import json

    date_from    = request.GET.get('date_from', '')
    date_to      = request.GET.get('date_to', '')
    op_type_id   = request.GET.get('op_type', '')
    anesthesia   = request.GET.get('anesthesia', '')

    qs = SurgicalOperation.objects.select_related(
        'operation_type', 'patient_card'
    ).filter(operation_type__isnull=False)

    if date_from:
        qs = qs.filter(operation_date__gte=date_from)
    if date_to:
        qs = qs.filter(operation_date__lte=date_to)
    if op_type_id:
        qs = qs.filter(operation_type_id=op_type_id)
    if anesthesia:
        qs = qs.filter(anesthesia=anesthesia)

    # Har bir operatsiya turi bo'yicha statistika
    op_stats = (
        qs.values(
            'operation_type__id',
            'operation_type__code',
            'operation_type__name',
        )
        .annotate(
            total_count=Count('id'),
            # Bemor kategoriyasi bo'yicha
            railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            # Narkoz bo'yicha
            anesthesia_yes=Count('id', filter=Q(anesthesia='yes')),
            anesthesia_no=Count('id', filter=Q(anesthesia='no')),
            anesthesia_local=Count('id', filter=Q(anesthesia='local')),
        )
        .order_by('-total_count')
    )

    # Xizmat narxlari bilan bog'lash (agar Service da OperationType bog'liq bo'lsa)
    # SurgicalOperation da narx yo'q - faqat PatientService orqali olish mumkin
    # Shu sababli PatientService orqali operatsiya xizmatlarini topamiz
    from apps.services.models import PatientService
    from django.db.models import Sum

    # Operatsiya xizmatlari summasi (service kategoriyasi 'surgery' bo'lganlar)
    svc_totals_by_cat = {}
    surgery_svcs = (
        PatientService.objects
        .filter(service__category__category_type='surgery')
        .values('patient_card__patient_category')
        .annotate(total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())), cnt=Sum('quantity'))
    )
    for s in surgery_svcs:
        cat = s['patient_card__patient_category']
        svc_totals_by_cat[cat] = {
            'total': float(s['total'] or 0),
            'count': s['cnt']
        }

    # Umumiy ko'rsatkichlar
    totals = qs.aggregate(
        total=Count('id'),
        railway=Count('id', filter=Q(patient_card__patient_category='railway')),
        paid=Count('id', filter=Q(patient_card__patient_category='paid')),
        nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        with_anesthesia=Count('id', filter=Q(anesthesia='yes')),
        local_anesthesia=Count('id', filter=Q(anesthesia='local')),
        no_anesthesia=Count('id', filter=Q(anesthesia='no')),
    )

    # Trend (oylik)
    from django.db.models.functions import TruncMonth
    trend = (
        qs.annotate(month=TruncMonth('operation_date'))
        .values('month')
        .annotate(cnt=Count('id'))
        .order_by('month')
    )
    trend_labels = [t['month'].strftime('%Y-%m') for t in trend if t['month']]
    trend_values = [t['cnt'] for t in trend if t['month']]

    op_types_list = OperationType.objects.filter(is_active=True).order_by('name')
    current_filters = request.GET.urlencode()

    return render(request, 'services/operation_statistics.html', {
        'op_stats': op_stats,
        'totals': totals,
        'svc_totals_by_cat': svc_totals_by_cat,
        'trend_labels': json.dumps(trend_labels),
        'trend_values': json.dumps(trend_values),
        'op_types_list': op_types_list,
        'date_from': date_from,
        'date_to': date_to,
        'selected_op_type': op_type_id,
        'selected_anesthesia': anesthesia,
        'current_filters': current_filters,
    })


@login_required
def export_operation_excel(request):
    """Operatsiya statistikasi Excel export"""
    from apps.patients.models import SurgicalOperation
    from django.db.models import Count, Sum, Q, ExpressionWrapper, DecimalField, F
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    op_type_id = request.GET.get('op_type', '')
    anesthesia = request.GET.get('anesthesia', '')

    qs = SurgicalOperation.objects.select_related(
        'operation_type', 'patient_card'
    ).filter(operation_type__isnull=False)

    if date_from:    qs = qs.filter(operation_date__gte=date_from)
    if date_to:      qs = qs.filter(operation_date__lte=date_to)
    if op_type_id:   qs = qs.filter(operation_type_id=op_type_id)
    if anesthesia:   qs = qs.filter(anesthesia=anesthesia)

    # Stillar
    BLUE  = PatternFill('solid', fgColor='1F4E79')
    LBLUE = PatternFill('solid', fgColor='D6E4F0')
    GREEN = PatternFill('solid', fgColor='E9F7EF')
    YELL  = PatternFill('solid', fgColor='FFF3CD')
    RED   = PatternFill('solid', fgColor='FDEDEC')
    WHITE = PatternFill('solid', fgColor='FFFFFF')
    TOTAL = PatternFill('solid', fgColor='145A32')

    WF   = Font(color='FFFFFF', bold=True, size=10)
    BOLD = Font(bold=True, size=10)
    NORM = Font(size=9)
    C    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    R    = Alignment(horizontal='right',  vertical='center')
    brd  = Side(style='thin', color='CCCCCC')
    BRD  = Border(left=brd, right=brd, top=brd, bottom=brd)

    wb = openpyxl.Workbook()

    # ===== SHEET 1: OPERATSIYA TURI BO'YICHA =====
    ws1 = wb.active
    ws1.title = "Operatsiya turlari"

    col_w = [5, 12, 35, 12, 12, 14, 16, 14, 14, 12, 12]
    for ci, w in enumerate(col_w, 1):
        from openpyxl.utils import get_column_letter
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.merge_cells('A1:K1')
    c = ws1.cell(row=1, column=1, value="OPERATSIYA TURLARI BO'YICHA STATISTIKA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = C
    ws1.row_dimensions[1].height = 30

    if date_from or date_to:
        ws1.merge_cells('A2:K2')
        c = ws1.cell(row=2, column=1,
            value=f"Davr: {date_from or '—'} dan {date_to or '—'} gacha")
        c.font = BOLD; c.alignment = C
        ws1.row_dimensions[2].height = 18
        hdr_row = 3
    else:
        hdr_row = 2

    headers = [
        '№', 'Kod', 'Operatsiya nomi',
        'Jami\nsoni',
        "Temir\nyo'lchi",
        'Pullik',
        'Norezident',
        "Narkoz\nbilan",
        'Mahalliy\nnarkoz',
        'Narkozsiz',
        'Asorat\nbor',
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=hdr_row, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws1.row_dimensions[hdr_row].height = 36

    op_stats = (
        qs.values(
            'operation_type__id',
            'operation_type__code',
            'operation_type__name',
        )
        .annotate(
            total_count=Count('id'),
            railway=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            anes_yes=Count('id', filter=Q(anesthesia='yes')),
            anes_local=Count('id', filter=Q(anesthesia='local')),
            anes_no=Count('id', filter=Q(anesthesia='no')),
            has_complication=Count('id', filter=~Q(complication='') & Q(complication__isnull=False)),
        )
        .order_by('-total_count')
    )

    totals_row = [0] * 9
    dr = hdr_row + 1
    for ri, op in enumerate(op_stats, 1):
        vals = [
            ri,
            op['operation_type__code'] or '—',
            op['operation_type__name'],
            op['total_count'],
            op['railway'],
            op['paid'],
            op['nonresident'],
            op['anes_yes'],
            op['anes_local'],
            op['anes_no'],
            op['has_complication'],
        ]
        for j in range(3, 11): totals_row[j-3] += vals[j] if j < len(vals) else 0
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=dr, column=col, value=val)
            c.font = NORM; c.border = BRD
            c.alignment = C if col != 3 else L
            fill = [LBLUE, YELL, WHITE, RED][ri % 4] if col > 3 else WHITE
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='F0F6FC')
        ws1.row_dimensions[dr].height = 18
        dr += 1

    # Jami
    tot_vals = ['', '', 'JAMI:'] + [qs.aggregate(t=Count('id'))['t']] + [
        qs.filter(patient_card__patient_category='railway').count(),
        qs.filter(patient_card__patient_category='paid').count(),
        qs.filter(patient_card__patient_category='non_resident').count(),
        qs.filter(anesthesia='yes').count(),
        qs.filter(anesthesia='local').count(),
        qs.filter(anesthesia='no').count(),
        qs.exclude(complication='').exclude(complication__isnull=True).count(),
    ]
    for col, val in enumerate(tot_vals, 1):
        c = ws1.cell(row=dr, column=col, value=val)
        c.fill = BLUE; c.font = WF; c.alignment = C if col != 3 else L; c.border = BRD
    ws1.row_dimensions[dr].height = 22

    # ===== SHEET 2: BATAFSIL RO'YXAT =====
    ws2 = wb.create_sheet("Batafsil ro'yxat")
    h2_widths = [5, 14, 28, 14, 35, 20, 16, 14, 30]
    for ci, w in enumerate(h2_widths, 1):
        from openpyxl.utils import get_column_letter
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells('A1:I1')
    c = ws2.cell(row=1, column=1, value="OPERATSIYALAR BATAFSIL RO'YXATI")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws2.row_dimensions[1].height = 26

    h2 = ['№', 'Sana', 'Bemor', 'Bayonnoma', 'Operatsiya turi', 'Bemor turi', 'Narkoz', "Asorat bor?", 'Asorat tavsifi']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws2.row_dimensions[2].height = 22

    for ri, op in enumerate(qs.order_by('-operation_date'), 1):
        cat_display = {
            'railway': "Temir yo'lchi",
            'paid': 'Pullik',
            'non_resident': 'Norezident',
        }.get(op.patient_card.patient_category, '—')

        has_comp = bool(op.complication and op.complication.strip())
        data = [
            ri,
            op.operation_date.strftime('%d.%m.%Y') if op.operation_date else '—',
            op.patient_card.full_name,
            op.patient_card.medical_record_number,
            str(op.operation_type) if op.operation_type else op.operation_name or '—',
            cat_display,
            op.get_anesthesia_display() if op.anesthesia else '—',
            'Ha' if has_comp else "Yo'q",
            op.complication or '—',
        ]
        for col, val in enumerate(data, 1):
            c = ws2.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD
            c.alignment = C if col in (1,2,4,6,7,8) else L
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='F0F6FC')
            if col == 8 and has_comp: c.fill = PatternFill('solid', fgColor='FDEDEC')
        ws2.row_dimensions[ri+2].height = 17

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="operatsiya_statistika.xlsx"'
    wb.save(response)
    return response


# ==================== XIZMAT OPERATSIYALARI STATISTIKASI ====================

@login_required
def operation_service_statistics(request):
    """is_operation=True bo'lgan xizmatlar bo'yicha statistika"""
    from django.db.models import Sum, Count, Q
    from django.db.models.functions import TruncMonth
    import json

    date_from    = request.GET.get('date_from', '')
    date_to      = request.GET.get('date_to', '')
    category_id  = request.GET.get('category', '')
    service_id   = request.GET.get('service', '')
    patient_cat  = request.GET.get('patient_category', '')

    # Faqat is_operation=True xizmatlar
    qs = PatientService.objects.filter(
        service__is_operation=True
    ).select_related('service__category', 'patient_card')

    if date_from:    qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:      qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id:  qs = qs.filter(service__category_id=category_id)
    if service_id:   qs = qs.filter(service_id=service_id)
    if patient_cat:  qs = qs.filter(patient_card__patient_category=patient_cat)

    # Umumiy ko'rsatkichlar
    totals = qs.aggregate(
        total_count=Sum('quantity'),
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        patients=Count('patient_card', distinct=True),
        railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
        paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
        nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        railway_sum=Sum('price', filter=Q(patient_card__patient_category='railway')),
        paid_sum=Sum('price', filter=Q(patient_card__patient_category='paid')),
        nonresident_sum=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
    )

    # Har bir operatsiya xizmati bo'yicha batafsil
    op_stats = (
        qs.values(
            'service__id',
            'service__name',
            'service__code',
            'service__category__name',
            'service__category__icon',
        )
        .annotate(
            total_count=Sum('quantity'),
            total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            patients=Count('patient_card', distinct=True),
            railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            railway_sum=Sum('price', filter=Q(patient_card__patient_category='railway')),
            paid_sum=Sum('price', filter=Q(patient_card__patient_category='paid')),
            nonresident_sum=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
        )
        .order_by('-total_count')
    )

    # Kategoriya bo'yicha guruhlash
    cat_stats = (
        qs.values('service__category__name', 'service__category__icon')
        .annotate(
            count=Count('id'),
            total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            railway=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        )
        .order_by('-total')
    )

    # Oylik trend
    trend = (
        qs.annotate(month=TruncMonth('ordered_at'))
        .values('month')
        .annotate(cnt=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())))
        .order_by('month')
    )
    trend_labels = [t['month'].strftime('%Y-%m') for t in trend if t['month']]
    trend_counts = [t['cnt'] for t in trend if t['month']]
    trend_sums   = [float(t['total'] or 0) for t in trend if t['month']]

    # Filter uchun ro'yxatlar
    from .models import ServiceCategory
    op_categories = ServiceCategory.objects.filter(
        services__is_operation=True, is_active=True
    ).distinct().order_by('name')

    op_services = Service.objects.filter(
        is_operation=True, is_active=True
    ).order_by('category__name', 'name')

    current_filters = request.GET.urlencode()

    return render(request, 'services/operation_service_statistics.html', {
        'totals': totals,
        'op_stats': op_stats,
        'cat_stats': cat_stats,
        'trend_labels': json.dumps(trend_labels),
        'trend_counts': json.dumps(trend_counts),
        'trend_sums':   json.dumps(trend_sums),
        'op_categories': op_categories,
        'op_services': op_services,
        'date_from': date_from,
        'date_to': date_to,
        'selected_category': category_id,
        'selected_service': service_id,
        'selected_patient_cat': patient_cat,
        'current_filters': current_filters,
    })


@login_required
def export_operation_service_excel(request):
    """Operatsiya xizmatlari statistikasi Excel"""
    from django.db.models import Sum, Count, Q
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    category_id = request.GET.get('category', '')
    service_id  = request.GET.get('service', '')
    patient_cat = request.GET.get('patient_category', '')
    visit_type  = request.GET.get('visit_type', '')

    qs = PatientService.objects.filter(
        service__is_operation=True
    ).select_related('service__category', 'patient_card')
    if date_from:   qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:     qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id: qs = qs.filter(service__category_id=category_id)
    if service_id:  qs = qs.filter(service_id=service_id)
    if patient_cat: qs = qs.filter(patient_card__patient_category=patient_cat)
    if visit_type:  qs = qs.filter(patient_card__visit_type=visit_type)

    # Stillar
    BLUE  = PatternFill('solid', fgColor='1F4E79')
    LBLUE = PatternFill('solid', fgColor='D6E4F0')
    GREEN = PatternFill('solid', fgColor='E9F7EF')
    YELL  = PatternFill('solid', fgColor='FFF3CD')
    RED   = PatternFill('solid', fgColor='FDEDEC')
    TOTAL = PatternFill('solid', fgColor='145A32')
    WHITE = PatternFill('solid', fgColor='FFFFFF')
    WF    = Font(color='FFFFFF', bold=True, size=10)
    BOLD  = Font(bold=True, size=10)
    NORM  = Font(size=9)
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    R = Alignment(horizontal='right',  vertical='center')
    brd = Side(style='thin', color='CCCCCC')
    BRD = Border(left=brd, right=brd, top=brd, bottom=brd)

    wb = openpyxl.Workbook()

    # ===== SHEET 1: OPERATSIYA XIZMATLARI =====
    ws1 = wb.active
    ws1.title = "Operatsiya xizmatlari"

    col_widths = [5, 12, 35, 18, 10, 10, 10, 10, 14, 14, 14, 16]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.merge_cells('A1:L1')
    c = ws1.cell(row=1, column=1,
        value="OPERATSIYA XIZMATLARI BO'YICHA STATISTIKA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = C
    ws1.row_dimensions[1].height = 30

    if date_from or date_to:
        ws1.merge_cells('A2:L2')
        ws1.cell(row=2, column=1,
            value=f"Davr: {date_from or '—'} dan {date_to or '—'} gacha").font = BOLD
        ws1.row_dimensions[2].height = 18
        hdr = 3
    else:
        hdr = 2

    headers = [
        '№', 'Kod', 'Operatsiya nomi', 'Kategoriya',
        'Jami\nsoni', 'Bemorlar',
        "TY\nsoni", "Pullik\nsoni", "Nores.\nsoni",
        "TY\nsumma", "Pullik\nsumma", "Nores.\nsumma",
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=hdr, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws1.row_dimensions[hdr].height = 36

    op_stats = (
        qs.values(
            'service__code', 'service__name', 'service__category__name'
        )
        .annotate(
            tc=Count('id'),
            pts=Count('patient_card', distinct=True),
            rc=Count('id', filter=Q(patient_card__patient_category='railway')),
            pc=Count('id', filter=Q(patient_card__patient_category='paid')),
            nc=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            rs=Sum('price', filter=Q(patient_card__patient_category='railway')),
            ps=Sum('price', filter=Q(patient_card__patient_category='paid')),
            ns=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
        )
        .order_by('service__category__name', '-tc')
    )

    dr = hdr + 1
    grand = {'tc':0,'pts':0,'rc':0,'pc':0,'nc':0,'rs':0,'ps':0,'ns':0}

    for ri, op in enumerate(op_stats, 1):
        vals = [
            ri,
            op['service__code'] or '—',
            op['service__name'],
            op['service__category__name'],
            op['tc'], op['pts'],
            op['rc'], op['pc'], op['nc'],
            float(op['rs'] or 0), float(op['ps'] or 0), float(op['ns'] or 0),
        ]
        for k in ('tc','pts','rc','pc','nc'):
            grand[k] += op[k] or 0
        for k, col in (('rs',10),('ps',11),('ns',12)):
            grand[k] += float(op[k] or 0)

        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=dr, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,5,6,7,8,9) else (R if col > 9 else L)
            if col > 9: c.number_format = '#,##0'
        ws1.row_dimensions[dr].height = 18
        dr += 1

    # Jami qatori
    tot = ['', '', 'JAMI:', '',
           grand['tc'], grand['pts'],
           grand['rc'], grand['pc'], grand['nc'],
           grand['rs'], grand['ps'], grand['ns']]
    for col, val in enumerate(tot, 1):
        c = ws1.cell(row=dr, column=col, value=val)
        c.fill = BLUE; c.font = WF; c.border = BRD
        c.alignment = C if col in (1,5,6,7,8,9) else (R if col > 9 else L)
        if col > 9: c.number_format = '#,##0'
    ws1.row_dimensions[dr].height = 24

    # Umumiy summa qatori
    dr += 1
    total_sum = grand['rs'] + grand['ps'] + grand['ns']
    ws1.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=9)
    c = ws1.cell(row=dr, column=1, value="UMUMIY DAROMAD:")
    c.fill = TOTAL; c.font = WF; c.alignment = L; c.border = BRD
    c2 = ws1.cell(row=dr, column=10, value=total_sum)
    c2.fill = TOTAL; c2.font = Font(color='FFFFFF', bold=True, size=12)
    c2.number_format = '#,##0'; c2.alignment = R; c2.border = BRD
    ws1.merge_cells(start_row=dr, start_column=10, end_row=dr, end_column=12)
    ws1.row_dimensions[dr].height = 26

    # ===== SHEET 2: KATEGORIYA BO'YICHA =====
    ws2 = wb.create_sheet("Kategoriya bo'yicha")
    for ci, w in enumerate([5, 25, 10, 12, 14, 14, 14, 16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells('A1:H1')
    c = ws2.cell(row=1, column=1, value="KATEGORIYA BO'YICHA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws2.row_dimensions[1].height = 26

    h2 = ['№', 'Kategoriya', 'Jami soni', 'Bemorlar',
          "TY soni", "Pullik soni", "Nores. soni", "Jami summa"]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws2.row_dimensions[2].height = 24

    cat_stats = (
        qs.values('service__category__name', 'service__category__icon')
        .annotate(
            cnt=Count('id'), pts=Count('patient_card', distinct=True),
            rc=Count('id', filter=Q(patient_card__patient_category='railway')),
            pc=Count('id', filter=Q(patient_card__patient_category='paid')),
            nc=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        )
        .order_by('-total')
    )

    cat_grand = 0
    for ri, cat in enumerate(cat_stats, 1):
        icon = cat['service__category__icon'] or ''
        name = f"{icon} {cat['service__category__name']}"
        vals = [ri, name, cat['cnt'], cat['pts'],
                cat['rc'], cat['pc'], cat['nc'],
                float(cat['total'] or 0)]
        cat_grand += float(cat['total'] or 0)
        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,3,4,5,6,7) else (R if col == 8 else L)
            if col == 8: c.number_format = '#,##0'
        ws2.row_dimensions[ri+2].height = 20

    last2 = cat_stats.count() + 3
    ws2.merge_cells(start_row=last2, start_column=1, end_row=last2, end_column=7)
    c = ws2.cell(row=last2, column=1, value="JAMI:")
    c.fill = BLUE; c.font = WF; c.alignment = L; c.border = BRD
    c8 = ws2.cell(row=last2, column=8, value=cat_grand)
    c8.fill = BLUE; c8.font = WF; c8.number_format = '#,##0'
    c8.alignment = R; c8.border = BRD
    ws2.row_dimensions[last2].height = 22

    # ===== SHEET 3: BATAFSIL =====
    ws3 = wb.create_sheet("Batafsil")
    for ci, w in enumerate([5, 14, 28, 28, 16, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws3.merge_cells('A1:G1')
    c = ws3.cell(row=1, column=1, value="BATAFSIL RO'YXAT")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws3.row_dimensions[1].height = 26

    h3 = ['№', 'Sana', 'Bemor', 'Operatsiya nomi', 'Bemor turi', "Narx (so'm)", "Jami (so'm)"]
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws3.row_dimensions[2].height = 22

    CAT_MAP = {'railway': "Temir yo'lchi", 'paid': 'Pullik', 'non_resident': 'Norezident'}
    for ri, ps in enumerate(qs.order_by('-ordered_at'), 1):
        data = [
            ri,
            ps.ordered_at.strftime('%d.%m.%Y'),
            ps.patient_card.full_name,
            ps.service.name,
            CAT_MAP.get(ps.patient_card.patient_category, '—'),
            float(ps.price),
            float(ps.total_price),
        ]
        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(data, 1):
            c = ws3.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,2,5) else (R if col > 5 else L)
            if col > 5: c.number_format = '#,##0'
        ws3.row_dimensions[ri+2].height = 17

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="operatsiya_xizmatlari.xlsx"'
    wb.save(response)
    return response


# ==================== BIRLASHTIRILGAN STATISTIKA ====================

@login_required
def statistics_combined(request):
    """Barcha statistikalarni bitta sahifada — tab bilan"""
    import json
    from django.db.models import Sum, Count, Q
    from django.db.models.functions import TruncMonth, TruncDay, TruncYear
    from .models import PatientMedicine, Medicine, ServiceCategory

    active_tab = request.GET.get('tab', 'services')

    # --- UMUMIY FILTERLAR ---
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    patient_cat = request.GET.get('patient_category', '')
    period      = request.GET.get('period', 'month')
    visit_type  = request.GET.get('visit_type', '')

    # ==================== 1. XIZMATLAR ====================
    svc_qs = PatientService.objects.exclude(
        status='cancelled'
    ).select_related('service__category', 'patient_card')
    if date_from:    svc_qs = svc_qs.filter(ordered_at__date__gte=date_from)
    if date_to:      svc_qs = svc_qs.filter(ordered_at__date__lte=date_to)
    if patient_cat:  svc_qs = svc_qs.filter(patient_category_at_order=patient_cat)
    if visit_type:   svc_qs = svc_qs.filter(patient_card__visit_type=visit_type)
    if request.GET.get('svc_category'):
        svc_qs = svc_qs.filter(service__category_id=request.GET['svc_category'])

    svc_totals = svc_qs.aggregate(
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        count=Sum('quantity'),
        railway_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()), filter=Q(patient_category_at_order='railway')),
        paid_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()), filter=Q(patient_category_at_order='paid')),
        nonresident_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()), filter=Q(patient_category_at_order='non_resident')),
        railway_count=Sum('quantity', filter=Q(patient_category_at_order='railway')),
        paid_count=Sum('quantity', filter=Q(patient_category_at_order='paid')),
        nonresident_count=Sum('quantity', filter=Q(patient_category_at_order='non_resident')),
    )
    svc_cat_stats = svc_qs.values(
        'service__category__name', 'service__category__icon', 'service__category__id'
    ).annotate(count=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()))).order_by('-total')

    top_services = svc_qs.values(
        'service__name', 'service__category__name'
    ).annotate(count=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField()))).order_by('-count')[:15]

    if period == 'day':
        from django.db.models.functions import TruncDay
        trunc_fn = TruncDay
        fmt = '%d.%m.%Y'
    elif period == 'year':
        from django.db.models.functions import TruncYear
        trunc_fn = TruncYear
        fmt = '%Y'
    else:
        trunc_fn = TruncMonth
        fmt = '%Y-%m'

    svc_trend = (
        svc_qs.annotate(p=trunc_fn('ordered_at'))
        .values('p').annotate(cnt=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())))
        .order_by('p')
    )
    svc_trend_labels = [t['p'].strftime(fmt) for t in svc_trend if t['p']]
    svc_trend_values = [float(t['total'] or 0) for t in svc_trend if t['p']]

    # ==================== 2. DORILAR ====================
    med_qs = PatientMedicine.objects.select_related('medicine', 'patient_card')
    if date_from:   med_qs = med_qs.filter(ordered_at__date__gte=date_from)
    if date_to:     med_qs = med_qs.filter(ordered_at__date__lte=date_to)
    if patient_cat: med_qs = med_qs.filter(patient_card__patient_category=patient_cat)
    if visit_type:  med_qs = med_qs.filter(patient_card__visit_type=visit_type)
    if request.GET.get('medicine'):
        med_qs = med_qs.filter(medicine_id=request.GET['medicine'])

    med_totals = med_qs.aggregate(
        total_records=Sum('quantity'),
        total_qty=Sum('quantity'),
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        patients=Count('patient_card', distinct=True),
    )
    top_medicines = (
        med_qs.values('medicine__name', 'medicine__unit')
        .annotate(
            rec_count=Sum('quantity'),
            total_qty=Sum('quantity'),
            total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            patient_count=Count('patient_card', distinct=True),
        ).order_by('-total_sum')[:15]
    )
    med_trend = (
        med_qs.annotate(p=TruncMonth('ordered_at'))
        .values('p').annotate(total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())), qty=Sum('quantity'))
        .order_by('p')
    )
    med_trend_labels = [t['p'].strftime('%Y-%m') for t in med_trend if t['p']]
    med_trend_values = [float(t['total'] or 0) for t in med_trend if t['p']]

    # ==================== 3. OPERATSIYA (SurgicalOperation) ====================
    from apps.patients.models import SurgicalOperation, OperationType
    op_qs = SurgicalOperation.objects.filter(
        operation_type__isnull=False
    ).select_related('operation_type', 'patient_card')
    if date_from:   op_qs = op_qs.filter(operation_date__gte=date_from)
    if date_to:     op_qs = op_qs.filter(operation_date__lte=date_to)
    if patient_cat: op_qs = op_qs.filter(patient_card__patient_category=patient_cat)
    if request.GET.get('op_type'):
        op_qs = op_qs.filter(operation_type_id=request.GET['op_type'])

    op_totals = op_qs.aggregate(
        total=Count('id'),
        railway=Count('id', filter=Q(patient_card__patient_category='railway')),
        paid=Count('id', filter=Q(patient_card__patient_category='paid')),
        nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        with_anesthesia=Count('id', filter=Q(anesthesia='yes')),
        local_anesthesia=Count('id', filter=Q(anesthesia='local')),
        no_anesthesia=Count('id', filter=Q(anesthesia='no')),
    )
    op_stats = (
        op_qs.values('operation_type__id', 'operation_type__code', 'operation_type__name')
        .annotate(
            total_count=Count('id'),
            railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            anesthesia_yes=Count('id', filter=Q(anesthesia='yes')),
            anesthesia_local=Count('id', filter=Q(anesthesia='local')),
            anesthesia_no=Count('id', filter=Q(anesthesia='no')),
        ).order_by('-total_count')
    )
    op_trend = (
        op_qs.annotate(p=TruncMonth('operation_date'))
        .values('p').annotate(cnt=Count('id'))
        .order_by('p')
    )
    op_trend_labels = [t['p'].strftime('%Y-%m') for t in op_trend if t['p']]
    op_trend_values = [t['cnt'] for t in op_trend if t['p']]

    # ==================== 4. OPERATSIYA XIZMATLARI ====================
    opx_qs = PatientService.objects.filter(
        service__is_operation=True
    ).select_related('service__category', 'patient_card')
    if date_from:   opx_qs = opx_qs.filter(ordered_at__date__gte=date_from)
    if date_to:     opx_qs = opx_qs.filter(ordered_at__date__lte=date_to)
    if patient_cat: opx_qs = opx_qs.filter(patient_card__patient_category=patient_cat)
    if request.GET.get('opx_category'):
        opx_qs = opx_qs.filter(service__category_id=request.GET['opx_category'])

    opx_totals = opx_qs.aggregate(
        total_count=Sum('quantity'),
        total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        patients=Count('patient_card', distinct=True),
        railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
        paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
        nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        railway_sum=Sum('price', filter=Q(patient_card__patient_category='railway')),
        paid_sum=Sum('price', filter=Q(patient_card__patient_category='paid')),
        nonresident_sum=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
    )
    opx_stats = (
        opx_qs.values(
            'service__id', 'service__name', 'service__code',
            'service__category__name', 'service__category__icon',
        ).annotate(
            total_count=Sum('quantity'),
            total_sum=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            patients=Count('patient_card', distinct=True),
            railway_count=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid_count=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident_count=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            railway_sum=Sum('price', filter=Q(patient_card__patient_category='railway')),
            paid_sum=Sum('price', filter=Q(patient_card__patient_category='paid')),
            nonresident_sum=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
        ).order_by('-total_count')
    )
    opx_cat_stats = (
        opx_qs.values('service__category__name', 'service__category__icon')
        .annotate(count=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
            railway=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
        ).order_by('-total')
    )
    opx_trend = (
        opx_qs.annotate(p=TruncMonth('ordered_at'))
        .values('p').annotate(cnt=Sum('quantity'), total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())))
        .order_by('p')
    )
    opx_trend_labels = [t['p'].strftime('%Y-%m') for t in opx_trend if t['p']]
    opx_trend_counts = [t['cnt'] for t in opx_trend if t['p']]
    opx_trend_sums   = [float(t['total'] or 0) for t in opx_trend if t['p']]

    # Filter ro'yxatlar
    categories       = ServiceCategory.objects.filter(is_active=True)
    op_categories    = ServiceCategory.objects.filter(services__is_operation=True, is_active=True).distinct()
    medicines_list   = Medicine.objects.filter(is_active=True).order_by('name')
    op_types_list    = OperationType.objects.filter(is_active=True).order_by('name')

    current_filters = request.GET.urlencode()

    return render(request, 'services/statistics_combined.html', {
        'active_tab': active_tab,
        'date_from': date_from,
        'date_to': date_to,
        'patient_cat': patient_cat,
        'period': period,
        'visit_type': visit_type,
        'current_filters': current_filters,

        # Xizmatlar
        'svc_totals': svc_totals,
        'svc_cat_stats': svc_cat_stats,
        'top_services': top_services,
        'svc_trend_labels': json.dumps(svc_trend_labels),
        'svc_trend_values': json.dumps(svc_trend_values),
        'categories': categories,
        'selected_svc_category': request.GET.get('svc_category', ''),
        'selected_period': period,

        # Dorilar
        'med_totals': med_totals,
        'top_medicines': top_medicines,
        'med_trend_labels': json.dumps(med_trend_labels),
        'med_trend_values': json.dumps(med_trend_values),
        'medicines_list': medicines_list,
        'selected_medicine': request.GET.get('medicine', ''),

        # Operatsiyalar (SurgicalOperation)
        'op_totals': op_totals,
        'op_stats': op_stats,
        'op_trend_labels': json.dumps(op_trend_labels),
        'op_trend_values': json.dumps(op_trend_values),
        'op_types_list': op_types_list,
        'selected_op_type': request.GET.get('op_type', ''),

        # Operatsiya xizmatlari
        'opx_totals': opx_totals,
        'opx_stats': opx_stats,
        'opx_cat_stats': opx_cat_stats,
        'opx_trend_labels': json.dumps(opx_trend_labels),
        'opx_trend_counts': json.dumps(opx_trend_counts),
        'opx_trend_sums':   json.dumps(opx_trend_sums),
        'op_categories': op_categories,
        'selected_opx_category': request.GET.get('opx_category', ''),
        'selected_patient_cat': patient_cat,
    })


# ==================== OPERATSIYALARNI BELGILASH ====================

@login_required
@role_required('admin', 'statistician')
def mark_operations(request):
    """Xizmatlarni operatsiya deb belgilash sahifasi"""
    from .models import ServiceCategory

    if request.method == 'POST':
        action     = request.POST.get('action')
        service_ids = request.POST.getlist('service_ids')
        if service_ids:
            is_op = (action == 'mark')
            updated = Service.objects.filter(pk__in=service_ids).update(is_operation=is_op)
            if is_op:
                messages.success(request, _("✅ %(count)s ta xizmat operatsiya deb belgilandi.") % {'count': updated})
            else:
                messages.warning(request, _("❌ %(count)s ta xizmatdan operatsiya belgisi olib tashlandi.") % {'count': updated})
        else:
            messages.error(request, _("Hech bir xizmat tanlanmadi."))
        return redirect('mark_operations')

    # Kategoriyalar va ularning xizmatlari (bitta so'rov bilan)
    from django.db.models import Prefetch
    categories_raw = ServiceCategory.objects.filter(is_active=True).order_by('name').prefetch_related(
        Prefetch('services', queryset=Service.objects.filter(is_active=True).order_by('name'), to_attr='services_list')
    )
    categories = []
    total_services = 0
    total_operations = 0

    for cat in categories_raw:
        svcs = cat.services_list
        if not svcs:
            continue
        op_count  = sum(1 for s in svcs if s.is_operation)
        cat.total_count   = len(svcs)
        cat.op_count      = op_count
        categories.append(cat)
        total_services   += len(svcs)
        total_operations += op_count

    return render(request, 'services/mark_operations.html', {
        'categories':       categories,
        'total_services':   total_services,
        'total_operations': total_operations,
    })

# ===== PAKET TIZIMI =====

@login_required
def package_list(request):
    from apps.services.models import ServicePackage
    packages = ServicePackage.objects.filter(
        owner=request.user, is_active=True
    ).prefetch_related('items__service')
    patient_id  = request.GET.get('patient_id')
    patient_cat = 'railway'
    if patient_id:
        try:
            from apps.patients.models import PatientCard
            p = PatientCard.objects.get(pk=patient_id)
            patient_cat = p.patient_category or 'railway'
        except Exception:
            pass
    data = []
    for pkg in packages:
        items = []
        for item in pkg.items.all():
            price = item.service.price_for_patient(patient_cat)
            items.append({
                'service_id':   item.service.pk,
                'service_name': item.service.name,
                'category':     item.service.category.name if item.service.category else '',
                'quantity':     item.quantity,
                'price':        float(price),
            })
        data.append({'id': pkg.pk, 'name': pkg.name, 'items': items})
    return JsonResponse(data, safe=False)


@login_required
@require_POST
def package_add(request):
    from apps.services.models import ServicePackage, ServicePackageItem, Service
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': _('JSON xato')})
    name     = body.get('name', '').strip()
    services = body.get('services', [])
    if not name:
        return JsonResponse({'success': False, 'error': _('Nom kiritilmagan')})
    pkg = ServicePackage.objects.create(name=name, owner=request.user)
    for i, svc in enumerate(services):
        try:
            service = Service.objects.get(pk=svc['id'], is_active=True)
            ServicePackageItem.objects.create(
                package=pkg, service=service,
                quantity=svc.get('qty', 1), sort_order=i
            )
        except Exception:
            continue
    return JsonResponse({'success': True, 'id': pkg.pk, 'name': pkg.name})


@login_required
@require_POST
def package_delete(request, pk):
    from apps.services.models import ServicePackage
    ServicePackage.objects.filter(pk=pk, owner=request.user).update(is_active=False)
    return JsonResponse({'success': True})


@login_required
@require_POST
def package_remove_service(request, pk):
    from apps.services.models import ServicePackage, ServicePackageItem
    try:
        body = json.loads(request.body)
        svc_id = body.get('service_id')
    except Exception:
        return JsonResponse({'success': False})
    ServicePackageItem.objects.filter(
        package__pk=pk, package__owner=request.user, service_id=svc_id
    ).delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def package_toggle_service(request):
    """Xizmatni shifokorning shaxsiy paketiga qo'shish yoki o'chirish (toggle)"""
    from apps.services.models import ServicePackage, ServicePackageItem
    try:
        body = json.loads(request.body)
        service_id = body.get('service_id')
    except Exception:
        return JsonResponse({'success': False, 'error': _('JSON xato')})

    if not service_id:
        return JsonResponse({'success': False, 'error': _('service_id yo\'q')})

    service = get_object_or_404(Service, pk=service_id, is_active=True)

    # Foydalanuvchining shaxsiy paketini topish yoki yaratish
    package = ServicePackage.objects.filter(
        owner=request.user, is_active=True
    ).first()

    if not package:
        package = ServicePackage.objects.create(
            name='Mening paketim',
            owner=request.user,
            is_active=True
        )

    # Toggle: bor bo'lsa o'chirish, yo'q bo'lsa qo'shish
    existing = ServicePackageItem.objects.filter(package=package, service=service).first()
    if existing:
        existing.delete()
        in_package = False
    else:
        sort_order = ServicePackageItem.objects.filter(package=package).count()
        ServicePackageItem.objects.create(
            package=package,
            service=service,
            quantity=1,
            sort_order=sort_order
        )
        in_package = True

    return JsonResponse({
        'success': True,
        'in_package': in_package,
        'package_id': package.pk,
        'service_name': service.name,
    })
