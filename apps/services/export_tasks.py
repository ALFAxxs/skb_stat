"""
Background Excel export tasks (Celery + pandas).
"""
import os
import uuid
import logging

from celery import shared_task
from django.conf import settings

logger = logging.getLogger(__name__)


def _export_dir():
    d = os.path.join(settings.MEDIA_ROOT, 'temp_exports')
    os.makedirs(d, exist_ok=True)
    return d


# ──────────────────────────────────────────────────────────────────────────────
# Xizmatlar hisoboti
# ──────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, max_retries=0, time_limit=600, soft_time_limit=560,
             name='services.export_services_excel')
def generate_services_excel(self, filters: dict) -> str:
    """
    Xizmatlar ro'yxatini Excel ga eksport qiladi.
    Qaytaradi: fayl nomi (MEDIA_ROOT/temp_exports/ papkasida).
    """
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from apps.services.models import PatientService

    qs = PatientService.objects.exclude(status='cancelled').values(
        'ordered_at',
        'patient_card__full_name',
        'patient_category_at_order',
        'service__category__name',
        'service__name',
        'quantity',
        'price',
        'status',
        'is_paid',
        'ordered_by__first_name',
        'ordered_by__last_name',
        'performed_by__first_name',
        'performed_by__last_name',
        'result',
    )

    if filters.get('date_from'):
        qs = qs.filter(ordered_at__date__gte=filters['date_from'])
    if filters.get('date_to'):
        qs = qs.filter(ordered_at__date__lte=filters['date_to'])
    if filters.get('category'):
        qs = qs.filter(service__category_id=filters['category'])
    if filters.get('patient_category'):
        qs = qs.filter(patient_category_at_order=filters['patient_category'])
    if filters.get('visit_type'):
        qs = qs.filter(patient_card__visit_type=filters['visit_type'])

    qs = qs.order_by('-ordered_at')

    # pandas orqali yuklab olish (ORM iterator'dan 3-5x tez)
    df = pd.DataFrame.from_records(qs.iterator(chunk_size=2000))

    if df.empty:
        # Bo'sh fayl
        df = pd.DataFrame(columns=[
            '№', 'Sana', 'Bemor', 'Bemor kategoriyasi', 'Kategoriya',
            'Xizmat', 'Miqdori', 'Narx', 'Jami', 'Holat', "To'langan",
            'Buyurtma bergan', 'Bajargan', 'Natija',
        ])
    else:
        cat_display = {
            'railway': "Temir yo'lchi", 'paid': 'Pullik',
            'non_resident': 'Norezident', 'foreign': 'Chet el',
        }
        status_display = {
            'ordered': 'Buyurtma berildi', 'completed': 'Bajarildi',
            'cancelled': 'Bekor qilindi', 'in_progress': 'Jarayonda',
        }

        df['№'] = range(1, len(df) + 1)
        df['Sana'] = df['ordered_at'].dt.tz_convert('Asia/Tashkent').dt.strftime('%d.%m.%Y %H:%M')
        df['Bemor'] = df['patient_card__full_name']
        df['Bemor kategoriyasi'] = df['patient_category_at_order'].map(cat_display).fillna(df['patient_category_at_order'])
        df['Kategoriya'] = df['service__category__name']
        df['Xizmat'] = df['service__name']
        df['Miqdori'] = df['quantity']
        df['Narx'] = df['price'].astype(float)
        df['Jami'] = (df['price'] * df['quantity']).astype(float)
        df['Holat'] = df['status'].map(status_display).fillna(df['status'])
        df["To'langan"] = df['is_paid'].map({True: 'Ha', False: "Yo'q"})
        df['Buyurtma bergan'] = (
            df['ordered_by__first_name'].fillna('') + ' ' +
            df['ordered_by__last_name'].fillna('')
        ).str.strip().replace('', '—')
        df['Bajargan'] = (
            df['performed_by__first_name'].fillna('') + ' ' +
            df['performed_by__last_name'].fillna('')
        ).str.strip().replace('', '—')
        df['Natija'] = df['result'].fillna('—')

        df = df[['№', 'Sana', 'Bemor', 'Bemor kategoriyasi', 'Kategoriya',
                 'Xizmat', 'Miqdori', 'Narx', 'Jami', 'Holat', "To'langan",
                 'Buyurtma bergan', 'Bajargan', 'Natija']]

    filename = f'services_{uuid.uuid4().hex[:12]}.xlsx'
    filepath = os.path.join(_export_dir(), filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Xizmatlar ro'yxati", index=False)
        ws = writer.sheets["Xizmatlar ro'yxati"]
        # Ustun kengliklari
        for col_idx, width in enumerate([4, 16, 25, 16, 18, 30, 8, 12, 12, 14, 10, 22, 22, 30], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Sarlavha formatlash
        hf = Font(bold=True, color='FFFFFF', size=10)
        hfill = PatternFill('solid', fgColor='1F4E79')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        brd = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.font = hf; cell.fill = hfill
            cell.alignment = center; cell.border = brd

    logger.info(f"Xizmatlar Excel yaratildi: {filename}, qatorlar: {len(df)}")
    return filename


# ──────────────────────────────────────────────────────────────────────────────
# Dori-darmonlar hisoboti
# ──────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, max_retries=0, time_limit=600, soft_time_limit=560,
             name='services.export_medicine_excel')
def generate_medicine_excel(self, filters: dict) -> str:
    """Dori-darmonlar ro'yxatini Excel ga eksport qiladi."""
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from apps.services.models import PatientMedicine

    qs = PatientMedicine.objects.values(
        'ordered_at',
        'patient_card__full_name',
        'medicine__name',
        'medicine__unit',
        'quantity',
        'price',
        'ordered_by__first_name',
        'ordered_by__last_name',
    )
    if filters.get('date_from'):
        qs = qs.filter(ordered_at__date__gte=filters['date_from'])
    if filters.get('date_to'):
        qs = qs.filter(ordered_at__date__lte=filters['date_to'])
    if filters.get('medicine'):
        qs = qs.filter(medicine_id=filters['medicine'])
    if filters.get('patient_category'):
        qs = qs.filter(patient_card__patient_category=filters['patient_category'])
    if filters.get('visit_type'):
        qs = qs.filter(patient_card__visit_type=filters['visit_type'])

    qs = qs.order_by('-ordered_at')
    df = pd.DataFrame.from_records(qs.iterator(chunk_size=2000))

    if df.empty:
        df = pd.DataFrame(columns=['№', 'Sana', 'Bemor', 'Dori nomi', 'Birlik', 'Miqdori', 'Narxi', 'Jami', 'Buyurtma bergan'])
    else:
        df['№'] = range(1, len(df) + 1)
        df['Sana'] = df['ordered_at'].dt.tz_convert('Asia/Tashkent').dt.strftime('%d.%m.%Y')
        df['Bemor'] = df['patient_card__full_name']
        df['Dori nomi'] = df['medicine__name']
        df['Birlik'] = df['medicine__unit']
        df['Miqdori'] = df['quantity'].astype(float)
        df['Narxi'] = df['price'].astype(float)
        df['Jami'] = (df['quantity'] * df['price']).astype(float)
        df['Buyurtma bergan'] = (
            df['ordered_by__first_name'].fillna('') + ' ' +
            df['ordered_by__last_name'].fillna('')
        ).str.strip().replace('', '—')
        df = df[['№', 'Sana', 'Bemor', 'Dori nomi', 'Birlik', 'Miqdori', 'Narxi', 'Jami', 'Buyurtma bergan']]

    filename = f'medicine_{uuid.uuid4().hex[:12]}.xlsx'
    filepath = os.path.join(_export_dir(), filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dori-darmonlar', index=False)
        ws = writer.sheets['Dori-darmonlar']
        for col_idx, width in enumerate([4, 14, 28, 28, 10, 12, 14, 14, 22], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        hf = Font(bold=True, color='FFFFFF', size=10)
        hfill = PatternFill('solid', fgColor='856404')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.font = hf; cell.fill = hfill; cell.alignment = center

    logger.info(f"Dori Excel yaratildi: {filename}, qatorlar: {len(df)}")
    return filename


# ──────────────────────────────────────────────────────────────────────────────
# Operatsiyalar statistikasi
# ──────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, max_retries=0, time_limit=600, soft_time_limit=560,
             name='services.export_operations_excel')
def generate_operations_excel(self, filters: dict) -> str:
    """Operatsiya statistikasi Excel faylini Celery'da yaratadi."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Count, Q
    from apps.patients.models import SurgicalOperation

    date_from  = filters.get('date_from', '')
    date_to    = filters.get('date_to', '')
    op_type_id = filters.get('op_type', '')
    anesthesia = filters.get('anesthesia', '')

    qs = SurgicalOperation.objects.select_related(
        'operation_type', 'patient_card'
    ).filter(operation_type__isnull=False)

    if date_from:    qs = qs.filter(operation_date__gte=date_from)
    if date_to:      qs = qs.filter(operation_date__lte=date_to)
    if op_type_id:   qs = qs.filter(operation_type_id=op_type_id)
    if anesthesia:   qs = qs.filter(anesthesia=anesthesia)

    # Stillar — bir marta yaratilib, har bir katakka qayta ishlatiladi
    BLUE  = PatternFill('solid', fgColor='1F4E79')
    RED   = PatternFill('solid', fgColor='FDEDEC')
    ZEBRA = PatternFill('solid', fgColor='F0F6FC')

    WF   = Font(color='FFFFFF', bold=True, size=10)
    BOLD = Font(bold=True, size=10)
    NORM = Font(size=9)
    C    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    brd  = Side(style='thin', color='CCCCCC')
    BRD  = Border(left=brd, right=brd, top=brd, bottom=brd)

    wb = openpyxl.Workbook()

    # ===== SHEET 1: OPERATSIYA TURI BO'YICHA =====
    ws1 = wb.active
    ws1.title = "Operatsiya turlari"

    col_w = [5, 12, 35, 12, 12, 14, 16, 14, 14, 12, 12]
    for ci, w in enumerate(col_w, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.merge_cells('A1:K1')
    c = ws1.cell(row=1, column=1, value="OPERATSIYA TURLARI BO'YICHA STATISTIKA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = C
    ws1.row_dimensions[1].height = 30

    if date_from or date_to:
        ws1.merge_cells('A2:K2')
        c = ws1.cell(row=2, column=1,
            value=f"Davr: {date_from or '—'} dan {date_to or '—'} gacha")
        c.font = BOLD; c.alignment = C
        ws1.row_dimensions[2].height = 18
        hdr_row = 3
    else:
        hdr_row = 2

    headers = [
        '№', 'Kod', 'Operatsiya nomi',
        'Jami\nsoni',
        "Temir\nyo'lchi",
        'Pullik',
        'Norezident',
        "Narkoz\nbilan",
        'Mahalliy\nnarkoz',
        'Narkozsiz',
        'Asorat\nbor',
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=hdr_row, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws1.row_dimensions[hdr_row].height = 36

    op_stats = (
        qs.values(
            'operation_type__id',
            'operation_type__code',
            'operation_type__name',
        )
        .annotate(
            total_count=Count('id'),
            railway=Count('id', filter=Q(patient_card__patient_category='railway')),
            paid=Count('id', filter=Q(patient_card__patient_category='paid')),
            nonresident=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            anes_yes=Count('id', filter=Q(anesthesia='yes')),
            anes_local=Count('id', filter=Q(anesthesia='local')),
            anes_no=Count('id', filter=Q(anesthesia='no')),
            has_complication=Count('id', filter=~Q(complication='') & Q(complication__isnull=False)),
        )
        .order_by('-total_count')
    )

    dr = hdr_row + 1
    for ri, op in enumerate(op_stats, 1):
        vals = [
            ri,
            op['operation_type__code'] or '—',
            op['operation_type__name'],
            op['total_count'],
            op['railway'],
            op['paid'],
            op['nonresident'],
            op['anes_yes'],
            op['anes_local'],
            op['anes_no'],
            op['has_complication'],
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=dr, column=col, value=val)
            c.font = NORM; c.border = BRD
            c.alignment = C if col != 3 else L
            if ri % 2 == 0: c.fill = ZEBRA
        ws1.row_dimensions[dr].height = 18
        dr += 1

    # Jami
    tot_vals = ['', '', 'JAMI:'] + [qs.aggregate(t=Count('id'))['t']] + [
        qs.filter(patient_card__patient_category='railway').count(),
        qs.filter(patient_card__patient_category='paid').count(),
        qs.filter(patient_card__patient_category='non_resident').count(),
        qs.filter(anesthesia='yes').count(),
        qs.filter(anesthesia='local').count(),
        qs.filter(anesthesia='no').count(),
        qs.exclude(complication='').exclude(complication__isnull=True).count(),
    ]
    for col, val in enumerate(tot_vals, 1):
        c = ws1.cell(row=dr, column=col, value=val)
        c.fill = BLUE; c.font = WF; c.alignment = C if col != 3 else L; c.border = BRD
    ws1.row_dimensions[dr].height = 22

    # ===== SHEET 2: BATAFSIL RO'YXAT =====
    ws2 = wb.create_sheet("Batafsil ro'yxat")
    h2_widths = [5, 14, 28, 14, 35, 20, 16, 14, 30]
    for ci, w in enumerate(h2_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells('A1:I1')
    c = ws2.cell(row=1, column=1, value="OPERATSIYALAR BATAFSIL RO'YXATI")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws2.row_dimensions[1].height = 26

    h2 = ['№', 'Sana', 'Bemor', 'Bayonnoma', 'Operatsiya turi', 'Bemor turi', 'Narkoz', "Asorat bor?", 'Asorat tavsifi']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws2.row_dimensions[2].height = 22

    cat_display_map = {
        'railway': "Temir yo'lchi",
        'paid': 'Pullik',
        'non_resident': 'Norezident',
    }
    for ri, op in enumerate(qs.order_by('-operation_date'), 1):
        cat_display = cat_display_map.get(op.patient_card.patient_category, '—')

        has_comp = bool(op.complication and op.complication.strip())
        data = [
            ri,
            op.operation_date.strftime('%d.%m.%Y') if op.operation_date else '—',
            op.patient_card.full_name,
            op.patient_card.medical_record_number,
            str(op.operation_type) if op.operation_type else op.operation_name or '—',
            cat_display,
            op.get_anesthesia_display() if op.anesthesia else '—',
            'Ha' if has_comp else "Yo'q",
            op.complication or '—',
        ]
        for col, val in enumerate(data, 1):
            c = ws2.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD
            c.alignment = C if col in (1,2,4,6,7,8) else L
            if col == 8 and has_comp:
                c.fill = RED
            elif ri % 2 == 0:
                c.fill = ZEBRA
        ws2.row_dimensions[ri+2].height = 17

    filename = f'operations_{uuid.uuid4().hex[:12]}.xlsx'
    filepath = os.path.join(_export_dir(), filename)
    wb.save(filepath)

    logger.info(f"Operatsiyalar Excel yaratildi: {filename}")
    return filename


# ──────────────────────────────────────────────────────────────────────────────
# Operatsiya xizmatlari statistikasi
# ──────────────────────────────────────────────────────────────────────────────

@shared_task(bind=True, max_retries=0, time_limit=600, soft_time_limit=560,
             name='services.export_operation_services_excel')
def generate_operation_services_excel(self, filters: dict) -> str:
    """Operatsiya xizmatlari statistikasi Excel faylini Celery'da yaratadi."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count, Q, ExpressionWrapper, DecimalField, F
    from apps.services.models import PatientService

    date_from   = filters.get('date_from', '')
    date_to     = filters.get('date_to', '')
    category_id = filters.get('category', '')
    service_id  = filters.get('service', '')
    patient_cat = filters.get('patient_category', '')
    visit_type  = filters.get('visit_type', '')

    qs = PatientService.objects.filter(
        service__is_operation=True
    ).select_related('service__category', 'patient_card')
    if date_from:   qs = qs.filter(ordered_at__date__gte=date_from)
    if date_to:     qs = qs.filter(ordered_at__date__lte=date_to)
    if category_id: qs = qs.filter(service__category_id=category_id)
    if service_id:  qs = qs.filter(service_id=service_id)
    if patient_cat: qs = qs.filter(patient_card__patient_category=patient_cat)
    if visit_type:  qs = qs.filter(patient_card__visit_type=visit_type)

    # Stillar — bir marta yaratilib, har bir katakka qayta ishlatiladi
    BLUE  = PatternFill('solid', fgColor='1F4E79')
    LBLUE = PatternFill('solid', fgColor='D6E4F0')
    TOTAL = PatternFill('solid', fgColor='145A32')
    WHITE = PatternFill('solid', fgColor='FFFFFF')
    WF    = Font(color='FFFFFF', bold=True, size=10)
    BOLD  = Font(bold=True, size=10)
    NORM  = Font(size=9)
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    R = Alignment(horizontal='right',  vertical='center')
    brd = Side(style='thin', color='CCCCCC')
    BRD = Border(left=brd, right=brd, top=brd, bottom=brd)

    wb = openpyxl.Workbook()

    # ===== SHEET 1: OPERATSIYA XIZMATLARI =====
    ws1 = wb.active
    ws1.title = "Operatsiya xizmatlari"

    col_widths = [5, 12, 35, 18, 10, 10, 10, 10, 14, 14, 14, 16]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.merge_cells('A1:L1')
    c = ws1.cell(row=1, column=1,
        value="OPERATSIYA XIZMATLARI BO'YICHA STATISTIKA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = C
    ws1.row_dimensions[1].height = 30

    if date_from or date_to:
        ws1.merge_cells('A2:L2')
        ws1.cell(row=2, column=1,
            value=f"Davr: {date_from or '—'} dan {date_to or '—'} gacha").font = BOLD
        ws1.row_dimensions[2].height = 18
        hdr = 3
    else:
        hdr = 2

    headers = [
        '№', 'Kod', 'Operatsiya nomi', 'Kategoriya',
        'Jami\nsoni', 'Bemorlar',
        "TY\nsoni", "Pullik\nsoni", "Nores.\nsoni",
        "TY\nsumma", "Pullik\nsumma", "Nores.\nsumma",
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=hdr, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws1.row_dimensions[hdr].height = 36

    op_stats = (
        qs.values(
            'service__code', 'service__name', 'service__category__name'
        )
        .annotate(
            tc=Count('id'),
            pts=Count('patient_card', distinct=True),
            rc=Count('id', filter=Q(patient_card__patient_category='railway')),
            pc=Count('id', filter=Q(patient_card__patient_category='paid')),
            nc=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            rs=Sum('price', filter=Q(patient_card__patient_category='railway')),
            ps=Sum('price', filter=Q(patient_card__patient_category='paid')),
            ns=Sum('price', filter=Q(patient_card__patient_category='non_resident')),
        )
        .order_by('service__category__name', '-tc')
    )

    dr = hdr + 1
    grand = {'tc':0,'pts':0,'rc':0,'pc':0,'nc':0,'rs':0,'ps':0,'ns':0}

    for ri, op in enumerate(op_stats, 1):
        vals = [
            ri,
            op['service__code'] or '—',
            op['service__name'],
            op['service__category__name'],
            op['tc'], op['pts'],
            op['rc'], op['pc'], op['nc'],
            float(op['rs'] or 0), float(op['ps'] or 0), float(op['ns'] or 0),
        ]
        for k in ('tc','pts','rc','pc','nc'):
            grand[k] += op[k] or 0
        for k, col in (('rs',10),('ps',11),('ns',12)):
            grand[k] += float(op[k] or 0)

        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=dr, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,5,6,7,8,9) else (R if col > 9 else L)
            if col > 9: c.number_format = '#,##0'
        ws1.row_dimensions[dr].height = 18
        dr += 1

    # Jami qatori
    tot = ['', '', 'JAMI:', '',
           grand['tc'], grand['pts'],
           grand['rc'], grand['pc'], grand['nc'],
           grand['rs'], grand['ps'], grand['ns']]
    for col, val in enumerate(tot, 1):
        c = ws1.cell(row=dr, column=col, value=val)
        c.fill = BLUE; c.font = WF; c.border = BRD
        c.alignment = C if col in (1,5,6,7,8,9) else (R if col > 9 else L)
        if col > 9: c.number_format = '#,##0'
    ws1.row_dimensions[dr].height = 24

    # Umumiy summa qatori
    dr += 1
    total_sum = grand['rs'] + grand['ps'] + grand['ns']
    ws1.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=9)
    c = ws1.cell(row=dr, column=1, value="UMUMIY DAROMAD:")
    c.fill = TOTAL; c.font = WF; c.alignment = L; c.border = BRD
    c2 = ws1.cell(row=dr, column=10, value=total_sum)
    c2.fill = TOTAL; c2.font = Font(color='FFFFFF', bold=True, size=12)
    c2.number_format = '#,##0'; c2.alignment = R; c2.border = BRD
    ws1.merge_cells(start_row=dr, start_column=10, end_row=dr, end_column=12)
    ws1.row_dimensions[dr].height = 26

    # ===== SHEET 2: KATEGORIYA BO'YICHA =====
    ws2 = wb.create_sheet("Kategoriya bo'yicha")
    for ci, w in enumerate([5, 25, 10, 12, 14, 14, 14, 16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.merge_cells('A1:H1')
    c = ws2.cell(row=1, column=1, value="KATEGORIYA BO'YICHA")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws2.row_dimensions[1].height = 26

    h2 = ['№', 'Kategoriya', 'Jami soni', 'Bemorlar',
          "TY soni", "Pullik soni", "Nores. soni", "Jami summa"]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws2.row_dimensions[2].height = 24

    cat_stats = (
        qs.values('service__category__name', 'service__category__icon')
        .annotate(
            cnt=Count('id'), pts=Count('patient_card', distinct=True),
            rc=Count('id', filter=Q(patient_card__patient_category='railway')),
            pc=Count('id', filter=Q(patient_card__patient_category='paid')),
            nc=Count('id', filter=Q(patient_card__patient_category='non_resident')),
            total=Sum(ExpressionWrapper(F('price') * F('quantity'), output_field=DecimalField())),
        )
        .order_by('-total')
    )
    cat_stats = list(cat_stats)

    cat_grand = 0
    for ri, cat in enumerate(cat_stats, 1):
        icon = cat['service__category__icon'] or ''
        name = f"{icon} {cat['service__category__name']}"
        vals = [ri, name, cat['cnt'], cat['pts'],
                cat['rc'], cat['pc'], cat['nc'],
                float(cat['total'] or 0)]
        cat_grand += float(cat['total'] or 0)
        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,3,4,5,6,7) else (R if col == 8 else L)
            if col == 8: c.number_format = '#,##0'
        ws2.row_dimensions[ri+2].height = 20

    last2 = len(cat_stats) + 3
    ws2.merge_cells(start_row=last2, start_column=1, end_row=last2, end_column=7)
    c = ws2.cell(row=last2, column=1, value="JAMI:")
    c.fill = BLUE; c.font = WF; c.alignment = L; c.border = BRD
    c8 = ws2.cell(row=last2, column=8, value=cat_grand)
    c8.fill = BLUE; c8.font = WF; c8.number_format = '#,##0'
    c8.alignment = R; c8.border = BRD
    ws2.row_dimensions[last2].height = 22

    # ===== SHEET 3: BATAFSIL =====
    ws3 = wb.create_sheet("Batafsil")
    for ci, w in enumerate([5, 14, 28, 28, 16, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws3.merge_cells('A1:G1')
    c = ws3.cell(row=1, column=1, value="BATAFSIL RO'YXAT")
    c.fill = BLUE; c.font = Font(color='FFFFFF', bold=True, size=12); c.alignment = C
    ws3.row_dimensions[1].height = 26

    h3 = ['№', 'Sana', 'Bemor', 'Operatsiya nomi', 'Bemor turi', "Narx (so'm)", "Jami (so'm)"]
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.fill = BLUE; c.font = WF; c.alignment = C; c.border = BRD
    ws3.row_dimensions[2].height = 22

    CAT_MAP = {'railway': "Temir yo'lchi", 'paid': 'Pullik', 'non_resident': 'Norezident'}
    for ri, ps in enumerate(qs.order_by('-ordered_at'), 1):
        data = [
            ri,
            ps.ordered_at.strftime('%d.%m.%Y'),
            ps.patient_card.full_name,
            ps.service.name,
            CAT_MAP.get(ps.patient_card.patient_category, '—'),
            float(ps.price),
            float(ps.total_price),
        ]
        fill = LBLUE if ri % 2 == 0 else WHITE
        for col, val in enumerate(data, 1):
            c = ws3.cell(row=ri+2, column=col, value=val)
            c.font = NORM; c.border = BRD; c.fill = fill
            c.alignment = C if col in (1,2,5) else (R if col > 5 else L)
            if col > 5: c.number_format = '#,##0'
        ws3.row_dimensions[ri+2].height = 17

    filename = f'opservices_{uuid.uuid4().hex[:12]}.xlsx'
    filepath = os.path.join(_export_dir(), filename)
    wb.save(filepath)

    logger.info(f"Operatsiya xizmatlari Excel yaratildi: {filename}")
    return filename
