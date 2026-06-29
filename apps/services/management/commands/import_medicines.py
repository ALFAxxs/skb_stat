# apps/services/management/commands/import_medicines.py

from django.core.management.base import BaseCommand
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Dori-darmonlarni Excel fayldan import qilish"

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help="Excel fayl yo'li")
        parser.add_argument(
            '--sheet', type=str, default=None,
            help="Sheet nomi (belgilanmasa — birinchi sheet)"
        )
        parser.add_argument(
            '--skip-header', action='store_true', default=True,
            help="Birinchi qatorni sarlavha sifatida o'tkazib yuborish"
        )

    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument(
            '--col-mnn', type=int, default=None,
            help="МНН ustun indeksi (0 dan boshlab). Belgilanmasa bo'sh qoladi."
        )
        parser.add_argument(
            '--col-form', type=int, default=None,
            help="Chiqarish shakli ustun indeksi."
        )
        parser.add_argument(
            '--col-strength', type=int, default=None,
            help="Doza/konsentratsiya ustun indeksi."
        )
        parser.add_argument(
            '--col-category', type=int, default=None,
            help="Kategoriya ustun indeksi (drug/lab/other)."
        )
        parser.add_argument(
            '--default-category', type=str, default='drug',
            choices=['drug', 'lab', 'other'],
            help="Ustun ko'rsatilmasa ishlatiladi (default: drug)."
        )
        parser.add_argument(
            '--update', action='store_true', default=False,
            help="Mavjud yozuvlarni yangilash (МНН, forma, doza)."
        )

    def handle(self, *args, **options):
        from apps.services.models import Medicine

        filepath    = options['filepath']
        sheet_name  = options.get('sheet')
        col_mnn     = options['col_mnn']
        col_form    = options['col_form']
        col_str     = options['col_strength']
        col_cat     = options['col_category']
        default_cat = options['default_category']
        do_update   = options['update']

        try:
            wb = load_workbook(filepath, read_only=True)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Fayl ochilmadi: {e}"))
            return

        ws = wb[sheet_name] if sheet_name else wb.active
        self.stdout.write(f"Sheet: '{ws.title}'")

        UNIT_MAP = {
            'qadoq': 'dona', 'пар': 'dona', 'рул': 'dona',
            'к-т': 'dona', 'К-т': 'dona', 'тюб': 'tuba',
            'flakon': 'shisha', 'flakon ': 'shisha',
            'доз/ampula': 'ampula', 'канис': 'shisha',
            'литр': 'l', 'kub.metr': 'l', 'kg': 'g',
            "nabor/to'plam": 'dona', ' ': 'dona', '': 'dona',
        }
        ALLOWED_UNITS = {
            'dona', 'ml', 'mg', 'g', 'l', 'ampula',
            'kapsula', 'tabletka', 'paket', 'shisha', 'tuba',
        }

        def cell(row, idx):
            if idx is None or idx >= len(row):
                return ''
            return str(row[idx]).strip() if row[idx] else ''

        created = updated = skipped = errors = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            name = cell(row, 0)
            if not name:
                skipped += 1
                continue

            unit = cell(row, 1) or 'dona'
            unit = UNIT_MAP.get(unit, unit)
            if unit not in ALLOWED_UNITS:
                unit = 'dona'

            mnn      = cell(row, col_mnn)      if col_mnn     is not None else ''
            form_raw = cell(row, col_form)     if col_form    is not None else ''
            strength = cell(row, col_str)      if col_str     is not None else ''
            cat_raw  = cell(row, col_cat)      if col_cat     is not None else default_cat

            # Форма выпуска — qiymat choices ichidami tekshir
            valid_forms = {c[0] for c in Medicine.DOSAGE_FORM_CHOICES}
            dosage_form = form_raw if form_raw in valid_forms else ''

            cat_raw = cat_raw.lower()
            valid_cats = {c[0] for c in Medicine.CATEGORY_CHOICES}
            category = cat_raw if cat_raw in valid_cats else default_cat

            try:
                obj, is_new = Medicine.objects.get_or_create(
                    name=name,
                    defaults={
                        'unit': unit, 'mnn': mnn, 'dosage_form': dosage_form,
                        'strength': strength, 'category': category, 'is_active': True,
                    }
                )
                if is_new:
                    created += 1
                    if created <= 5:
                        self.stdout.write(self.style.SUCCESS(f"  + {name}"))
                elif do_update:
                    obj.unit = unit
                    if mnn:      obj.mnn = mnn
                    if dosage_form: obj.dosage_form = dosage_form
                    if strength: obj.strength = strength
                    obj.category = category
                    obj.save()
                    updated += 1
                else:
                    skipped += 1

            except Exception as e:
                errors += 1
                self.stderr.write(f"  Qator {row_num} xato: {e}")

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"✅ Yakunlandi: {created} ta qo'shildi, {updated} ta yangilandi, "
            f"{skipped} ta o'tkazildi, {errors} ta xato"
        ))