from django.views.decorators.http import require_POST
# apps/patients/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext as _
from django.utils.translation import gettext_lazy as _l
from apps.users.decorators import role_required, department_filter
import json
import uuid

from django.utils import timezone

from datetime import datetime, timedelta

from .forms import PatientCardForm, DeathCauseForm, SurgicalOperationFormSet, ReceptionForm
from .models import (
    PatientCard, ICD10Code, DischargeConclusion,
    Region, District, City, Village, Country, OperationType,
    Department, InitialExamination, EpisodeDiagnosis,
    MedicalExamination, DoctorNotification,
    TreatmentProcedure, ProcedureExecutionLog,
    LabTestAssignment, LabTestResultLog,
    DiagnosticAssignment, DiagnosticResultLog,
    ConsultationRequest, ConsultationResponse,
    AmbulatoryConsultation, DoctorTextTemplate,
    ServiceSchedule, Prescription,
)
from apps.users.models import CustomUser


def _notify_doctor(doctor, patient, message):
    """Shifokor (CustomUser)ga bildirishnoma yuboradi."""
    if not doctor:
        return
    DoctorNotification.objects.create(
        recipient=doctor,
        patient_card=patient,
        message=message,
    )


def _notify_user(user, patient, message):
    """Istalgan tizim foydalanuvchisiga (masalan hamshiraga) bildirishnoma yuboradi."""
    if not user:
        return
    DoctorNotification.objects.create(recipient=user, patient_card=patient, message=message)


def _notify_new_admission(patient):
    """Yangi statsionar bemor qabul qilinganda — davolovchi shifokor va bo'lim mudiriga xabar."""
    if patient.visit_type != 'inpatient':
        return
    if patient.attending_doctor_id:
        _notify_doctor(
            patient.attending_doctor, patient,
            f"Sizga yangi bemor biriktirildi: {patient.full_name} ({patient.medical_record_number})"
        )
    head = patient.department_head or (
        CustomUser.objects.filter(department=patient.department, role__in=('doctor', 'old'), is_head=True, is_active=True).first()
        if patient.department_id else None
    )
    if head and (not patient.attending_doctor_id or head.pk != patient.attending_doctor_id):
        _notify_doctor(
            head, patient,
            f"Bo'limingizga yangi bemor yotqizildi: {patient.full_name} ({patient.medical_record_number})"
        )


def _find_recent_duplicate(full_name, birth_date, seconds=120):
    """Bir xil ism+tug'ilgan sana bilan so'nggi N soniyada yaratilgan bemor."""
    if not full_name or not birth_date:
        return None
    cutoff = timezone.now() - timedelta(seconds=seconds)
    return PatientCard.objects.filter(
        full_name__iexact=full_name,
        birth_date=birth_date,
        created_at__gte=cutoff,
    ).order_by('-created_at').first()


# ==================== AJAX VIEWS ====================

@login_required
def add_conclusion(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            if not name:
                return JsonResponse({'success': False, 'error': _("Nom bo'sh")})
            obj, created = DischargeConclusion.objects.get_or_create(name=name)
            return JsonResponse({
                'success': True,
                'id': obj.id,
                'name': obj.name,
                'created': created
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': _('Faqat POST')})


def get_regions(request):
    country_id = request.GET.get('country_id')
    regions = Region.objects.filter(country_id=country_id).values('id', 'name')
    return JsonResponse(list(regions), safe=False)


def get_districts(request):
    region_id = request.GET.get('region_id')
    districts = District.objects.filter(region_id=region_id).values('id', 'name')
    return JsonResponse(list(districts), safe=False)


def get_cities(request):
    district_id = request.GET.get('district_id')
    cities = City.objects.filter(district_id=district_id).values('id', 'name')
    return JsonResponse(list(cities), safe=False)


def get_villages(request):
    district_id = request.GET.get('district_id')
    villages = Village.objects.filter(district_id=district_id).values('id', 'name')
    return JsonResponse(list(villages), safe=False)


def icd10_search(request):
    q = request.GET.get('q', '')
    if len(q) < 2:
        return JsonResponse([], safe=False)
    results = ICD10Code.objects.filter(
        Q(code__icontains=q) | Q(title_uz__icontains=q) | Q(title_ru__icontains=q)
    )[:20]
    data = [{'code': r.code, 'title_uz': r.title_uz, 'title_ru': r.title_ru} for r in results]
    return JsonResponse(data, safe=False)


@login_required
def episode_diagnoses(request, patient_id):
    patient = get_object_or_404(PatientCard, pk=patient_id)

    if request.method == 'GET':
        diags = patient.episode_diagnoses.select_related('icd10_code').all()
        data = [{
            'id':             d.id,
            'icd10_code':     d.icd10_code.code if d.icd10_code else '',
            'icd10_title':    d.icd10_code.title_uz if d.icd10_code else '',
            'diagnosis_type': d.diagnosis_type,
            'diagnosis_type_display': d.get_diagnosis_type_display(),
            'diagnosis_role': d.diagnosis_role,
            'diagnosis_role_display': d.get_diagnosis_role_display(),
            'clinical_text':  d.clinical_text,
            'sort_order':     d.sort_order,
        } for d in diags]
        return JsonResponse({'diagnoses': data})

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            icd10_id   = body.get('icd10_code')
            diag_type  = body.get('diagnosis_type', 'preliminary')
            diag_role  = body.get('diagnosis_role', 'main')
            clin_text  = body.get('clinical_text', '')
            sort_order = int(body.get('sort_order', 0))

            icd = ICD10Code.objects.filter(code=icd10_id).first() if icd10_id else None
            d = EpisodeDiagnosis.objects.create(
                patient_card=patient,
                icd10_code=icd,
                diagnosis_type=diag_type,
                diagnosis_role=diag_role,
                disease_course=body.get('disease_course', ''),
                clinical_text=clin_text,
                sort_order=sort_order,
            )
            return JsonResponse({'success': True, 'id': d.id,
                                 'icd10_code': d.icd10_code.code if d.icd10_code else '',
                                 'icd10_title': d.icd10_code.title_uz if d.icd10_code else '',
                                 'diagnosis_type_display': d.get_diagnosis_type_display(),
                                 'diagnosis_role_display': d.get_diagnosis_role_display(),
                                 'clinical_text': d.clinical_text})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'error': _('Faqat GET/POST')}, status=405)


@login_required
@require_POST
def episode_diagnosis_delete(request, pk):
    d = get_object_or_404(EpisodeDiagnosis, pk=pk)
    d.delete()
    return JsonResponse({'success': True})


@login_required
def initial_examination(request, patient_id):
    patient = get_object_or_404(PatientCard, pk=patient_id)
    exam, _ = InitialExamination.objects.get_or_create(patient_card=patient)

    if request.method == 'GET':
        return JsonResponse({
            'complaints':                exam.complaints,
            'anamnesis_morbi':           exam.anamnesis_morbi,
            'anamnesis_vitae':           exam.anamnesis_vitae,
            'status_localis':            exam.status_localis,
            'epidemiological_anamnesis': exam.epidemiological_anamnesis,
            'status_praesens':           exam.status_praesens,
            'allergy_anamnesis':         exam.allergy_anamnesis,
            'neurological_status':       exam.neurological_status,
            'lab_investigations':        exam.lab_investigations,
        })

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            fields = [
                'complaints', 'anamnesis_morbi', 'anamnesis_vitae',
                'status_localis', 'epidemiological_anamnesis', 'status_praesens',
                'allergy_anamnesis', 'neurological_status', 'lab_investigations',
            ]
            for f in fields:
                if f in body:
                    setattr(exam, f, body[f])
            exam.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'error': 'Faqat GET/POST'}, status=405)


# ==================== TIBBIY KO'RIK VIEWLARI ====================

EXAM_FIELDS = [
    ('complaints',               _l("Ko'rik jarayonidagi shikoyatlar")),
    ('anamnesis_morbi',          _l('Kasallik tarixidan (Anamnesis morbi)')),
    ('anamnesis_vitae',          _l('Hayot anamnezi (Anamnesis vitae)')),
    ('status_praesens',          _l("Ob'ektiv holat (Status praesens)")),
    ('neurological_status',      _l('Nevrologik holat')),
    ('status_localis',           _l('Mahalliy holat (Status localis)')),
    ('epidemiological_anamnesis',_l('Epidemiologik anamnez')),
    ('lab_investigations',       _l('Laboratoriya va instrumental tadqiqotlar va maslahatlar')),
    ('specialist_consultations', _l('Turdosh mutaxassislar maslahatlari')),
    ('allergy_anamnesis',        _l('Allergoanamnesis')),
    ('conclusion',               _l('Tavsiyalar')),
    ('drug_justification',       _l('Dori vositalari uchun asoslar')),
]

# Epikriz turlarida yo'q bo'lgan maydonlar
EPICRISIS_HIDDEN = {
    'stage_epicrisis': {'epidemiological_anamnesis', 'anamnesis_vitae', 'neurological_status'},
    'discharge':       {'epidemiological_anamnesis', 'anamnesis_vitae', 'neurological_status'},
    'clinical_basis':  set(),
}
# Ko'rik turlari uchun drug_justification ko'rinmaydi
CHECKUP_TYPES = {'initial', 'ward', 'daily', 'specialist', 'consilium', 'anesthesia'}


def _safe_next_url(request, fallback_url):
    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url
    return fallback_url


@login_required
def medical_examination_page(request, patient_id, exam_pk=None):
    """To'liq sahifa — ko'rik yozish/tahrirlash."""
    patient = get_object_or_404(PatientCard, pk=patient_id)
    exam    = get_object_or_404(MedicalExamination, pk=exam_pk, patient_card=patient) if exam_pk else None
    valid_exam_types = dict(MedicalExamination.EXAM_TYPE_CHOICES)
    fallback_url = reverse('patient_detail', args=[patient_id])
    next_url = _safe_next_url(request, fallback_url)

    if request.method == 'POST':
        exam_type = request.POST.get('examination_type') or (exam.examination_type if exam else None)
        if exam_type not in valid_exam_types:
            messages.error(request, _("Ko'rik turi tanlanmagan yoki noto'g'ri."))
            return redirect(next_url)
        exam_dt   = request.POST.get('examination_datetime') or None

        if not exam:
            exam = MedicalExamination(patient_card=patient, examination_type=exam_type)

        from django.utils.dateparse import parse_datetime
        exam.examination_datetime  = parse_datetime(exam_dt) if exam_dt else None
        exam.created_by            = request.user
        exam.department_head_name  = request.POST.get('department_head_name', '')
        for fname, _label in EXAM_FIELDS:
            if fname == 'lab_investigations':
                continue
            setattr(exam, fname, request.POST.get(fname, ''))

        # Laboratoriya/diagnostika natijalarini checkbox orqali tanlash
        selected_lab_ids       = request.POST.getlist('selected_lab_ids')
        selected_diag_ids      = request.POST.getlist('selected_diag_ids')
        selected_labresult_ids = request.POST.getlist('selected_labresult_ids')
        lab_qs = LabTestAssignment.objects.filter(
            patient_card=patient, status='done', pk__in=selected_lab_ids
        ).prefetch_related('result_logs')
        diag_qs = DiagnosticAssignment.objects.filter(
            patient_card=patient, status='done', pk__in=selected_diag_ids
        ).prefetch_related('result_logs')
        from apps.laboratory.models import LabResult
        labresult_qs = LabResult.objects.filter(
            patient_card=patient, status__in=('done', 'verified', 'printed'), pk__in=selected_labresult_ids
        ).select_related('template').prefetch_related('values__parameter')

        parts = []
        for lab in lab_qs:
            log = lab.result_logs.first()
            text = f"🧪 {lab.test_name}"
            if log and log.result_text:
                text += f": {log.result_text}"
            if log and log.recommendation:
                text += f"\n{_('Tavsiya')}: {log.recommendation}"
            parts.append(text)
        for diag in diag_qs:
            log = diag.result_logs.first()
            text = f"🔬 {diag.get_diagnostic_type_display()}"
            if log and log.conclusion:
                text += f": {log.conclusion}"
            if log and log.recommendation:
                text += f"\n{_('Tavsiya')}: {log.recommendation}"
            parts.append(text)
        for res in labresult_qs:
            text = f"🧪 {res.template.name}"
            values = [f"{v.parameter.name}: {v.value}" for v in res.values.all() if v.value]
            if values:
                text += " — " + "; ".join(values)
            if res.conclusion:
                text += f"\n{_('Xulosa')}: {res.conclusion}"
            parts.append(text)
        exam.lab_investigations = '\n\n'.join(parts)

        exam.save()
        exam.selected_lab_tests.set(lab_qs)
        exam.selected_diagnostics.set(diag_qs)
        exam.selected_lab_results.set(labresult_qs)
        messages.success(request, _("✅ Ko'rik saqlandi."))
        return redirect(next_url)

    exam_type_val = exam.examination_type if exam else request.GET.get('type', '')
    if not exam and exam_type_val not in valid_exam_types:
        messages.error(request, _("Ko'rik turini tanlang."))
        return redirect(next_url)
    exam_type_display = valid_exam_types.get(exam_type_val, '')
    is_epicrisis  = exam_type_val in EPICRISIS_HIDDEN
    hidden        = EPICRISIS_HIDDEN.get(exam_type_val, set())
    show_drug     = exam_type_val not in CHECKUP_TYPES

    # Bemorning bo'lim mudiri — tahrirlashda saqlanganini, yangi ko'rikda avtomatik topamiz
    if exam and exam.department_head_name:
        dept_head_default = exam.department_head_name
    else:
        dept_head_obj = CustomUser.objects.filter(
            department=patient.department,
            role__in=('doctor', 'old'),
            is_head=True,
            is_active=True
        ).first()
        dept_head_default = dept_head_obj.full_name if dept_head_obj else ''

    # Template uchun faqat ko'rinadigan maydonlarni yuboramiz
    visible_fields = [
        (fname, label) for fname, label in EXAM_FIELDS
        if fname not in hidden and (fname != 'drug_justification' or show_drug)
    ]

    doctor = _get_doctor_profile(request.user)
    visible_fields_with_templates = []
    for fname, label in visible_fields:
        tmpls = list(
            DoctorTextTemplate.objects.filter(doctor=doctor, kind=fname).order_by('title')
        ) if doctor else []
        visible_fields_with_templates.append((fname, label, tmpls))

    available_lab_results = patient.lab_test_assignments.filter(status='done').prefetch_related('result_logs')
    available_diag_results = patient.diagnostic_assignments.filter(status='done').prefetch_related('result_logs')
    available_labresult_results = patient.lab_results.filter(
        status__in=('done', 'verified', 'printed')
    ).select_related('template').prefetch_related('values__parameter')
    selected_lab_ids  = set(exam.selected_lab_tests.values_list('pk', flat=True)) if exam else set()
    selected_diag_ids = set(exam.selected_diagnostics.values_list('pk', flat=True)) if exam else set()
    selected_labresult_ids = set(exam.selected_lab_results.values_list('pk', flat=True)) if exam else set()

    return render(request, 'patients/examination_form.html', {
        'visible_fields_with_templates': visible_fields_with_templates,
        'patient':              patient,
        'exam':                 exam,
        'exam_type':            exam_type_val,
        'exam_type_display':    exam_type_display,
        'next_url':             next_url,
        'visible_fields':       visible_fields,
        'is_epicrisis':         is_epicrisis,
        'dept_head_default':    dept_head_default,
        'dept_heads':           CustomUser.objects.filter(role__in=('doctor', 'old'), is_head=True, is_active=True)
                                              .select_related('department')
                                              .order_by('department__name', 'first_name'),
        'now':                  timezone.localtime(),
        'diag_type_choices':    EpisodeDiagnosis.DIAGNOSIS_TYPE_CHOICES,
        'diag_role_choices':    EpisodeDiagnosis.DIAGNOSIS_ROLE_CHOICES,
        'disease_course_choices': EpisodeDiagnosis.DISEASE_COURSE_CHOICES,
        'existing_diagnoses':   patient.episode_diagnoses.select_related('icd10_code').all(),
        'available_lab_results':  available_lab_results,
        'available_diag_results': available_diag_results,
        'available_labresult_results': available_labresult_results,
        'selected_lab_ids':  selected_lab_ids,
        'selected_diag_ids': selected_diag_ids,
        'selected_labresult_ids': selected_labresult_ids,
    })


@login_required
def medical_examination_print(request, patient_id, exam_pk):
    """Ko'rikni chop etish uchun toza (shapkali) ko'rinish."""
    patient = get_object_or_404(PatientCard, pk=patient_id)
    exam = get_object_or_404(MedicalExamination, pk=exam_pk, patient_card=patient)

    valid_exam_types = dict(MedicalExamination.EXAM_TYPE_CHOICES)
    hidden = EPICRISIS_HIDDEN.get(exam.examination_type, set())
    show_drug = exam.examination_type not in CHECKUP_TYPES

    fields = [
        (label, getattr(exam, fname))
        for fname, label in EXAM_FIELDS
        if fname not in hidden and (fname != 'drug_justification' or show_drug)
    ]

    import base64
    from django.conf import settings as dj_settings
    import os as _os

    header_b64 = ''
    for header_path in [
        _os.path.join(dj_settings.STATIC_ROOT, 'img', 'hospital_header.png'),
        _os.path.join(dj_settings.BASE_DIR, 'static', 'img', 'hospital_header.png'),
    ]:
        if _os.path.exists(header_path):
            with open(header_path, 'rb') as f:
                header_b64 = base64.b64encode(f.read()).decode()
            break

    return render(request, 'patients/examination_print.html', {
        'patient':           patient,
        'exam':              exam,
        'exam_type_display': valid_exam_types.get(exam.examination_type, ''),
        'fields':            fields,
        'diagnoses':         patient.episode_diagnoses.select_related('icd10_code').all(),
        'header_b64':        header_b64,
        'print_date':        timezone.now(),
    })


@login_required
def medical_examinations_list(request, patient_id):
    """AJAX — ko'riklar ro'yxati JSON."""
    patient = get_object_or_404(PatientCard, pk=patient_id)
    exams   = patient.medical_examinations.select_related('doctor').all()
    data    = [{
        'id':           e.id,
        'type':         e.examination_type,
        'type_display': e.get_examination_type_display(),
        'datetime':     e.examination_datetime.strftime('%d.%m.%Y %H:%M') if e.examination_datetime else '',
        'doctor':       str(e.doctor) if e.doctor else '',
        'created_at':   e.created_at.strftime('%d.%m.%Y %H:%M'),
    } for e in exams]
    return JsonResponse({'exams': data})


@login_required
@require_POST
def medical_examination_delete(request, pk):
    exam = get_object_or_404(MedicalExamination, pk=pk)
    patient_id = exam.patient_card_id
    exam.delete()
    return JsonResponse({'success': True, 'patient_id': patient_id})


def operation_type_search(request):
    q = request.GET.get('q', '')
    if len(q) < 1:
        results = OperationType.objects.filter(is_active=True)[:20]
    else:
        results = OperationType.objects.filter(
            Q(name__icontains=q) | Q(code__icontains=q),
            is_active=True
        )[:15]
    data = [{'id': r.id, 'name': str(r)} for r in results]
    return JsonResponse(data, safe=False)


# ==================== PATIENT VIEWS ====================

# apps/patients/views.py — patient_list view

@login_required
def patient_list(request):
    qs = PatientCard.objects.select_related(
        'department', 'attending_doctor', 'registered_by'
    ).order_by('-admission_date')

    # Oddiy shifokor (bo'lim mudiri emas) — faqat o'ziga biriktirilgan bemorlarni ko'radi
    # (bo'lim filteridan ko'ra aniqroq — o'tkazilgan bemorlarni ham qamrab oladi)
    doctor_profile = None
    if request.user.role == 'doctor' and not request.user.is_superuser:
        doctor_profile = request.user

    if doctor_profile and doctor_profile.is_general_practitioner:
        pass  # Terapevt — barcha ro'yxatga olingan bemorlarni (bo'limidan qat'iy nazar) ko'radi
    elif doctor_profile and not doctor_profile.is_head:
        qs = qs.filter(attending_doctor=doctor_profile)
    elif request.user.role == 'doctor' and not request.user.is_superuser and not doctor_profile:
        qs = qs.none()
    else:
        # Bo'lim filteri — boshqa rollar va bo'lim mudirlari uchun
        qs = department_filter(qs, request.user)

    # Qabulxona barcha bemorlarni ko'radi (filtrsiz)

    # Qidiruv
    query = request.GET.get('q', '')
    if query:
        qs = qs.filter(
            Q(full_name__icontains=query) |
            Q(medical_record_number__icontains=query) |
            Q(case_sheet_number__icontains=query) |
            Q(passport_serial__icontains=query)
        )

    # Status filteri — yakunlangan/chiqarilgan faqat aniq filter qilinganda chiqadi
    status = request.GET.get('status', '')
    if status == 'discharged':
        # "Chiqarilgan" — bazada bu holat status='completed' sifatida saqlanadi
        qs = qs.filter(status='completed')
    elif status:
        qs = qs.filter(status=status)
    else:
        qs = qs.exclude(status='completed')

    # Yakun filteri
    outcome = request.GET.get('outcome', '')
    if outcome:
        qs = qs.filter(outcome=outcome)

    # Bo'lim filteri (faqat admin)
    dept_filter = request.GET.get('department', '')
    if dept_filter and (request.user.is_superuser or request.user.role == 'admin'):
        qs = qs.filter(department_id=dept_filter)

    # Visit type filteri (ambulator/statsionar)
    visit_type = request.GET.get('visit_type', '')
    if visit_type:
        qs = qs.filter(visit_type=visit_type)

    # Shifokor filteri
    doctor_filter = request.GET.get('doctor', '')
    if doctor_filter:
        qs = qs.filter(attending_doctor_id=doctor_filter)

    # Registrator o'zi qo'shgan bemorlar filteri
    my_patients = request.GET.get('my_patients', '')
    if my_patients and request.user.role == 'reception':
        qs = qs.filter(registered_by=request.user)

    # Bemor kategoriyasi filteri
    category_filter = request.GET.get('category', '')
    if category_filter:
        qs = qs.filter(patient_category=category_filter)

    # Sana filteri — "Chiqarilgan/Yakunlangan" status tanlanganda chiqish sanasi,
    # aks holda registratsiya (qabul) sanasi bo'yicha filtrlanadi
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    date_field = 'discharge_date' if (status in ('discharged', 'completed') or outcome) else 'admission_date'
    if date_from:
        qs = qs.filter(**{f'{date_field}__date__gte': date_from})
    if date_to:
        qs = qs.filter(**{f'{date_field}__date__lte': date_to})

    # Filter uchun ro'yxatlar
    from .models import Department
    if request.user.is_superuser or request.user.role in ('admin', 'reception', 'statistician'):
        departments = Department.objects.filter(is_active=True)
    else:
        dept_ids = request.user.get_all_department_ids()
        departments = Department.objects.filter(pk__in=dept_ids) if dept_ids else Department.objects.none()

    doctors = CustomUser.objects.filter(role__in=('doctor', 'old'), is_active=True).select_related('department')
    if not request.user.is_superuser and request.user.role not in ('admin', 'reception', 'statistician'):
        dept_ids = request.user.get_all_department_ids()
        if dept_ids:
            doctors = doctors.filter(department_id__in=dept_ids)

    # Sahifa o'lchami
    per_page = request.GET.get('per_page', '20')
    try:
        per_page = int(per_page)
        if per_page not in (20, 50, 100):
            per_page = 20
    except ValueError:
        per_page = 20

    paginator = Paginator(qs, per_page)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'patients/patient_list.html', {
        'page_obj': page,
        'query': query,
        'selected_status': status,
        'selected_visit_type': visit_type,
        'selected_outcome': outcome,
        'selected_dept': dept_filter,
        'selected_doctor': doctor_filter,
        'selected_category': category_filter,
        'selected_date_from': date_from,
        'selected_date_to': date_to,
        'selected_per_page': per_page,
        'my_patients': my_patients,
        'departments': departments,
        'doctors': doctors,
        'total_count': qs.count(),
    })


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(
        PatientCard.objects.select_related(
            'department', 'attending_doctor', 'department_head',
            'referral_organization', 'country', 'region', 'district',
            'city', 'discharge_conclusion'
        ).prefetch_related('operations__operation_type'),
        pk=pk
    )

    # Bo'lim tekshiruvi
    if not request.user.is_superuser and request.user.role != 'admin':
        if request.user.role == 'reception':
            if patient.registered_by != request.user:
                messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
                return redirect('patient_list')
        else:
            dept_ids = request.user.get_all_department_ids()
            if dept_ids and patient.department_id not in dept_ids:
                messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
                return redirect('patient_list')

    death_cause = getattr(patient, 'death_cause', None)
    address_parts = filter(None, [
        str(patient.country) if patient.country else '',
        str(patient.region) if patient.region else '',
        str(patient.district) if patient.district else '',
        str(patient.city) if patient.city else '',
        patient.street_address or '',
    ])
    full_address = ', '.join(address_parts) or '—'

    # Xizmatlar va jami narx
    from apps.services.models import PatientService
    from django.db.models import Sum
    patient_services = PatientService.objects.filter(
        patient_card=patient
    ).select_related('service__category', 'ordered_by').order_by('-ordered_at')
    services_total = sum(s.price * s.quantity for s in patient_services) or 0

    # Bemor tarixi — oldingi tashriflar
    prev_visits = []
    if patient.JSHSHIR:
        prev_visits = PatientCard.objects.filter(
            JSHSHIR=patient.JSHSHIR
        ).exclude(pk=patient.pk).order_by('-admission_date')
    elif patient.passport_serial:
        prev_visits = PatientCard.objects.filter(
            passport_serial=patient.passport_serial
        ).exclude(pk=patient.pk).exclude(
            passport_serial=''
        ).order_by('-admission_date')

    # Dorilar
    from apps.services.models import PatientMedicine
    patient_medicines = PatientMedicine.objects.filter(
        patient_card=patient
    ).select_related('medicine', 'ordered_by').order_by('-ordered_at')
    medicines_total = sum(m.total_price for m in patient_medicines) or 0

    grand_total = float(services_total or 0) + float(medicines_total or 0)

    # Ko'chirish tarixi va xizmatlarni davrlarga bo'lib guruhlash
    transfers = list(
        patient.patient_transfers.all().select_related(
            'from_department', 'from_doctor', 'from_dept_head',
            'to_department', 'to_doctor', 'to_dept_head', 'transferred_by'
        ).order_by('transferred_at')
    )

    # Xizmatlarni vaqt bo'yicha tartiblash (ascending)
    all_svc_asc = list(PatientService.objects.filter(
        patient_card=patient
    ).select_related('service__category', 'ordered_by').order_by('ordered_at'))

    transfer_periods = []
    if transfers:
        # Birinchi davr: yotqizilishdan birinchi ko'chirishgacha
        boundaries = [None] + [t.transferred_at for t in transfers]
        for i, t in enumerate(transfers):
            t_start = boundaries[i]
            t_end   = t.transferred_at
            if t_start is None:
                period_svcs = [s for s in all_svc_asc if s.ordered_at < t_end]
            else:
                period_svcs = [s for s in all_svc_asc if t_start <= s.ordered_at < t_end]
            transfer_periods.append({
                'period_num':    i + 1,
                'start':         t_start,
                'end':           t_end,
                'transfer':      t,
                'dept':          t.from_department,
                'doctor':        t.from_doctor,
                'dept_head':     t.from_dept_head,
                'services':      period_svcs,
                'services_total': sum(s.price * s.quantity for s in period_svcs),
            })
        # Oxirgi davr: so'nggi ko'chirishdan hozirga qadar
        last_t = transfers[-1]
        last_svcs = [s for s in all_svc_asc if s.ordered_at >= last_t.transferred_at]
        transfer_periods.append({
            'period_num':    len(transfers) + 1,
            'start':         last_t.transferred_at,
            'end':           None,
            'transfer':      None,
            'dept':          last_t.to_department,
            'doctor':        last_t.to_doctor,
            'dept_head':     last_t.to_dept_head,
            'services':      last_svcs,
            'services_total': sum(s.price * s.quantity for s in last_svcs),
        })

    return render(request, 'patients/patient_detail.html', {
        'patient': patient,
        'death_cause': death_cause,
        'full_address': full_address,
        'patient_services': patient_services,
        'services_total': services_total,
        'patient_medicines': patient_medicines,
        'medicines_total': medicines_total,
        'grand_total': grand_total,
        'prev_visits': prev_visits,
        'departments': Department.objects.filter(is_active=True).order_by('name'),
        'doctors':     CustomUser.objects.filter(role__in=('doctor', 'old'), is_active=True).select_related('department').order_by('department__name', 'first_name'),
        'dept_heads':  CustomUser.objects.filter(role__in=('doctor', 'old'), is_active=True, is_head=True).order_by('first_name'),
        'today':       timezone.localdate(),
        'now':         timezone.localtime(),
        'transfers':   transfers,
        'transfer_periods': transfer_periods,
        'discharge_conclusions': DischargeConclusion.objects.filter(is_active=True).order_by('name'),
        'outcome_choices':    PatientCard.OUTCOME_CHOICES,
        'status_choices':     PatientCard.STATUS_CHOICES,
        'exam_type_choices':  MedicalExamination.EXAM_TYPE_CHOICES,
    })


def get_doctors(request):
    """AJAX — bo'lim bo'yicha shifokorlarni qaytarish"""
    department_id = request.GET.get('department_id')
    if not department_id:
        return JsonResponse([], safe=False)

    doctors = CustomUser.objects.filter(
        role__in=('doctor', 'old'),
        department_id=department_id,
        is_active=True
    ).values('id', 'first_name', 'last_name', 'is_head').order_by('-is_head', 'first_name')

    result = [
        {'id': d['id'], 'full_name': f"{d['first_name']} {d['last_name']}".strip(), 'is_head': d['is_head']}
        for d in doctors
    ]
    return JsonResponse(result, safe=False)

@login_required
@role_required('admin', 'doctor', 'statistician')
def patient_card_create(request):
    if request.method == 'POST':
        form = PatientCardForm(request.POST)
        death_form = DeathCauseForm(request.POST)
        surgery_formset = SurgicalOperationFormSet(request.POST)

        is_deceased = request.POST.get('outcome') == 'deceased'
        forms_valid = form.is_valid() and surgery_formset.is_valid()
        if is_deceased:
            forms_valid = forms_valid and death_form.is_valid()

        if forms_valid:
            dup = _find_recent_duplicate(
                form.cleaned_data.get('full_name', ''),
                form.cleaned_data.get('birth_date'),
            )
            if dup:
                messages.warning(request, _("⚠️ Bu bemor 2 daqiqa ichida allaqachon ro'yxatga olingan: %(name)s (#%(record_number)s)") % {'name': dup.full_name, 'record_number': dup.medical_record_number})
                return redirect('patient_detail', pk=dup.pk)

            patient = form.save(commit=False)
            if not request.user.is_superuser and request.user.role != 'admin':
                if request.user.department:
                    patient.department = request.user.department
                elif request.user.departments.exists():
                    patient.department = request.user.departments.first()
            patient.save()
            form.save_m2m()

            surgeries = surgery_formset.save(commit=False)
            for s in surgeries:
                s.patient_card = patient
                s.save()
            for s in surgery_formset.deleted_objects:
                s.delete()

            if is_deceased:
                death = death_form.save(commit=False)
                death.patient_card = patient
                death.save()

            messages.success(request, _("Bemor kartasi saqlandi!"))
            return redirect('patient_list')
        else:
            messages.error(request, _("Formada xatoliklar bor. Tekshiring."))
    else:
        form = PatientCardForm()
        if not request.user.is_superuser and request.user.role != 'admin':
            primary = request.user.department or request.user.departments.first()
            if primary:
                form.initial['department'] = primary
        death_form = DeathCauseForm()
        surgery_formset = SurgicalOperationFormSet()

    return render(request, 'patients/patient_form.html', {
        'form': form,
        'death_form': death_form,
        'surgery_formset': surgery_formset,
        'title': 'Yangi bemor kartasi',
    })

def sevgi_maktubi(request):
    return render(request, 'patients/sevgim.html')

@login_required
@role_required('admin', 'doctor', 'statistician', 'reception', 'old')
def patient_card_edit(request, pk):
    patient = get_object_or_404(PatientCard, pk=pk)

    # Ruxsat tekshiruvi — shifokor faqat o'z bo'limini tahrirlaydi
    if not request.user.is_superuser and request.user.role not in ('admin', 'reception'):
        if request.user.department and patient.department != request.user.department:
            messages.error(request, _("Siz bu bemorni tahrirlay olmaysiz."))
            return redirect('patient_detail', pk=pk)

    is_ambulatory = patient.visit_type == 'ambulatory'
    is_reception  = request.user.role == 'reception'

    # Ambulator bemor uchun ReceptionForm (soddalashtirilgan)
    # Statsionar uchun rol ga qarab forma
    if is_ambulatory:
        FormClass = ReceptionForm
    elif is_reception:
        FormClass = ReceptionForm
    else:
        FormClass = PatientCardForm

    death_instance = getattr(patient, 'death_cause', None)

    if request.method == 'POST':
        # Chiqarish modal dan kelgan so'rov — faqat admin/doctor/statistician, faqat statsionar
        if request.POST.get('_discharge'):
            next_url = _safe_next_url(request, reverse('patient_detail', args=[pk]))
            if patient.visit_type == 'ambulatory':
                messages.error(request, _("Ambulator bemor chiqarilmaydi."))
                return redirect(next_url)
            if request.user.role == 'reception' and not request.user.is_superuser:
                messages.error(request, _("Sizda bemorni chiqarish huquqi yo'q."))
                return redirect(next_url)

        if request.POST.get('_discharge'):
            patient.outcome = request.POST.get('outcome', '')
            patient.status  = 'completed'
            patient.final_diagnosis = request.POST.get('final_diagnosis', '')
            patient.discharge_note  = request.POST.get('discharge_conclusion_text', '')
            discharge_date = request.POST.get('discharge_date', '')
            if discharge_date:
                from datetime import date
                try:
                    patient.discharge_date = date.fromisoformat(discharge_date)
                except ValueError:
                    pass
            doc_id  = request.POST.get('attending_doctor')
            head_id = request.POST.get('department_head')
            if doc_id:
                patient.attending_doctor = CustomUser.objects.filter(pk=doc_id).first()
            if head_id:
                patient.department_head = CustomUser.objects.filter(pk=head_id).first()
            days_str = request.POST.get('days_in_hospital', '')
            if days_str:
                try:
                    patient.days_in_hospital = int(days_str)
                except ValueError:
                    pass
            conclusion_id = request.POST.get('discharge_conclusion')
            if conclusion_id:
                patient.discharge_conclusion = DischargeConclusion.objects.filter(pk=conclusion_id).first()
            patient.save()
            messages.success(request, _("✅ Bemor chiqarildi: %(name)s") % {'name': patient.full_name})
            return redirect(next_url)

        form = FormClass(request.POST, instance=patient)

        if is_ambulatory or is_reception:
            if form.is_valid():
                obj = form.save(commit=False)
                # visit_type o'zgarmasin
                obj.visit_type = patient.visit_type or ('ambulatory' if is_ambulatory else 'inpatient')
                obj.save()
                messages.success(request, _("Ma'lumotlar yangilandi!"))
                return redirect('patient_detail', pk=pk)
            else:
                # Ambulator uchun required bo'lmagan xatolarni o'chirish
                if is_ambulatory:
                    skip = ['medical_record_number','resident_status','admission_date',
                            'department','days_in_hospital','hospital_type']
                    for f in skip:
                        if f in form.errors:
                            del form.errors[f]
                    if not form.errors:
                        obj = form.save(commit=False)
                        obj.visit_type = 'ambulatory'
                        obj.save()
                        messages.success(request, _("Ma'lumotlar yangilandi!"))
                        return redirect('patient_detail', pk=pk)
                messages.error(request, _("Formada xatoliklar bor."))
        else:
            death_form = DeathCauseForm(request.POST, instance=death_instance)
            surgery_formset = SurgicalOperationFormSet(request.POST, instance=patient)

            is_deceased = request.POST.get('outcome') == 'deceased'
            forms_valid = form.is_valid() and surgery_formset.is_valid()
            if is_deceased:
                forms_valid = forms_valid and death_form.is_valid()

            if forms_valid:
                obj = form.save(commit=False)
                # visit_type formada yo'q — mavjud qiymatni saqlash
                if not obj.visit_type:
                    obj.visit_type = patient.visit_type or 'inpatient'
                obj.save()
                patient = obj
                surgeries = surgery_formset.save(commit=False)
                for s in surgeries:
                    s.patient_card = patient
                    s.save()
                for s in surgery_formset.deleted_objects:
                    s.delete()

                if is_deceased:
                    death = death_form.save(commit=False)
                    death.patient_card = patient
                    death.save()
                elif death_instance:
                    death_instance.delete()

                messages.success(request, _("Bemor kartasi yangilandi!"))
                return redirect('patient_detail', pk=pk)
            else:
                messages.error(request, _("Formada xatoliklar bor."))
    else:
        form = FormClass(instance=patient)
        if not is_reception:
            death_form = DeathCauseForm(instance=death_instance)
            surgery_formset = SurgicalOperationFormSet(instance=patient)

    if is_ambulatory:
        return render(request, 'patients/ambulatory_form.html', {
            'form': form,
            'title': f"Tahrirlash: {patient.full_name}",
            'patient': patient,
            'auto_record_number': patient.medical_record_number,
            'now': patient.admission_date.strftime('%Y-%m-%dT%H:%M') if patient.admission_date else '',
        })

    if is_reception:
        return render(request, 'patients/reception_form.html', {
            'form': form,
            'title': f"Tahrirlash: {patient.full_name}",
            'patient': patient,
        })

    return render(request, 'patients/patient_form.html', {
        'form': form,
        'death_form': death_form,
        'surgery_formset': surgery_formset,
        'title': f"Tahrirlash: {patient.full_name}",
        'patient': patient,
        'doctors': CustomUser.objects.filter(role__in=('doctor', 'old'), is_active=True).select_related('department').order_by('department__name', 'first_name'),
    })


@login_required
@role_required('admin')
def patient_delete(request, pk):
    patient = get_object_or_404(PatientCard, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.success(request, _("Bemor kartasi o'chirildi."))
        return redirect('patient_list')
    return render(request, 'patients/patient_confirm_delete.html', {'patient': patient})


@login_required
@role_required('admin', 'reception')
def reception_create(request):
    if request.method == 'POST':
        form = ReceptionForm(request.POST)
        if form.is_valid():
            # Takroriy yuborish himoyasi — 2 daqiqa ichida bir xil ism+sana
            dup = _find_recent_duplicate(
                form.cleaned_data.get('full_name', ''),
                form.cleaned_data.get('birth_date'),
            )
            if dup:
                messages.warning(request, _("⚠️ Bu bemor 2 daqiqa ichida allaqachon qabul qilingan: %(name)s (#%(record_number)s)") % {'name': dup.full_name, 'record_number': dup.medical_record_number})
                return redirect('patient_detail', pk=dup.pk)

            patient = form.save(commit=False)

            # Bo'limni avtomatik qo'yish
            if not request.user.is_superuser and request.user.role != 'admin':
                if request.user.department:
                    patient.department = request.user.department
                elif request.user.departments.exists():
                    patient.department = request.user.departments.first()

            # Avtomatik bayonnoma raqami
            year = timezone.now().year
            while True:
                record_number = f"{year}-{str(uuid.uuid4())[:6].upper()}"
                if not PatientCard.objects.filter(
                    medical_record_number=record_number
                ).exists():
                    break

            patient.medical_record_number = record_number
            patient.status = 'registered'
            patient.registered_by = request.user

            # Qo'lda kiritilgan shartnoma raqamini saqlash
            manual_contract_number = request.POST.get('manual_contract_number', '').strip()
            if manual_contract_number:
                patient._manual_contract_number = manual_contract_number
            patient.save()
            InitialExamination.objects.get_or_create(patient_card=patient)
            _notify_new_admission(patient)

            messages.success(
                request,
                _("✅ Bemor qabul qilindi! Bayonnoma: %(record_number)s") % {'record_number': patient.medical_record_number}
            )
            return redirect('patient_detail', pk=patient.pk)
        else:
            messages.error(request, _("Formada xatoliklar bor."))
    else:
        form = ReceptionForm()
        primary = request.user.department or request.user.departments.first()
        if primary:
            form.initial['department'] = primary

    return render(request, 'patients/reception_form.html', {
        'form': form,
        'title': 'Bemor qabul qilish',
    })
@login_required
def patient_card_excel(request, pk):
    """Bemor kartasi + xizmatlar Excel export"""
    patient = get_object_or_404(
        PatientCard.objects.select_related(
            'department', 'attending_doctor', 'department_head',
            'referral_organization', 'country', 'region',
            'district', 'city', 'discharge_conclusion'
        ),
        pk=pk
    )

    from apps.services.models import PatientService
    from django.db.models import Sum, Count
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Stillar
    BLUE   = PatternFill('solid', fgColor='1F4E79')
    LBLUE  = PatternFill('solid', fgColor='D6E4F0')
    LLBLUE = PatternFill('solid', fgColor='EBF5FB')
    GREEN  = PatternFill('solid', fgColor='E9F7EF')
    TOTAL  = PatternFill('solid', fgColor='1E8449')

    W_FONT  = Font(color='FFFFFF', bold=True, size=10)
    BOLD    = Font(bold=True, size=10)
    NORMAL  = Font(size=10)
    CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT    = Alignment(horizontal='left', vertical='center', wrap_text=True)
    RIGHT   = Alignment(horizontal='right', vertical='center')
    thin    = Side(style='thin')
    BRD     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(ws, row, col, value='', fill=None, font=None, align=LEFT):
        c = ws.cell(row=row, column=col, value=value)
        if fill:  c.fill  = fill
        if font:  c.font  = font
        c.alignment = align
        c.border    = BRD
        return c

    def section_header(ws, row, title, ncols=2):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=title)
        c.fill = LBLUE; c.font = BOLD; c.alignment = LEFT; c.border = BRD
        ws.row_dimensions[row].height = 20
        return row + 1

    def info_row(ws, row, label, value):
        cell(ws, row, 1, label, font=BOLD)
        cell(ws, row, 2, str(value) if value else '—', font=NORMAL)
        ws.row_dimensions[row].height = 18
        return row + 1

    # ===== 1-SAHIFA: BEMOR MA'LUMOTLARI =====
    ws1 = wb.active
    ws1.title = "Bemor ma'lumotlari"
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 42

    address_parts = list(filter(None, [
        str(patient.country)  if patient.country  else '',
        str(patient.region)   if patient.region   else '',
        str(patient.district) if patient.district else '',
        str(patient.city)     if patient.city     else '',
        patient.street_address or '',
    ]))
    full_address = ', '.join(address_parts) or '—'

    ws1.merge_cells('A1:B1')
    c = ws1.cell(row=1, column=1,
                 value=f"BEMOR KARTASI — {patient.full_name} ({patient.medical_record_number})")
    c.fill = BLUE; c.font = W_FONT; c.alignment = CENTER; c.border = BRD
    ws1.row_dimensions[1].height = 30

    r = 2
    r = section_header(ws1, r, "SHAXSIY MA'LUMOTLAR")
    r = info_row(ws1, r, "Bayonnoma raqami", patient.medical_record_number)
    r = info_row(ws1, r, "Ism-familiya", patient.full_name)
    r = info_row(ws1, r, "Jinsi", patient.get_gender_display())
    r = info_row(ws1, r, "Tug'ilgan sana", patient.birth_date.strftime('%d.%m.%Y') if patient.birth_date else '—')
    r = info_row(ws1, r, "Rezidentlik", patient.get_resident_status_display())
    r = info_row(ws1, r, "Bemor kategoriyasi", patient.get_patient_category_display())
    r = info_row(ws1, r, "Telefon", patient.phone or '—')
    r = info_row(ws1, r, "Manzil", full_address)
    r = info_row(ws1, r, "Ijtimoiy holat", patient.get_social_status_display() if patient.social_status else '—')
    if patient.social_status == 'dependent' and patient.parent_name:
        r = info_row(ws1, r, "Ota-ona ismi", patient.parent_name)
        if patient.parent_jshshir:
            r = info_row(ws1, r, "Ota-ona JSHSHIR", patient.parent_jshshir)
        if patient.parent_workplace_org:
            r = info_row(ws1, r, "Ota-ona ish joyi", str(patient.parent_workplace_org))
    r = info_row(ws1, r, "Ish joyi", patient.workplace or '—')
    r = info_row(ws1, r, "Lavozim", patient.position or '—')
    r = info_row(ws1, r, "Passport", patient.passport_serial or '—')
    r = info_row(ws1, r, "JSHSHIR", patient.JSHSHIR or '—')

    r = section_header(ws1, r, "QABUL MA'LUMOTLARI")
    r = info_row(ws1, r, "Qabul sanasi", patient.admission_date.strftime('%d.%m.%Y %H:%M') if patient.admission_date else '—')
    r = info_row(ws1, r, "Bo'lim", str(patient.department) if patient.department else '—')
    r = info_row(ws1, r, "Kim olib kelgan", patient.get_referral_type_display() if patient.referral_type else '—')
    r = info_row(ws1, r, "Yo'llagan muassasa", str(patient.referral_organization) if patient.referral_organization else '—')
    r = info_row(ws1, r, "Yo'llanma tashxisi", patient.referring_diagnosis or '—')
    r = info_row(ws1, r, "Qabul tashxisi", patient.admission_diagnosis or '—')
    r = info_row(ws1, r, "Kasallanishdan keyin", patient.get_hours_after_illness_display() if patient.hours_after_illness else '—')
    r = info_row(ws1, r, "Shoshilinch", 'Ha' if patient.is_emergency else "Yo'q")
    r = info_row(ws1, r, "Shifoxona turi", str(patient.hospital_type) if patient.hospital_type else '—')

    r = section_header(ws1, r, "CHIQISH MA'LUMOTLARI")
    r = info_row(ws1, r, "Yotgan kunlar", f"{patient.days_in_hospital} kun")
    r = info_row(ws1, r, "Yakun", patient.get_outcome_display() if patient.outcome else '—')
    r = info_row(ws1, r, "Chiqish xulosasi", str(patient.discharge_conclusion) if patient.discharge_conclusion else '—')
    r = info_row(ws1, r, "Chiqish sanasi", patient.discharge_date.strftime('%d.%m.%Y') if patient.discharge_date else '—')

    r = section_header(ws1, r, "YAKUNIY TASHXIS")
    r = info_row(ws1, r, "Klinik tashxis (MKB-10)",
                 f"{patient.clinical_main_diagnosis or ''} {patient.clinical_main_diagnosis_text or ''}".strip() or '—')
    r = info_row(ws1, r, "Yo'ldosh kasalliklar", patient.clinical_comorbidities or '—')

    r = section_header(ws1, r, "SHIFOKORLAR")
    r = info_row(ws1, r, "Davolovchi shifokor", str(patient.attending_doctor) if patient.attending_doctor else '—')
    r = info_row(ws1, r, "Bo'lim mudiri", str(patient.department_head) if patient.department_head else '—')

    # ===== 2-SAHIFA: XIZMATLAR =====
    ws2 = wb.create_sheet("Xizmatlar")
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    # Sarlavha
    ws2.merge_cells('A1:D1')
    c = ws2.cell(row=1, column=1,
                 value=f"XIZMATLAR HISOBOTI — {patient.full_name}")
    c.fill = BLUE; c.font = W_FONT; c.alignment = CENTER; c.border = BRD
    ws2.row_dimensions[1].height = 30

    # Bemor info
    ws2.merge_cells('A2:D2')
    c = ws2.cell(row=2, column=1,
                 value=f"Bayonnoma: {patient.medical_record_number} | "
                       f"Kategoriya: {patient.get_patient_category_display()} | "
                       f"Qabul: {patient.admission_date.strftime('%d.%m.%Y') if patient.admission_date else '—'}")
    c.font = NORMAL; c.alignment = LEFT; c.border = BRD
    ws2.row_dimensions[2].height = 18

    # Ustun sarlavhalar
    headers = ['Xizmat nomi', 'Miqdor', "Narx (so'm)", "Jami (so'm)"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.fill = BLUE; c.font = W_FONT
        c.alignment = CENTER if col > 1 else LEFT
        c.border = BRD
    ws2.row_dimensions[3].height = 22

    # Xizmatlar
    patient_services = PatientService.objects.filter(
        patient_card=patient
    ).select_related('service__category').order_by('service__category__name', 'service__name')

    from collections import defaultdict as _dd2
    _cm2 = _dd2(lambda: {'count': 0, 'total': 0.0})
    for _s in patient_services:
        _k = _s.service.category.name if _s.service.category else 'Boshqa'
        _cm2[_k]['count'] += 1
        _cm2[_k]['total'] += float((_s.price or 0) * (_s.quantity or 1))
    cat_stats = [
        {'service__category__name': k, 'count': v['count'], 'total': v['total']}
        for k, v in sorted(_cm2.items())
    ]

    r = 4
    grand_total = 0

    for cat in cat_stats:
        cat_name = cat['service__category__name']
        cat_total = float(cat['total'] or 0)
        grand_total += cat_total

        # Kategoriya sarlavhasi
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws2.cell(row=r, column=1, value=f"▶  {cat_name}")
        c.fill = LBLUE; c.font = BOLD; c.alignment = LEFT; c.border = BRD
        ws2.row_dimensions[r].height = 20
        r += 1

        # Xizmatlar
        cat_svcs = patient_services.filter(service__category__name=cat_name)
        for svc in cat_svcs:
            cell(ws2, r, 1, f"   {svc.service.name}", font=NORMAL)
            cell(ws2, r, 2, svc.quantity, font=NORMAL, align=CENTER)
            cell(ws2, r, 3, float(svc.price), font=NORMAL, align=RIGHT)
            c4 = cell(ws2, r, 4, float(svc.total_price), font=NORMAL, align=RIGHT)
            ws2.cell(row=r, column=3).number_format = '#,##0'
            ws2.cell(row=r, column=4).number_format = '#,##0'
            ws2.row_dimensions[r].height = 18
            r += 1

        # Kategoriya jami
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws2.cell(row=r, column=1,
                     value=f"   {cat_name} bo'yicha jami ({cat['count']} ta xizmat):")
        c.fill = LLBLUE; c.font = BOLD; c.alignment = LEFT; c.border = BRD
        c4 = ws2.cell(row=r, column=4, value=cat_total)
        c4.fill = LLBLUE; c4.font = BOLD; c4.alignment = RIGHT
        c4.border = BRD; c4.number_format = '#,##0'
        ws2.row_dimensions[r].height = 20
        r += 1

    # Umumiy jami
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws2.cell(row=r, column=1, value="UMUMIY JAMI:")
    c.fill = TOTAL; c.font = W_FONT; c.alignment = LEFT; c.border = BRD
    c4 = ws2.cell(row=r, column=4, value=grand_total)
    c4.fill = TOTAL; c4.font = W_FONT; c4.alignment = RIGHT
    c4.border = BRD; c4.number_format = '#,##0'
    ws2.row_dimensions[r].height = 25

    if not patient_services.exists():
        ws2.merge_cells('A4:D4')
        c = ws2.cell(row=4, column=1, value="Xizmatlar qo'shilmagan")
        c.font = NORMAL; c.alignment = CENTER; c.border = BRD

    # ===== SHEET 3: DORI-DARMONLAR =====
    from apps.services.models import PatientMedicine as PMed
    patient_medicines = PMed.objects.filter(
        patient_card=patient
    ).select_related('medicine', 'ordered_by').order_by('medicine__name')

    ws3 = wb.create_sheet("Dori-darmonlar")
    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 20

    # Sarlavha
    ws3.merge_cells('A1:F1')
    c = ws3.cell(row=1, column=1, value=f"DORI-DARMONLAR — {patient.full_name}")
    c.fill = PatternFill('solid', fgColor='856404')
    c.font = Font(color='FFFFFF', bold=True, size=12)
    c.alignment = CENTER; c.border = BRD
    ws3.row_dimensions[1].height = 28

    heads3 = ['№', 'Dori nomi', 'Birlik', 'Miqdori', "Narxi (so'm)", "Jami (so'm)"]
    for col, h in enumerate(heads3, 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.fill = PatternFill('solid', fgColor='856404')
        c.font = Font(color='FFFFFF', bold=True, size=10)
        c.alignment = CENTER; c.border = BRD
    ws3.row_dimensions[2].height = 22

    med_grand = 0
    for ri, med in enumerate(patient_medicines, 1):
        tp = float(med.total_price)
        med_grand += tp
        row_data = [ri, med.medicine.name, med.medicine.unit,
                    float(med.quantity), float(med.price), tp]
        for col, val in enumerate(row_data, 1):
            c = ws3.cell(row=ri+2, column=col, value=val)
            c.font = NORMAL
            c.alignment = CENTER if col in (1,3,4) else (RIGHT if col in (5,6) else LEFT)
            c.border = BRD
            if col in (5,6): c.number_format = '#,##0'
            if ri % 2 == 0: c.fill = PatternFill('solid', fgColor='FFF8E1')
        ws3.row_dimensions[ri+2].height = 18

    # Jami
    last = patient_medicines.count() + 3
    if last > 3:
        ws3.merge_cells(start_row=last, start_column=1, end_row=last, end_column=5)
        c = ws3.cell(row=last, column=1, value="JAMI:")
        c.fill = PatternFill('solid', fgColor='856404')
        c.font = Font(color='FFFFFF', bold=True, size=10)
        c.alignment = LEFT; c.border = BRD
        c6 = ws3.cell(row=last, column=6, value=med_grand)
        c6.fill = PatternFill('solid', fgColor='856404')
        c6.font = Font(color='FFFFFF', bold=True, size=10)
        c6.alignment = RIGHT; c6.border = BRD
        c6.number_format = '#,##0'
        ws3.row_dimensions[last].height = 22
    else:
        ws3.merge_cells('A3:F3')
        c = ws3.cell(row=3, column=1, value="Dori-darmonlar qo'shilmagan")
        c.font = NORMAL; c.alignment = CENTER; c.border = BRD

    # ===== SHEET 4: UMUMIY HISOB =====
    ws4 = wb.create_sheet("Umumiy hisob")
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 25

    ws4.merge_cells('A1:B1')
    c = ws4.cell(row=1, column=1, value="UMUMIY HISOB")
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.font = Font(color='FFFFFF', bold=True, size=13)
    c.alignment = CENTER; c.border = BRD
    ws4.row_dimensions[1].height = 28

    rows4 = [
        ("Xizmatlar jami:", grand_total),
        ("Dori-darmonlar jami:", med_grand),
        ("UMUMIY JAMI:", grand_total + med_grand),
    ]
    fills4 = ['D6E4F0', 'FFF8E1', '145A32']
    fonts4 = [
        Font(bold=True, size=11),
        Font(bold=True, size=11),
        Font(color='FFFFFF', bold=True, size=13),
    ]
    for ri, ((label, val), fill, font) in enumerate(zip(rows4, fills4, fonts4), 2):
        c1 = ws4.cell(row=ri, column=1, value=label)
        c1.fill = PatternFill('solid', fgColor=fill)
        c1.font = font; c1.alignment = LEFT; c1.border = BRD
        c2 = ws4.cell(row=ri, column=2, value=val)
        c2.fill = PatternFill('solid', fgColor=fill)
        c2.font = font; c2.alignment = RIGHT; c2.border = BRD
        c2.number_format = '#,##0'
        ws4.row_dimensions[ri].height = 24

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"bemor_{patient.medical_record_number}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def organization_search(request):
    """AJAX — ish joyi qidirish (temir yo'lchilar uchun)"""
    q = request.GET.get('q', '').strip()
    org_id = request.GET.get('id', '').strip()

    from .models import Organization
    from django.db.models import Q

    qs = Organization.objects.filter(is_active=True)

    if org_id:
        qs = qs.filter(pk=org_id)
    elif q:
        qs = qs.filter(
            Q(enterprise_name__icontains=q) |
            Q(branch_name__icontains=q) |
            Q(enterprise_code__icontains=q) |
            Q(branch_code__icontains=q) |
            Q(enterprise_inn__icontains=q)
        )

    qs = qs.order_by('enterprise_name', 'branch_name')[:30]

    data = []
    for org in qs:
        data.append({
            'id': org.id,
            'enterprise_code': org.enterprise_code,
            'enterprise_inn': org.enterprise_inn,
            'enterprise_name': org.enterprise_name,
            'branch_code': org.branch_code,
            'branch_name': org.branch_name,
            'display': org.display_name,
        })

    return JsonResponse(data, safe=False)


@login_required
def check_existing_patient(request):
    """AJAX — JSHSHIR yoki passport bo'yicha oldingi tashriflarni tekshirish"""
    jshshir   = request.GET.get('jshshir', '').strip()
    passport  = request.GET.get('passport', '').strip()
    exclude_pk = request.GET.get('exclude', '')

    qs = PatientCard.objects.select_related('department', 'attending_doctor')

    if jshshir:
        matches = qs.filter(JSHSHIR=jshshir)
    elif passport and len(passport) >= 5:
        matches = qs.filter(passport_serial__iexact=passport)
    else:
        return JsonResponse({'found': False, 'patients': []})

    if exclude_pk:
        matches = matches.exclude(pk=exclude_pk)

    matches = matches.order_by('-admission_date')[:5]

    data = []
    for p in matches:
        data.append({
            'pk':                   p.pk,
            'full_name':            p.full_name,
            'last_name':            p.last_name  if hasattr(p, 'last_name')  else '',
            'first_name':           p.first_name if hasattr(p, 'first_name') else '',
            'middle_name':          p.middle_name if hasattr(p, 'middle_name') else '',
            'birth_date_raw':       p.birth_date.strftime('%Y-%m-%d') if p.birth_date else '',
            'gender_code':          p.gender or '',
            'jshshir':              p.JSHSHIR or '',
            'passport':             p.passport_serial or '',
            'phone':                p.phone or '',
            'address':              p.street_address or '',
            'medical_record_number': p.medical_record_number,
            'admission_date':       p.admission_date.strftime('%d.%m.%Y') if p.admission_date else '—',
            'department':           str(p.department) if p.department else '—',
            'status':               p.get_status_display(),
            'status_code':          p.status,
            'outcome':              p.get_outcome_display() if p.outcome else '—',
            'diagnosis':            p.admission_diagnosis[:60] if p.admission_diagnosis else '—',
        })

    return JsonResponse({'found': bool(data), 'patients': data})

@login_required
def patient_invoice(request, pk):
    """Eski hisob-faktura sahifasi — yangi billing moduliga yo'naltiriladi"""
    return redirect('invoice_detail', pk=pk)



@login_required
@role_required('admin', 'doctor', 'statistician')
def transfer_department(request, pk):
    """Bemorni boshqa bo'limga ko'chirish"""
    from .models import DepartmentTransfer, Department
    patient = get_object_or_404(PatientCard, pk=pk)

    if request.method == 'POST':
        new_dept_id = request.POST.get('department')
        reason      = request.POST.get('reason', '')

        if new_dept_id and str(new_dept_id) != str(patient.department_id):
            try:
                new_dept = Department.objects.get(pk=new_dept_id)
                DepartmentTransfer.objects.create(
                    patient_card=patient,
                    from_department=patient.department,
                    to_department=new_dept,
                    transferred_by=request.user,
                    reason=reason,
                )
                patient.department = new_dept
                patient.save(update_fields=['department'])
                messages.success(
                    request,
                    _("Bemor %(dept_name)s bo'limiga ko'chirildi.") % {'dept_name': new_dept.name}
                )
            except Department.DoesNotExist:
                messages.error(request, _("Bo'lim topilmadi."))
        else:
            messages.warning(request, _("Yangi bo'lim tanlang."))

    return redirect('patient_detail', pk=pk)


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception')


@login_required
@role_required('admin', 'doctor', 'statistician', 'reception')
def ambulatory_create(request):
    """Ambulator (kunlik) bemor qabuli"""
    from django.utils import timezone

    def gen_record_number():
        year = timezone.now().year
        while True:
            num = f"AMB-{year}-{str(uuid.uuid4())[:6].upper()}"
            if not PatientCard.objects.filter(medical_record_number=num).exists():
                return num

    if request.method == 'POST':
        # POST da forma validatsiyasiz kerakli maydonlarni to'g'ridan saqlash
        from django.utils.dateparse import parse_date

        record_number   = request.POST.get('medical_record_number') or gen_record_number()
        full_name       = request.POST.get('full_name', '').strip()
        gender          = request.POST.get('gender', '')
        birth_date_raw  = request.POST.get('birth_date', '').strip()
        jshshir         = request.POST.get('JSHSHIR', '').strip()
        phone           = request.POST.get('phone', '').strip()
        category        = request.POST.get('patient_category', 'paid')
        now             = timezone.now()

        errors = {}
        if not full_name:
            errors['full_name'] = "Ism-familiyani kiriting"
        if not gender:
            errors['gender'] = "Jinsni tanlang"

        birth_date = None
        if birth_date_raw:
            birth_date = parse_date(birth_date_raw)
            if birth_date is None or birth_date.year < 1900 or birth_date > now.date():
                errors['birth_date'] = "Tug'ilgan sana noto'g'ri formatda (YYYY-MM-DD)"
                birth_date = None

        if errors:
            auto_number = record_number
            form = ReceptionForm(request.POST)
            # Required xatolarni o'chirish
            for field in ['medical_record_number','resident_status','admission_date',
                          'department','days_in_hospital']:
                if field in form.errors:
                    del form.errors[field]
            return render(request, 'patients/ambulatory_form.html', {
                'form': form,
                'errors': errors,
                'auto_record_number': auto_number,
                'now': now.strftime('%Y-%m-%dT%H:%M'),
            })

        # Takroriy yuborish himoyasi — 2 daqiqa ichida bir xil ism+sana
        if birth_date:
            dup = _find_recent_duplicate(full_name, birth_date)
            if dup:
                messages.warning(request, _("⚠️ Bu bemor 2 daqiqa ichida allaqachon qabul qilingan: %(name)s (#%(record_number)s)") % {'name': dup.full_name, 'record_number': dup.medical_record_number})
                return redirect('patient_detail', pk=dup.pk)

        patient = PatientCard(
            medical_record_number = record_number,
            full_name     = full_name,
            gender        = gender,
            birth_date    = birth_date,
            JSHSHIR       = jshshir,
            phone         = phone,
            patient_category = category,
            visit_type    = 'ambulatory',
            status        = 'registered',
            registered_by = request.user,
            admission_date = now,
        )
        patient.save()
        messages.success(request, _('Ambulator bemor %(name)s qabul qilindi. Bayonnoma: %(record_number)s') % {'name': patient.full_name, 'record_number': record_number})
        return redirect('patient_detail', pk=patient.pk)

    # GET
    auto_number = gen_record_number()
    now = timezone.now()
    form = ReceptionForm()
    return render(request, 'patients/ambulatory_form.html', {
        'form': form,
        'auto_record_number': auto_number,
        'now': now.strftime('%Y-%m-%dT%H:%M'),
    })

@login_required
@role_required('admin', 'doctor', 'statistician', 'old')
@require_POST
def patient_transfer(request, pk):
    """Bemorni boshqa bo'limga ko'chirish"""
    from apps.patients.models import PatientTransfer, Department

    patient      = get_object_or_404(PatientCard, pk=pk)

    if patient.visit_type == 'ambulatory':
        messages.error(request, _("Ambulator bemor ko'chirilmaydi."))
        return redirect('patient_detail', pk=pk)

    to_dept_id   = request.POST.get('to_department')
    to_doc_id    = request.POST.get('to_doctor')
    to_head_id   = request.POST.get('to_dept_head')
    reason       = request.POST.get('reason', '').strip()

    if not to_dept_id:
        messages.error(request, _("Bo'lim tanlanmagan."))
        return redirect('patient_detail', pk=pk)

    transfer_date_str = request.POST.get('transfer_date', '').strip()
    transfer_date = None
    if transfer_date_str:
        from datetime import date as _date
        try:
            transfer_date = _date.fromisoformat(transfer_date_str)
        except ValueError:
            pass

    PatientTransfer.objects.create(
        patient_card    = patient,
        from_department = patient.department,
        from_doctor     = patient.attending_doctor,
        from_dept_head  = patient.department_head,
        to_department   = Department.objects.filter(pk=to_dept_id).first(),
        to_doctor       = CustomUser.objects.filter(pk=to_doc_id).first() if to_doc_id else None,
        to_dept_head    = CustomUser.objects.filter(pk=to_head_id).first() if to_head_id else None,
        reason          = reason,
        transfer_date   = transfer_date,
        transferred_by  = request.user,
    )

    update_fields = ['department']
    patient.department = Department.objects.filter(pk=to_dept_id).first()
    if to_doc_id:
        patient.attending_doctor = CustomUser.objects.filter(pk=to_doc_id).first()
        update_fields.append('attending_doctor')
    if to_head_id:
        patient.department_head = CustomUser.objects.filter(pk=to_head_id).first()
        update_fields.append('department_head')
    patient.save(update_fields=update_fields)

    messages.success(request, _("✅ Bemor %(dept)s bo'limiga ko'chirildi.") % {'dept': patient.department})
    return redirect('patient_detail', pk=pk)


# ==================== SHIFOKOR KABINETI ====================

def _get_doctor_profile(user):
    """Shifokor sifatida ishlay oladigan foydalanuvchini qaytaradi."""
    if getattr(user, 'role', None) in ('doctor', 'old'):
        return user
    return None


def _doctor_scope(doctor):
    """Bo'lim mudiri — butun bo'lim + shaxsan biriktirilgan bemorlar (masalan,
    bo'limsiz ambulator bemorlar); terapevt — barcha ro'yxatga olingan bemorlar
    (bo'limidan qat'iy nazar); oddiy shifokor — tasdiqlangan biriktirish YOKI
    xizmat/konsultatsiya orqali biriktirilgan bemorlar."""
    qs = PatientCard.objects.select_related('department', 'attending_doctor', 'department_head')
    if doctor.is_general_practitioner:
        return qs.distinct()
    personal = (
        Q(attending_doctor=doctor, attending_doctor_confirmed=True) |
        Q(consultation_requests__consultants=doctor) |
        Q(diagnostic_assignments__assigned_by=doctor) |
        Q(lab_test_assignments__assigned_by=doctor) |
        Q(treatment_procedures__assigned_by=doctor)
    )
    if doctor.is_head and doctor.department_id:
        return qs.filter(Q(department_id=doctor.department_id) | personal).distinct()
    return qs.filter(personal).distinct()


@login_required
def doctor_dashboard(request):
    if request.user.role != 'doctor' and not request.user.is_superuser:
        return redirect('access_denied')
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        messages.warning(request, _("Sizning shifokor profilingiz hali yaratilmagan. Administratorga murojaat qiling."))
        return render(request, 'patients/doctor/dashboard.html', {'doctor': None})

    scope = _doctor_scope(doctor)
    today = timezone.now().date()

    stats = {
        'total':          scope.count(),
        'inpatient':      scope.filter(visit_type='inpatient').count(),
        'outpatient':     scope.filter(visit_type='ambulatory').count(),
        'railway':        scope.filter(patient_category='railway').count(),
        'paid':           scope.filter(patient_category='paid').count(),
        'non_resident':   scope.filter(patient_category='non_resident').count(),
        'new_today':      scope.filter(created_at__date=today).count(),
        'admitted_today': scope.filter(admission_date__date=today).count(),
        'discharged':     scope.filter(status='completed').count(),
        'active':         scope.exclude(status='completed').count(),
    }

    if doctor.is_head and doctor.department_id:
        stats['doctors_count'] = CustomUser.objects.filter(role__in=('doctor', 'old'), department_id=doctor.department_id, is_active=True).count()
        stats['unassigned'] = scope.filter(visit_type='inpatient').exclude(status='completed').filter(
            Q(attending_doctor__isnull=True) | Q(attending_doctor_confirmed=False)
        ).count()

    # So'nggi 7 kunlik qabul oqimi (diagramma uchun)
    flow_labels, flow_counts = [], []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        flow_labels.append(d.strftime('%d.%m'))
        flow_counts.append(scope.filter(admission_date__date=d).count())

    recent_patients = scope.order_by('-created_at')[:8]
    notifications = DoctorNotification.objects.filter(recipient=request.user).order_by('-created_at')[:8]

    return render(request, 'patients/doctor/dashboard.html', {
        'doctor': doctor,
        'stats': stats,
        'flow_labels': json.dumps(flow_labels),
        'flow_counts': json.dumps(flow_counts),
        'category_data': json.dumps([stats['railway'], stats['paid'], stats['non_resident']]),
        'recent_patients': recent_patients,
        'notifications': notifications,
    })


@login_required
def doctor_patient_list(request, visit_type):
    if request.user.role != 'doctor' and not request.user.is_superuser:
        return redirect('access_denied')
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        messages.warning(request, _("Sizning shifokor profilingiz hali yaratilmagan."))
        return redirect('doctor_dashboard')

    qs = _doctor_scope(doctor).filter(visit_type=visit_type)

    query = request.GET.get('q', '')
    if query:
        qs = qs.filter(
            Q(full_name__icontains=query) |
            Q(medical_record_number__icontains=query) |
            Q(case_sheet_number__icontains=query)
        )

    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)

    sort = request.GET.get('sort', '-admission_date')
    if sort not in ('full_name', '-full_name', 'admission_date', '-admission_date', 'status', '-status'):
        sort = '-admission_date'
    qs = qs.order_by(sort)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'patients/doctor/patient_list.html', {
        'doctor': doctor,
        'page_obj': page_obj,
        'visit_type': visit_type,
        'visit_type_label': 'Statsionar' if visit_type == 'inpatient' else 'Ambulator',
        'query': query,
        'status': status,
        'sort': sort,
        'status_choices': PatientCard.STATUS_CHOICES,
    })


@login_required
def doctor_assign_patients(request):
    """Bo'lim mudiri — yangi statsionar bemorlarni o'z bo'limidagi shifokorlarga biriktiradi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor or not doctor.is_head:
        return redirect('access_denied')

    from django.db.models import Count

    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        doctor_id  = request.POST.get('doctor_id')
        patient = get_object_or_404(PatientCard, pk=patient_id, department_id=doctor.department_id)
        target_doctor = get_object_or_404(CustomUser, pk=doctor_id, role__in=('doctor', 'old'), department_id=doctor.department_id, is_active=True)
        previous_doctor = patient.attending_doctor
        previous_doctor_id = patient.attending_doctor_id
        patient.attending_doctor = target_doctor
        patient.attending_doctor_confirmed = True
        if not patient.department_head_id:
            patient.department_head = doctor
        patient.save()

        if previous_doctor_id and previous_doctor_id != target_doctor.pk:
            # Qayta biriktirish — eski shifokorga xabar, yangi shifokorga xabar
            _notify_doctor(
                previous_doctor, patient,
                f"Bemor boshqa shifokorga o'tkazildi: {patient.full_name} ({patient.medical_record_number})"
            )
            _notify_doctor(
                target_doctor, patient,
                f"Sizga bemor qayta biriktirildi: {patient.full_name} ({patient.medical_record_number})"
            )
            messages.success(request, _("🔄 %(patient)s — %(prev_doctor)s dan %(target_doctor)s ga qayta biriktirildi.") % {'patient': patient.full_name, 'prev_doctor': previous_doctor.full_name if previous_doctor else '—', 'target_doctor': target_doctor.full_name})
        else:
            _notify_doctor(
                target_doctor, patient,
                f"Sizga yangi bemor biriktirildi: {patient.full_name} ({patient.medical_record_number})"
            )
            messages.success(request, _("✅ %(patient)s — %(target_doctor)s ga biriktirildi.") % {'patient': patient.full_name, 'target_doctor': target_doctor.full_name})
        return redirect('doctor_assign_patients')

    pending = PatientCard.objects.filter(
        department_id=doctor.department_id,
        visit_type='inpatient',
    ).exclude(status='completed').select_related('attending_doctor').order_by(
        'attending_doctor_id', '-admission_date'
    )

    dept_doctors = CustomUser.objects.filter(
        role__in=('doctor', 'old'), department_id=doctor.department_id, is_active=True
    ).exclude(pk=doctor.pk).annotate(
        workload=Count('attending_cards', filter=Q(attending_cards__status__in=['registered', 'admitted']))
    ).order_by('first_name')

    history = PatientCard.objects.filter(
        department_id=doctor.department_id,
        attending_doctor__isnull=False,
    ).exclude(attending_doctor=doctor).select_related('attending_doctor').order_by('-updated_at')[:15]

    return render(request, 'patients/doctor/assign.html', {
        'doctor': doctor,
        'pending': pending,
        'dept_doctors': dept_doctors,
        'history': history,
    })


REMINDER_LEAD_MINUTES = 30


def _schedule_recipients(occ):
    """Rejalashtirilgan band turi bo'yicha bildirishnoma oluvchilarni aniqlaydi."""
    if occ.treatment_procedure_id:
        patient = occ.treatment_procedure.patient_card
        if not patient.department_id:
            return CustomUser.objects.none()
        return CustomUser.objects.filter(role__in=['nurse', 'head_nurse'], department_id=patient.department_id, is_active=True)
    if occ.lab_test_assignment_id:
        return CustomUser.objects.filter(role='laborant', is_active=True)
    if occ.diagnostic_assignment_id:
        return CustomUser.objects.filter(role='diagnostician', is_active=True)
    if occ.consultation_request_id:
        user_ids = [c.user_id for c in occ.consultation_request.consultants.all() if c.user_id]
        return CustomUser.objects.filter(pk__in=user_ids)
    return CustomUser.objects.none()


def _send_due_schedule_reminders():
    """Rejalashtirilgan vaqti yaqinlashgan ServiceSchedule bandlari uchun bir martalik eslatma yuboradi.
    Bir nechta foydalanuvchi 30s polling orqali bir vaqtda chaqirsa ham, cache lock orqali
    haqiqiy so'rov faqat ~20 soniyada bir marta bajariladi (to'g'rilik reminder_sent_at'dan keladi)."""
    from django.core.cache import cache
    if not cache.add('schedule_reminder_sweep_lock', True, timeout=20):
        return
    now = timezone.now()
    due = ServiceSchedule.objects.filter(
        status='pending', reminder_sent_at__isnull=True,
        scheduled_at__lte=now + timedelta(minutes=REMINDER_LEAD_MINUTES),
    ).select_related(
        'treatment_procedure__patient_card', 'lab_test_assignment__patient_card',
        'diagnostic_assignment__patient_card', 'consultation_request__patient_card',
    ).prefetch_related('consultation_request__consultants')
    for occ in due:
        patient = occ.patient_card
        if not patient:
            continue
        when = timezone.localtime(occ.scheduled_at).strftime('%d.%m.%Y %H:%M')
        message = f"⏰ {occ.label} vaqti yaqinlashdi: {patient.full_name} — {when}"
        for user in _schedule_recipients(occ):
            _notify_user(user, patient, message)
        occ.reminder_sent_at = now
        occ.save(update_fields=['reminder_sent_at'])


@require_POST
@login_required
def schedule_occurrence_update(request, pk):
    """Hamshira/laborant/diagnost/konsultant — rejalashtirilgan bandning holatini
    (bajarildi/o'tkazib yuborildi/bekor qilindi) belgilaydi. Bu ota-obyektning
    umumiy holatiga (status) yoki hisob-fakturaga ta'sir qilmaydi — mustaqil
    tashrif kuzatuvi qatlami."""
    occ = get_object_or_404(ServiceSchedule, pk=pk)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    def _error(message, status=400):
        if is_ajax:
            return JsonResponse({'success': False, 'error': message}, status=status)
        messages.error(request, message)
        return redirect(next_url)

    def _occ_payload():
        return {
            'success': True,
            'status': occ.status,
            'status_display': occ.get_status_display(),
            'reason': occ.reason,
            'comment': occ.comment,
            'performed_by': occ.performed_by.get_full_name() if occ.performed_by else '',
        }

    doctor = _get_doctor_profile(request.user)
    allowed = request.user.is_superuser or (
        (occ.treatment_procedure_id and request.user.role in ('nurse', 'head_nurse')) or
        (occ.lab_test_assignment_id and request.user.role == 'laborant') or
        (occ.diagnostic_assignment_id and request.user.role == 'diagnostician') or
        (occ.consultation_request_id and doctor and occ.consultation_request.consultants.filter(pk=doctor.pk).exists())
    )
    next_url = request.POST.get('next') or 'nurse_dashboard'
    if not allowed:
        return _error(_("Ruxsat yo'q"), status=403) if is_ajax else redirect('access_denied')
    if occ.status != 'pending':
        return JsonResponse(_occ_payload()) if is_ajax else redirect(next_url)
    new_status = request.POST.get('status')
    if new_status not in dict(ServiceSchedule.STATUS_CHOICES):
        return _error(_("Noto'g'ri holat"))
    occ.status = new_status
    if new_status == 'done':
        occ.completed_at = timezone.now()
        occ.performed_by = request.user
        occ.comment = (request.POST.get('comment') or '').strip()
        occ.save(update_fields=['status', 'completed_at', 'performed_by', 'comment'])
    elif new_status == 'stopped':
        # "To'xtatdim" faqat hamshira plitka UI'sidan keladi va u yerda sabab
        # HTML5 `required` bilan ham majburiy — bu yerda server tomonida ham
        # qat'iy talab qilinadi.
        reason = (request.POST.get('reason') or '').strip()
        if not reason:
            return _error(_("Sabab kiritilishi shart."))
        occ.completed_at = timezone.now()
        occ.performed_by = request.user
        occ.reason = reason
        occ.save(update_fields=['status', 'completed_at', 'performed_by', 'reason'])
    elif new_status == 'cancelled':
        # "cancelled" ham hamshira plitka modalidan (sabab majburiy, frontendda
        # talab qilinadi), ham laborant/diagnost/eski oddiy dropdown'lardan
        # (sababsiz) kelishi mumkin — shu sabab bu yerda sabab ixtiyoriy.
        occ.completed_at = timezone.now()
        occ.performed_by = request.user
        occ.reason = (request.POST.get('reason') or '').strip()
        occ.save(update_fields=['status', 'completed_at', 'performed_by', 'reason'])
    else:
        occ.save(update_fields=['status'])
    return JsonResponse(_occ_payload()) if is_ajax else redirect(next_url)


@login_required
def doctor_notifications_ajax(request):
    _send_due_schedule_reminders()
    base = DoctorNotification.objects.filter(recipient=request.user)
    unread = base.filter(is_read=False).count()
    notifications = base.order_by('-created_at')[:15]
    data = [{
        'id':          n.id,
        'message':     n.message,
        'is_read':     n.is_read,
        'created_at':  n.created_at.strftime('%d.%m.%Y %H:%M'),
        'patient_id':  n.patient_card_id,
    } for n in notifications]
    return JsonResponse({'unread': unread, 'notifications': data})


@require_POST
@login_required
def doctor_notifications_read(request):
    DoctorNotification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'success': True})


def doctor_patient_card(request, pk):
    """Shifokor kabineti — bemor kartasi (8 bo'limli ko'rinish)."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')

    patient = get_object_or_404(
        PatientCard.objects.select_related(
            'department', 'attending_doctor', 'department_head',
            'discharge_conclusion', 'registered_by', 'referral_organization',
            'country', 'region', 'district', 'city'
        ),
        pk=pk
    )

    # Ruxsat: bo'lim mudiri — bo'lim bo'yicha, davolovchi shifokor — o'z bemori,
    # yoki taklif qilingan konsultant — faqat shu bemor bo'yicha
    if not _doctor_card_access(doctor, patient):
        messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
        return redirect('doctor_dashboard')

    from apps.services.models import PatientService, PatientMedicine, ServiceCategory
    from apps.laboratory.models import LabOrderItem

    service_categories = list(
        ServiceCategory.objects.filter(is_active=True).order_by('name').values(
            'id', 'name', 'category_type', 'icon',
        )
    )

    address_parts = filter(None, [
        str(patient.country) if patient.country else '',
        str(patient.region) if patient.region else '',
        str(patient.district) if patient.district else '',
        str(patient.city) if patient.city else '',
        patient.street_address or '',
    ])
    full_address = ', '.join(address_parts) or '—'

    # 2-tab: Shifokor ko'riklari
    examinations = patient.medical_examinations.select_related('created_by').order_by('-examination_datetime', '-created_at')
    initial_exam = getattr(patient, 'initial_examination', None)

    # 3-tab: Tibbiy muolajalar — barcha tayinlangan muolaja/dori, tahlil/laboratoriya va diagnostika
    # tekshiruvlari bitta ro'yxatga birlashtiriladi, har biri o'z holati va natijasi bilan.
    operations = patient.operations.select_related('operation_type').order_by('-operation_date')
    diagnoses = patient.episode_diagnoses.select_related('icd10_code').all()

    treatment_procedures = patient.treatment_procedures.select_related('assigned_by').prefetch_related('schedule_occurrences').order_by('-created_at')
    lab_test_assignments = patient.lab_test_assignments.select_related('assigned_by').prefetch_related('result_logs__performed_by', 'schedule_occurrences').order_by('-created_at')
    lab_order_items = LabOrderItem.objects.filter(order__patient_card=patient).select_related(
        'template', 'result', 'patient_service__service', 'order'
    ).order_by('-created_at')
    diagnostic_assignments = patient.diagnostic_assignments.select_related('assigned_by').prefetch_related('result_logs__performed_by', 'schedule_occurrences').order_by('-created_at')

    def _schedule_summary(occurrences):
        occurrences = list(occurrences)
        if not occurrences:
            return ''
        done_count = sum(1 for o in occurrences if o.status == 'done')
        next_pending = next((o for o in occurrences if o.status == 'pending'), None)
        when = timezone.localtime(next_pending.scheduled_at).strftime('%d.%m %H:%M') if next_pending else ''
        if len(occurrences) == 1:
            return f"Reja: {when}" if when else ''
        suffix = f" — keyingisi {when}" if when else ''
        return f"Reja: {done_count}/{len(occurrences)} bajarildi{suffix}"

    ACTIVITY_BADGE_CLASS = {
        'done': 'success', 'completed': 'success', 'verified': 'success', 'printed': 'success',
        'in_progress': 'info', 'sample_taken': 'primary', 'result_entering': 'info',
        'cancelled': 'secondary', 'rejected': 'secondary', 'recollect': 'secondary',
    }

    assigned_activity = []
    for proc in treatment_procedures:
        detail = proc.dosage or ''
        if proc.quantity > 1:
            detail = f"{detail} ×{proc.quantity}".strip()
        sched = _schedule_summary(proc.schedule_occurrences.all())
        if sched:
            detail = f"{detail} · {sched}".strip(' ·')
        assigned_activity.append({
            'date': proc.created_at,
            'category': _("Dori-darmon"),
            'name': proc.medicine_name,
            'detail': detail,
            'status_display': proc.get_status_display(),
            'badge_class': ACTIVITY_BADGE_CLASS.get(proc.status, 'warning'),
            'view_url': None,
            'print_url': None,
        })
    for a in lab_test_assignments:
        is_done = a.status == 'done'
        detail = a.notes
        sched = _schedule_summary(a.schedule_occurrences.all())
        if sched:
            detail = f"{detail} · {sched}".strip(' ·')
        assigned_activity.append({
            'date': a.created_at,
            'category': _("Laboratoriya"),
            'name': a.test_name,
            'detail': detail,
            'status_display': a.get_status_display(),
            'badge_class': ACTIVITY_BADGE_CLASS.get(a.status, 'warning'),
            'view_url': reverse('lab_test_result_form', args=[a.pk]) if is_done else None,
            'print_url': reverse('lab_test_result_print', args=[a.pk]) if is_done else None,
        })
    for item in lab_order_items:
        result = item.result
        print_url = reverse('lab_result_print', args=[result.pk]) if result else None
        assigned_activity.append({
            'date': item.created_at,
            'category': _("Laboratoriya"),
            'name': item.template.name if item.template else item.patient_service.service.name,
            'detail': '',
            'status_display': item.get_status_display(),
            'badge_class': ACTIVITY_BADGE_CLASS.get(item.status, 'warning'),
            'view_url': print_url,
            'print_url': print_url,
        })
    for a in diagnostic_assignments:
        is_done = a.status == 'done'
        detail = a.notes
        sched = _schedule_summary(a.schedule_occurrences.all())
        if sched:
            detail = f"{detail} · {sched}".strip(' ·')
        assigned_activity.append({
            'date': a.created_at,
            'category': _("Diagnostika"),
            'name': a.get_diagnostic_type_display(),
            'detail': detail,
            'status_display': a.get_status_display(),
            'badge_class': ACTIVITY_BADGE_CLASS.get(a.status, 'warning'),
            'view_url': reverse('diagnostic_result_form', args=[a.pk]) if is_done else None,
            'print_url': reverse('diagnostic_result_print', args=[a.pk]) if is_done else None,
        })
    assigned_activity.sort(key=lambda x: x['date'], reverse=True)

    procedures_badge_count = len(assigned_activity) + operations.count()

    # 6-tab: Konsultatsiyalar — maxsus mutaxassis va konsilium ko'riklari + so'rovlar
    consultations = examinations.filter(examination_type__in=['specialist', 'consilium', 'anesthesia'])
    consultation_requests = patient.consultation_requests.select_related('requested_by').prefetch_related(
        'consultants', 'responses__responded_by', 'schedule_occurrences'
    ).order_by('-created_at')

    # Retsept tab
    prescriptions = patient.prescriptions.select_related('doctor').order_by('-created_at')

    # 7-tab: Davolash yakuni
    discharge_exam = examinations.filter(examination_type='discharge').first()

    # 8-tab: Hisob-faktura
    patient_services = PatientService.objects.filter(patient_card=patient).select_related('service__category', 'ordered_by').order_by('-ordered_at')
    patient_medicines = PatientMedicine.objects.filter(patient_card=patient).select_related('medicine', 'ordered_by').order_by('-ordered_at')
    services_total = sum(s.price * s.quantity for s in patient_services) or 0
    medicines_total = sum(m.total_price for m in patient_medicines) or 0
    grand_total = float(services_total or 0) + float(medicines_total or 0)

    return render(request, 'patients/doctor/patient_card.html', {
        'doctor': doctor,
        'patient': patient,
        'full_address': full_address,
        'examinations': examinations,
        'initial_exam': initial_exam,
        'exam_type_choices': MedicalExamination.EXAM_TYPE_CHOICES,
        'operations': operations,
        'procedures_badge_count': procedures_badge_count,
        'service_categories': service_categories,
        'assigned_activity': assigned_activity,
        'diagnoses': diagnoses,
        'consultations': consultations,
        'consultation_requests': consultation_requests,
        'discharge_exam': discharge_exam,
        'prescriptions': [_prescription_payload(p) for p in prescriptions],
        'patient_services': patient_services,
        'patient_medicines': patient_medicines,
        'services_total': services_total,
        'medicines_total': medicines_total,
        'grand_total': grand_total,
        'ambulatory_consultation': getattr(patient, 'ambulatory_consultation', None),
        'doctors':              CustomUser.objects.filter(role__in=('doctor', 'old'), is_active=True).select_related('department').order_by('department__name', 'first_name'),
        'discharge_conclusions': DischargeConclusion.objects.filter(is_active=True).order_by('name'),
        'outcome_choices':      PatientCard.OUTCOME_CHOICES,
        'today':                timezone.localdate(),
    })

# ==================== AMBULATOR QABUL ====================

@login_required
def ambulatory_consultation_form(request, patient_id):
    """Ambulator bemor uchun qabul natijasini kiritish sahifasi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')

    patient = get_object_or_404(PatientCard, pk=patient_id)
    if not _doctor_card_access(doctor, patient):
        messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
        return redirect('doctor_dashboard')

    consultation, _created = AmbulatoryConsultation.objects.get_or_create(
        patient_card=patient, defaults={'doctor': doctor}
    )

    templates_by_kind = {}
    for kind, _label in DoctorTextTemplate.KIND_CHOICES:
        templates_by_kind[kind] = list(
            DoctorTextTemplate.objects.filter(doctor=doctor, kind=kind).order_by('title')
        )

    from_consultation = ConsultationRequest.objects.filter(
        patient_card=patient, consultants=doctor, status='in_progress'
    ).exists()

    return render(request, 'patients/ambulatory_consultation_form.html', {
        'patient': patient,
        'consultation': consultation,
        'templates_result': templates_by_kind['result'],
        'templates_recommendation': templates_by_kind['recommendation'],
        'templates_conclusion': templates_by_kind['conclusion'],
        'from_consultation': from_consultation,
    })


@login_required
@require_POST
def ambulatory_consultation_save(request, pk):
    """AJAX: ambulator qabul natijasini saqlash / yakunlash."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return JsonResponse({'success': False, 'error': _("Shifokor profili topilmadi")}, status=403)

    consultation = get_object_or_404(AmbulatoryConsultation, pk=pk)
    if not _doctor_card_access(doctor, consultation.patient_card):
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)

    try:
        payload = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)

    consultation.result = payload.get('result', '')
    consultation.recommendation = payload.get('recommendation', '')
    consultation.conclusion = payload.get('conclusion', '')

    if payload.get('finish'):
        consultation.status = 'completed'
        consultation.finished_at = timezone.now()

    consultation.save()

    if payload.get('finish'):
        if consultation.patient_card.status != 'completed':
            consultation.patient_card.status = 'completed'
            consultation.patient_card.save(update_fields=['status'])

        ConsultationRequest.objects.filter(
            patient_card=consultation.patient_card, consultants=doctor, status='in_progress'
        ).update(status='done')

    return JsonResponse({'success': True, 'status': consultation.status})


@login_required
def ambulatory_consultation_print(request, pk):
    """Ambulator qabul natijasini chop etish uchun ko'rinish."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')

    consultation = get_object_or_404(
        AmbulatoryConsultation.objects.select_related('patient_card', 'doctor'), pk=pk
    )
    if not _doctor_card_access(doctor, consultation.patient_card):
        messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
        return redirect('doctor_dashboard')

    import base64
    from django.conf import settings as dj_settings
    import os as _os

    header_b64 = ''
    for header_path in [
        _os.path.join(dj_settings.STATIC_ROOT, 'img', 'hospital_header.png'),
        _os.path.join(dj_settings.BASE_DIR, 'static', 'img', 'hospital_header.png'),
    ]:
        if _os.path.exists(header_path):
            with open(header_path, 'rb') as f:
                header_b64 = base64.b64encode(f.read()).decode()
            break

    return render(request, 'patients/ambulatory_consultation_print.html', {
        'consultation': consultation,
        'patient': consultation.patient_card,
        'header_b64': header_b64,
        'print_date': timezone.now(),
    })


@login_required
@require_POST
def doctor_template_create(request):
    """AJAX: shifokor o'zi uchun yangi matn shabloni qo'shadi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return JsonResponse({'success': False, 'error': _("Shifokor profili topilmadi")}, status=403)

    try:
        payload = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)

    kind = payload.get('kind', '')
    title = (payload.get('title') or '').strip()
    body = (payload.get('body') or '').strip()

    valid_kinds = dict(DoctorTextTemplate.KIND_CHOICES)
    if kind not in valid_kinds:
        return JsonResponse({'success': False, 'error': _("Noto'g'ri shablon turi")}, status=400)
    if not title or not body:
        return JsonResponse({'success': False, 'error': _("Nomi va matnini kiriting")}, status=400)

    tmpl = DoctorTextTemplate.objects.create(doctor=doctor, kind=kind, title=title, body=body)
    return JsonResponse({'success': True, 'id': tmpl.pk, 'title': tmpl.title, 'body': tmpl.body})


# ==================== TIBBIY MUOLAJALAR — TAYINLASH (shifokor) ====================

@require_POST
@login_required
def procedure_assign(request, pk):
    """Shifokor bemorga muolaja tayinlaydi — bo'lim hamshiralariga avtomatik xabar boradi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')

    patient = get_object_or_404(PatientCard, pk=pk)
    if doctor.is_head:
        if patient.department_id != doctor.department_id:
            return redirect('access_denied')
    else:
        if patient.attending_doctor_id != doctor.pk:
            return redirect('access_denied')

    medicine_name = request.POST.get('medicine_name', '').strip()
    if not medicine_name:
        messages.error(request, _("Dori vositasi nomini kiriting."))
        return redirect('doctor_patient_card', pk=pk)

    try:
        quantity = int(request.POST.get('quantity') or 1)
    except ValueError:
        quantity = 1

    procedure = TreatmentProcedure.objects.create(
        patient_card=patient,
        assigned_by=doctor,
        medicine_name=medicine_name,
        dosage=request.POST.get('dosage', '').strip(),
        quantity=quantity,
        schedule_note=request.POST.get('schedule_note', '').strip(),
        medicine_source=request.POST.get('medicine_source') or 'clinic_stock',
        notes=request.POST.get('notes', '').strip(),
    )

    # Bo'lim hamshiralariga bildirishnoma
    if patient.department_id:
        nurses = CustomUser.objects.filter(role__in=['nurse', 'head_nurse'], department_id=patient.department_id, is_active=True)
        for nurse in nurses:
            _notify_user(nurse, patient, f"Yangi muolaja tayinlandi: {patient.full_name} — {procedure.medicine_name}")

    messages.success(request, _("✅ Muolaja tayinlandi: %(name)s") % {'name': procedure.medicine_name})
    return redirect('doctor_patient_card', pk=pk)


# ==================== HAMSHIRA KABINETI ====================

NURSE_OVERDUE_HOURS = 4


@login_required
def nurse_dashboard(request):
    if request.user.role not in ('nurse', 'head_nurse') and not request.user.is_superuser:
        return redirect('access_denied')

    is_head_nurse = request.user.role == 'head_nurse'

    view_mode = request.GET.get('view', 'service')
    if view_mode not in ('service', 'patient'):
        view_mode = 'service'

    qs = TreatmentProcedure.objects.select_related(
        'patient_card', 'assigned_by'
    ).prefetch_related('execution_logs__performed_by', 'schedule_occurrences')

    if request.user.department_id:
        qs = qs.filter(patient_card__department_id=request.user.department_id)

    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)
    else:
        qs = qs.exclude(status__in=['done', 'cancelled'])

    query = request.GET.get('q', '')
    if query:
        qs = qs.filter(
            Q(patient_card__full_name__icontains=query) |
            Q(medicine_name__icontains=query)
        )

    qs = qs.order_by('-created_at')

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    notifications = DoctorNotification.objects.filter(recipient=request.user).order_by('-created_at')[:8]

    try:
        highlight_patient_id = int(request.GET.get('patient') or 0) or None
    except ValueError:
        highlight_patient_id = None

    context = {
        'page_obj': page_obj,
        'status': status,
        'query': query,
        'status_choices': TreatmentProcedure.STATUS_CHOICES,
        'notifications': notifications,
        'is_head_nurse': is_head_nurse,
        'view_mode': view_mode,
        'highlight_patient_id': highlight_patient_id,
        'pending_count': TreatmentProcedure.objects.filter(
            patient_card__department_id=request.user.department_id, status='assigned'
        ).count() if request.user.department_id else 0,
    }

    if view_mode == 'patient':
        date_str = request.GET.get('date', '')
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localdate()
        except ValueError:
            selected_date = timezone.localdate()

        dept_patients = PatientCard.objects.filter(status='admitted')
        if request.user.department_id:
            dept_patients = dept_patients.filter(department_id=request.user.department_id)
        if query:
            dept_patients = dept_patients.filter(full_name__icontains=query)
        dept_patients = list(dept_patients.order_by('full_name'))

        occ_qs = ServiceSchedule.objects.filter(
            Q(treatment_procedure__patient_card__in=dept_patients) |
            Q(lab_test_assignment__patient_card__in=dept_patients) |
            Q(diagnostic_assignment__patient_card__in=dept_patients) |
            Q(consultation_request__patient_card__in=dept_patients),
            scheduled_at__date__gte=selected_date,
        ).select_related(
            'treatment_procedure__patient_card', 'lab_test_assignment__patient_card',
            'diagnostic_assignment__patient_card', 'consultation_request__patient_card',
        ).order_by('scheduled_at')

        by_patient = {}
        for occ in occ_qs:
            patient = occ.patient_card
            if patient:
                by_patient.setdefault(patient.pk, []).append(occ)

        patient_rows = []
        for p in dept_patients:
            occs = by_patient.get(p.pk, [])
            days = {}
            for occ in occs:
                d = timezone.localtime(occ.scheduled_at).date()
                days.setdefault(d, []).append(occ)
            day_groups = [{'date': d, 'occurrences': days[d]} for d in sorted(days.keys())]
            patient_rows.append({'patient': p, 'day_groups': day_groups, 'total': len(occs)})

        context.update({
            'selected_date': selected_date,
            'patient_rows': patient_rows,
        })

    if is_head_nurse and request.user.department_id:
        today_start = timezone.localtime().replace(hour=0, minute=0, second=0, microsecond=0)
        overdue_cutoff = timezone.now() - timedelta(hours=NURSE_OVERDUE_HOURS)

        dept_nurses = CustomUser.objects.filter(
            role__in=['nurse', 'head_nurse'], department_id=request.user.department_id, is_active=True,
        ).order_by('last_name', 'first_name')

        nurse_stats = []
        for n in dept_nurses:
            nurse_stats.append({
                'user': n,
                'done_today': ProcedureExecutionLog.objects.filter(
                    performed_by=n, performed_at__gte=today_start,
                ).count(),
                'done_total': ProcedureExecutionLog.objects.filter(performed_by=n).count(),
            })

        overdue_procedures = TreatmentProcedure.objects.select_related('patient_card', 'assigned_by').filter(
            patient_card__department_id=request.user.department_id,
            status='assigned',
            created_at__lte=overdue_cutoff,
        ).order_by('created_at')

        context.update({
            'nurse_stats': nurse_stats,
            'overdue_procedures': overdue_procedures,
            'overdue_hours': NURSE_OVERDUE_HOURS,
        })

    return render(request, 'patients/nurse/dashboard.html', context)


@require_POST
@login_required
def procedure_log_execution(request, pk):
    """Hamshira — muolajani bajarganini qayd qiladi (alohida log yozuvi)."""
    if request.user.role not in ('nurse', 'head_nurse') and not request.user.is_superuser:
        return redirect('access_denied')

    procedure = get_object_or_404(TreatmentProcedure, pk=pk)
    ProcedureExecutionLog.objects.create(
        procedure=procedure,
        performed_by=request.user,
        comment=request.POST.get('comment', '').strip(),
    )
    if procedure.status == 'assigned':
        procedure.status = 'in_progress'
        procedure.save(update_fields=['status'])

    if procedure.assigned_by:
        _notify_doctor(
            procedure.assigned_by, procedure.patient_card,
            f"Muolaja bajarildi: {procedure.patient_card.full_name} — {procedure.medicine_name} ({request.user.get_full_name()})"
        )

    messages.success(request, _("✅ Bajarilish qayd etildi."))
    return redirect('nurse_dashboard')


@require_POST
@login_required
def procedure_update_status(request, pk):
    """Hamshira — muolaja holatini o'zgartiradi (Bajarilgan / Bekor qilingan)."""
    if request.user.role not in ('nurse', 'head_nurse') and not request.user.is_superuser:
        return redirect('access_denied')

    procedure = get_object_or_404(TreatmentProcedure, pk=pk)
    new_status = request.POST.get('status')
    if new_status in dict(TreatmentProcedure.STATUS_CHOICES):
        procedure.status = new_status
        procedure.save(update_fields=['status'])
        _sync_billing_status(procedure)
        if new_status == 'cancelled':
            procedure.schedule_occurrences.filter(status='pending').update(status='cancelled')
        messages.success(request, _("Holat yangilandi: %(status)s") % {'status': procedure.get_status_display()})
    return redirect('nurse_dashboard')


def _doctor_card_access(doctor, patient):
    """Shifokor ushbu bemor kartasini ko'rishi mumkinmi."""
    if doctor.is_general_practitioner:
        return True
    if doctor.role == 'old' and patient.department_id == doctor.department_id:
        return True
    if doctor.is_head and patient.department_id == doctor.department_id:
        return True
    if patient.attending_doctor_id == doctor.pk and patient.attending_doctor_confirmed:
        return True
    if ConsultationRequest.objects.filter(patient_card=patient, consultants=doctor).exists():
        return True
    if DiagnosticAssignment.objects.filter(patient_card=patient, assigned_by=doctor).exists():
        return True
    if LabTestAssignment.objects.filter(patient_card=patient, assigned_by=doctor).exists():
        return True
    if TreatmentProcedure.objects.filter(patient_card=patient, assigned_by=doctor).exists():
        return True
    return False


# ==================== RETSEPT (shifokor -> bemor) ====================

def _to_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _to_date(val):
    if not val:
        return None
    from datetime import date
    try:
        return date.fromisoformat(val)
    except ValueError:
        return None


def _prescription_summary(p):
    bits = []
    if p.dose:
        bits.append(p.dose)
    if p.frequency_num:
        unit_label = dict(Prescription.FREQUENCY_UNIT_CHOICES).get(p.frequency_unit, '')
        bits.append(f"{unit_label} {p.frequency_num} marta".strip())
    if p.duration_days:
        bits.append(_("%(n)s kun") % {'n': p.duration_days})
    return f"{p.drug_name} — {', '.join(bits)}" if bits else p.drug_name


def _prescription_payload(p):
    return {
        'id': p.pk,
        'drug_name': p.drug_name,
        'dosage_form': p.dosage_form,
        'dose': p.dose,
        'frequency_num': p.frequency_num,
        'frequency_unit': p.frequency_unit,
        'single_dose': p.single_dose,
        'method': p.method,
        'duration_days': p.duration_days,
        'intake_time': p.intake_time,
        'date_start': p.date_start.isoformat() if p.date_start else '',
        'date_end': p.date_end.isoformat() if p.date_end else '',
        'total_quantity': p.total_quantity,
        'note': p.note,
        'created_at': timezone.localtime(p.created_at).strftime('%d.%m.%Y %H:%M'),
        'summary': _prescription_summary(p),
    }


def _prescription_body(request):
    try:
        body = json.loads(request.body)
    except (ValueError, TypeError):
        return None
    return body


def _apply_prescription_fields(p, body):
    p.drug_name      = (body.get('drug_name') or '').strip()
    p.dosage_form    = (body.get('dosage_form') or '').strip()
    p.dose           = (body.get('dose') or '').strip()
    p.frequency_num  = _to_int(body.get('frequency_num'))
    p.frequency_unit = body.get('frequency_unit') or 'День'
    p.single_dose    = (body.get('single_dose') or '').strip()
    p.method         = (body.get('method') or '').strip()
    p.duration_days  = _to_int(body.get('duration_days'))
    p.intake_time    = (body.get('intake_time') or '').strip()
    p.date_start     = _to_date(body.get('date_start'))
    p.date_end       = _to_date(body.get('date_end'))
    p.total_quantity = (body.get('total_quantity') or '').strip()
    p.note           = (body.get('note') or '').strip()


def _prescription_occurrence_times(p):
    """Retsept asosida hamshira uchun kunlik bandlar vaqtini hisoblaydi.
    Faqat 'Kuniga' (День) birligi uchun kuniga necha marta dozani teng
    taqsimlaydi; Haftada/Oyda uchun soddalashtirib kuniga 1 marta olinadi."""
    if not p.date_start:
        return []
    from datetime import time as _time
    duration = min(p.duration_days or 1, 60)
    doses_per_day = p.frequency_num or 1
    if p.frequency_unit != 'День' or doses_per_day < 1:
        doses_per_day = 1
    try:
        hour, minute = (int(x) for x in p.intake_time.split(':'))
    except (ValueError, AttributeError):
        hour, minute = 9, 0
    step_minutes = (24 * 60) // doses_per_day if doses_per_day > 1 else 0
    tz = timezone.get_current_timezone()
    times = []
    for day_offset in range(duration):
        day = p.date_start + timedelta(days=day_offset)
        for dose_idx in range(doses_per_day):
            total_minutes = (hour * 60 + minute + dose_idx * step_minutes) % (24 * 60)
            h, m = divmod(total_minutes, 60)
            times.append(timezone.make_aware(datetime.combine(day, _time(h, m)), tz))
    return sorted(times)


def _sync_prescription_schedule(p, patient, doctor):
    """Retsept saqlanganda/yangilanganda bog'langan TreatmentProcedure va
    bajarilmagan ServiceSchedule bandlarini retsept maydonlariga moslab
    qayta hosil qiladi — shunda hamshira buni 'Bajariladigan ishlar'
    ro'yxatida ko'radi."""
    occurrence_times = _prescription_occurrence_times(p)
    proc = p.treatment_procedure
    is_new = proc is None
    if is_new:
        if not occurrence_times:
            return
        proc = TreatmentProcedure.objects.create(
            patient_card=patient, assigned_by=doctor,
            medicine_name=p.drug_name, dosage=p.dose,
            quantity=len(occurrence_times), schedule_note=p.method, notes=p.note,
        )
        p.treatment_procedure = proc
        p.save(update_fields=['treatment_procedure'])
    else:
        proc.medicine_name = p.drug_name
        proc.dosage = p.dose
        proc.quantity = len(occurrence_times) or proc.quantity
        proc.schedule_note = p.method
        proc.notes = p.note
        proc.save(update_fields=['medicine_name', 'dosage', 'quantity', 'schedule_note', 'notes'])
        proc.schedule_occurrences.filter(status='pending').delete()

    _create_schedule_occurrences(occurrence_times, treatment_procedure=proc)

    if is_new and patient.department_id:
        for nurse in CustomUser.objects.filter(role__in=['nurse', 'head_nurse'], department_id=patient.department_id, is_active=True):
            _notify_user(nurse, patient, f"Retsept bo'yicha yangi muolaja: {patient.full_name} — {p.drug_name}")


@login_required
@require_POST
def prescription_add(request, pk):
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return JsonResponse({'success': False, 'error': _("Shifokor profili topilmadi")}, status=403)
    patient = get_object_or_404(PatientCard, pk=pk)
    if not _doctor_card_access(doctor, patient):
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)
    body = _prescription_body(request)
    if body is None:
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)
    if not (body.get('drug_name') or '').strip():
        return JsonResponse({'success': False, 'error': _("Dori nomini kiriting")}, status=400)

    p = Prescription(patient_card=patient, doctor=doctor)
    _apply_prescription_fields(p, body)
    p.save()
    _sync_prescription_schedule(p, patient, doctor)
    return JsonResponse({'success': True, 'prescription': _prescription_payload(p)})


@login_required
@require_POST
def prescription_update(request, pk):
    p = get_object_or_404(Prescription, pk=pk)
    doctor = _get_doctor_profile(request.user)
    if not doctor or not _doctor_card_access(doctor, p.patient_card):
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)
    body = _prescription_body(request)
    if body is None:
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)
    if not (body.get('drug_name') or '').strip():
        return JsonResponse({'success': False, 'error': _("Dori nomini kiriting")}, status=400)

    _apply_prescription_fields(p, body)
    p.save()
    _sync_prescription_schedule(p, p.patient_card, p.doctor or doctor)
    return JsonResponse({'success': True, 'prescription': _prescription_payload(p)})


@login_required
@require_POST
def prescription_delete(request, pk):
    p = get_object_or_404(Prescription, pk=pk)
    doctor = _get_doctor_profile(request.user)
    if not doctor or not _doctor_card_access(doctor, p.patient_card):
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)
    if p.treatment_procedure_id:
        p.treatment_procedure.delete()
    p.delete()
    return JsonResponse({'success': True})


@login_required
def prescription_print(request, pk):
    p = get_object_or_404(Prescription.objects.select_related('patient_card', 'doctor'), pk=pk)
    doctor = _get_doctor_profile(request.user)
    if not doctor or not _doctor_card_access(doctor, p.patient_card):
        messages.error(request, _("Siz bu bemorni ko'rishga ruxsatingiz yo'q."))
        return redirect('doctor_dashboard')

    import base64
    from django.conf import settings as dj_settings
    import os as _os

    header_b64 = ''
    for header_path in [
        _os.path.join(dj_settings.STATIC_ROOT, 'img', 'hospital_header.png'),
        _os.path.join(dj_settings.BASE_DIR, 'static', 'img', 'hospital_header.png'),
    ]:
        if _os.path.exists(header_path):
            with open(header_path, 'rb') as f:
                header_b64 = base64.b64encode(f.read()).decode()
            break

    return render(request, 'patients/prescription_print.html', {
        'p': p,
        'patient': p.patient_card,
        'header_b64': header_b64,
        'print_date': timezone.now(),
    })


# ==================== YOZUV QO'SHISH — XIZMATLAR KATALOGIDAN PAKET TAYINLASH ====================

_DIAGNOSTIC_TYPE_KEYWORDS = [
    ('rentgen',     ['rentgen']),
    ('uzi',         ['uzi', 'ultratovush']),
    ('mrt',         ['mrt']),
    ('ekg',         ['kardiologiya', 'ekg']),
    ('kt',          ['kompyuter tomografiya']),
    ('endoskopiya', ['endoskopiya']),
]


def _guess_diagnostic_type(category_name):
    name_lower = (category_name or '').lower()
    for dtype, keywords in _DIAGNOSTIC_TYPE_KEYWORDS:
        if any(kw in name_lower for kw in keywords):
            return dtype
    return 'other'


_BILLING_STATUS_MAP = {
    'assigned': 'ordered', 'sample_taken': 'in_progress', 'in_progress': 'in_progress',
    'done': 'completed', 'cancelled': 'cancelled',
}


def _sync_billing_status(obj):
    """Tayinlangan muolaja/tahlil/diagnostika/konsultatsiya holati o'zgarganda,
    unga bog'langan hisob-faktura (PatientService) yozuvi holatini ham moslashtiradi."""
    ps = getattr(obj, 'patient_service', None)
    if not ps:
        return
    new_status = _BILLING_STATUS_MAP.get(obj.status)
    if new_status and ps.status != new_status:
        ps.status = new_status
        ps.save(update_fields=['status'])


def _create_billing_record(patient, service, doctor, quantity, notes):
    from apps.services.models import PatientService
    return PatientService.objects.create(
        patient_card=patient, service=service, quantity=quantity,
        price=service.price_for_patient(patient.patient_category or 'railway'),
        patient_category_at_order=patient.patient_category or 'railway',
        ordered_by=doctor, notes=notes,
    )


def _parse_schedule_occurrences(item):
    """Savatcha elementidagi sana/vaqt/necha-kun maydonlaridan ServiceSchedule
    uchun rejalashtirilgan vaqtlar ro'yxatini hosil qiladi (kun sayin bittadan)."""
    sched_date_str = (item.get('scheduled_date') or '').strip()
    if not sched_date_str:
        return []
    sched_time_str = (item.get('scheduled_time') or '09:00').strip() or '09:00'
    try:
        base_date = datetime.strptime(sched_date_str, '%Y-%m-%d').date()
        base_time = datetime.strptime(sched_time_str, '%H:%M').time()
    except ValueError:
        return []
    try:
        repeat_days = max(1, min(60, int(item.get('repeat_days') or 1)))
    except (TypeError, ValueError):
        repeat_days = 1
    tz = timezone.get_current_timezone()
    return [
        timezone.make_aware(datetime.combine(base_date + timedelta(days=i), base_time), tz)
        for i in range(repeat_days)
    ]


def _create_schedule_occurrences(occurrence_times, **target_kwargs):
    if occurrence_times:
        ServiceSchedule.objects.bulk_create([
            ServiceSchedule(scheduled_at=dt, **target_kwargs) for dt in occurrence_times
        ])


@require_POST
@login_required
def assign_services(request, pk):
    """Shifokor — xizmatlar katalogidan (istalgan kategoriya) bir nechta yozuvni bir vaqtda tayinlaydi.
    Har bir xizmatga ixtiyoriy ravishda aniq bir biriktirilgan shifokor va izoh ko'rsatilishi mumkin."""
    from apps.services.models import Service

    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return JsonResponse({'success': False, 'error': _("Shifokor profili topilmadi")}, status=403)
    patient = get_object_or_404(PatientCard, pk=pk)
    if not _doctor_card_access(doctor, patient):
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)

    try:
        payload = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)

    items = payload.get('items') or []
    if not items:
        return JsonResponse({'success': False, 'error': _("Kamida bitta xizmat tanlang")}, status=400)

    created_lab, created_diag, created_proc, created_consult = [], [], [], []

    for item in items:
        service = Service.objects.filter(
            pk=item.get('service_id'), is_active=True
        ).select_related('category').first()
        if not service:
            continue
        notes = (item.get('notes') or '').strip()
        category_type = service.category.category_type

        chosen_doctor = None
        doctor_id = item.get('doctor_id')
        if doctor_id:
            chosen_doctor = service.assigned_doctors.filter(pk=doctor_id, is_active=True).first()

        occurrence_times = _parse_schedule_occurrences(item)
        quantity = len(occurrence_times) if occurrence_times else 1
        billing = _create_billing_record(patient, service, doctor, quantity, notes)

        if category_type == 'consultation':
            consultation = ConsultationRequest.objects.create(
                patient_card=patient, requested_by=doctor, service=service, reason=notes,
                patient_service=billing,
            )
            if chosen_doctor:
                consultation.consultants.set([chosen_doctor])
                _notify_doctor(chosen_doctor, patient, f"Sizga konsultatsiya so'rovi yuborildi: {patient.full_name} — {service.name}")
            _create_schedule_occurrences(occurrence_times, consultation_request=consultation)
            created_consult.append(consultation)
            continue

        if category_type == 'lab':
            obj = LabTestAssignment.objects.create(
                patient_card=patient, assigned_by=doctor, service=service,
                test_name=service.name, notes=notes, patient_service=billing,
            )
            _create_schedule_occurrences(occurrence_times, lab_test_assignment=obj)
            created_lab.append(obj)
        elif category_type == 'radiology':
            obj = DiagnosticAssignment.objects.create(
                patient_card=patient, assigned_by=doctor, service=service,
                diagnostic_type=_guess_diagnostic_type(service.category.name), notes=notes,
                patient_service=billing,
            )
            _create_schedule_occurrences(occurrence_times, diagnostic_assignment=obj)
            created_diag.append(obj)
        else:
            obj = TreatmentProcedure.objects.create(
                patient_card=patient, assigned_by=doctor, service=service,
                medicine_name=service.name, quantity=quantity, notes=notes, patient_service=billing,
            )
            _create_schedule_occurrences(occurrence_times, treatment_procedure=obj)
            created_proc.append(obj)

        if chosen_doctor:
            extra = f" — {notes}" if notes else ''
            _notify_doctor(chosen_doctor, patient, f"Sizga {service.name} bo'yicha bemor biriktirildi: {patient.full_name}{extra}")

    if created_lab:
        for laborant in CustomUser.objects.filter(role='laborant', is_active=True):
            _notify_user(laborant, patient, f"Yangi tahlil(lar) tayinlandi: {patient.full_name} ({len(created_lab)} ta)")
    if created_diag:
        for d in CustomUser.objects.filter(role='diagnostician', is_active=True):
            _notify_user(d, patient, f"Yangi diagnostika tekshiruv(lar)i tayinlandi: {patient.full_name} ({len(created_diag)} ta)")
    if created_proc and patient.department_id:
        for nurse in CustomUser.objects.filter(role__in=['nurse', 'head_nurse'], department_id=patient.department_id, is_active=True):
            _notify_user(nurse, patient, f"Yangi muolaja(lar) tayinlandi: {patient.full_name} ({len(created_proc)} ta)")
    total = len(created_lab) + len(created_diag) + len(created_proc) + len(created_consult)
    if total == 0:
        return JsonResponse({'success': False, 'error': _("Hech narsa tayinlanmadi")}, status=400)

    return JsonResponse({
        'success': True,
        'lab_count': len(created_lab),
        'diagnostic_count': len(created_diag),
        'procedure_count': len(created_proc),
        'consultation_count': len(created_consult),
    })


# ==================== LABORATORIYA — SHIFOKOR TAYINLASH VA LABORANT NAVBATI ====================

@require_POST
@login_required
def lab_test_assign(request, pk):
    """Shifokor — bemorga laboratoriya tahlilini tayinlaydi va laborantlarga xabar yuboradi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    patient = get_object_or_404(PatientCard, pk=pk)
    if not _doctor_card_access(doctor, patient):
        return redirect('access_denied')

    test_name = request.POST.get('test_name', '').strip()
    if not test_name:
        messages.error(request, _("Tahlil nomini kiriting."))
        return redirect('doctor_patient_card', pk=pk)

    assignment = LabTestAssignment.objects.create(
        patient_card=patient, assigned_by=doctor, test_name=test_name,
        notes=request.POST.get('notes', '').strip(),
    )
    for laborant in CustomUser.objects.filter(role='laborant', is_active=True):
        _notify_user(laborant, patient, f"Yangi tahlil tayinlandi: {patient.full_name} — {assignment.test_name}")
    messages.success(request, _("✅ Tahlil tayinlandi: %(name)s") % {'name': assignment.test_name})
    return redirect('doctor_patient_card', pk=pk)


@login_required
def lab_assignment_queue(request):
    """Laborant — shifokorlar tayinlagan tahlillar navbati."""
    if request.user.role != 'laborant' and not request.user.is_superuser:
        return redirect('access_denied')
    qs = LabTestAssignment.objects.select_related('patient_card', 'assigned_by').prefetch_related('schedule_occurrences')
    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)
    else:
        qs = qs.exclude(status__in=['done', 'cancelled'])
    query = request.GET.get('q', '')
    if query:
        qs = qs.filter(Q(patient_card__full_name__icontains=query) | Q(test_name__icontains=query))
    qs = qs.order_by('-created_at')
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    notifications = DoctorNotification.objects.filter(recipient=request.user).order_by('-created_at')[:8]
    return render(request, 'patients/laborant/queue.html', {
        'page_obj': page_obj, 'status': status, 'query': query,
        'status_choices': LabTestAssignment.STATUS_CHOICES, 'notifications': notifications,
        'pending_count': LabTestAssignment.objects.filter(status='assigned').count(),
    })


@login_required
def lab_test_result_form(request, pk):
    """Laborant — tahlil natijasini kiritish/ko'rish sahifasi (tayinlagan shifokor ham faqat ko'rish uchun kira oladi)."""
    assignment = get_object_or_404(LabTestAssignment.objects.select_related('patient_card', 'assigned_by'), pk=pk)
    is_laborant = request.user.role == 'laborant' or request.user.is_superuser
    doctor_viewer = _get_doctor_profile(request.user)
    if not is_laborant and not (doctor_viewer and _doctor_card_access(doctor_viewer, assignment.patient_card)):
        return redirect('access_denied')

    if is_laborant and assignment.status in ('assigned', 'sample_taken'):
        assignment.status = 'in_progress'
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)

    result_log, _created = LabTestResultLog.objects.get_or_create(
        assignment=assignment, defaults={'performed_by': request.user}
    )

    return render(request, 'patients/laborant/result_form.html', {
        'assignment': assignment,
        'patient': assignment.patient_card,
        'result_log': result_log,
    })


@require_POST
@login_required
def lab_test_result_save(request, pk):
    """AJAX — laborant natijani saqlaydi/yakunlaydi."""
    if request.user.role != 'laborant' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)
    assignment = get_object_or_404(LabTestAssignment, pk=pk)

    try:
        payload = json.loads(request.body)
    except ValueError:
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)

    result_log, _created = LabTestResultLog.objects.get_or_create(
        assignment=assignment, defaults={'performed_by': request.user}
    )
    result_log.result_text = payload.get('result_text', '').strip()
    result_log.recommendation = payload.get('recommendation', '').strip()
    result_log.comment = payload.get('comment', '').strip()
    result_log.performed_by = request.user
    result_log.save(update_fields=['result_text', 'recommendation', 'comment', 'performed_by'])

    if payload.get('finish'):
        assignment.status = 'done'
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)
        if assignment.assigned_by:
            _notify_doctor(assignment.assigned_by, assignment.patient_card, f"Tahlil natijasi kiritildi: {assignment.patient_card.full_name} — {assignment.test_name} ({request.user.get_full_name()})")

    return JsonResponse({'success': True, 'status': assignment.status})


@login_required
def lab_test_result_print(request, pk):
    """Oddiy laboratoriya tahlili natijasini chop etish uchun toza (shapkali) ko'rinish."""
    assignment = get_object_or_404(LabTestAssignment.objects.select_related('patient_card', 'assigned_by'), pk=pk)
    is_laborant = request.user.role == 'laborant' or request.user.is_superuser
    doctor = _get_doctor_profile(request.user)
    if not is_laborant and not (doctor and _doctor_card_access(doctor, assignment.patient_card)):
        return redirect('access_denied')
    result_log = assignment.result_logs.first()

    import base64
    from django.conf import settings as dj_settings
    import os as _os

    header_b64 = ''
    for header_path in [
        _os.path.join(dj_settings.STATIC_ROOT, 'img', 'hospital_header.png'),
        _os.path.join(dj_settings.BASE_DIR, 'static', 'img', 'hospital_header.png'),
    ]:
        if _os.path.exists(header_path):
            with open(header_path, 'rb') as f:
                header_b64 = base64.b64encode(f.read()).decode()
            break

    return render(request, 'patients/laborant/result_print.html', {
        'assignment': assignment,
        'patient': assignment.patient_card,
        'result_log': result_log,
        'header_b64': header_b64,
        'print_date': timezone.now(),
    })


@require_POST
@login_required
def lab_test_update_status(request, pk):
    """Laborant — tahlil holatini o'zgartiradi."""
    if request.user.role != 'laborant' and not request.user.is_superuser:
        return redirect('access_denied')
    assignment = get_object_or_404(LabTestAssignment, pk=pk)
    new_status = request.POST.get('status')
    if new_status in dict(LabTestAssignment.STATUS_CHOICES):
        assignment.status = new_status
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)
        if new_status == 'cancelled':
            assignment.schedule_occurrences.filter(status='pending').update(status='cancelled')
        messages.success(request, _("Holat yangilandi: %(status)s") % {'status': assignment.get_status_display()})
    return redirect('lab_assignment_queue')


# ==================== DIAGNOSTIKA — SHIFOKOR TAYINLASH VA DIAGNOST NAVBATI ====================

@require_POST
@login_required
def diagnostic_assign(request, pk):
    """Shifokor — bemorga diagnostika tekshiruvini tayinlaydi va diagnostlarga xabar yuboradi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    patient = get_object_or_404(PatientCard, pk=pk)
    if not _doctor_card_access(doctor, patient):
        return redirect('access_denied')

    diagnostic_type = request.POST.get('diagnostic_type')
    if diagnostic_type not in dict(DiagnosticAssignment.TYPE_CHOICES):
        messages.error(request, _("Tekshiruv turini tanlang."))
        return redirect('doctor_patient_card', pk=pk)

    assignment = DiagnosticAssignment.objects.create(
        patient_card=patient, assigned_by=doctor, diagnostic_type=diagnostic_type,
        notes=request.POST.get('notes', '').strip(),
    )
    for d in CustomUser.objects.filter(role='diagnostician', is_active=True):
        _notify_user(d, patient, f"Yangi diagnostika tayinlandi: {patient.full_name} — {assignment.get_diagnostic_type_display()}")
    messages.success(request, _("✅ Diagnostika tayinlandi: %(name)s") % {'name': assignment.get_diagnostic_type_display()})
    return redirect('doctor_patient_card', pk=pk)


@login_required
def diagnostic_queue(request):
    """Diagnost — shifokorlar tayinlagan diagnostika tekshiruvlari navbati."""
    if request.user.role != 'diagnostician' and not request.user.is_superuser:
        return redirect('access_denied')
    qs = DiagnosticAssignment.objects.select_related('patient_card', 'assigned_by').prefetch_related('schedule_occurrences')
    status = request.GET.get('status', '')
    if status:
        qs = qs.filter(status=status)
    else:
        qs = qs.exclude(status__in=['done', 'cancelled'])
    query = request.GET.get('q', '')
    if query:
        qs = qs.filter(Q(patient_card__full_name__icontains=query) | Q(diagnostic_type__icontains=query))
    qs = qs.order_by('-created_at')
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    notifications = DoctorNotification.objects.filter(recipient=request.user).order_by('-created_at')[:8]
    return render(request, 'patients/diagnostician/queue.html', {
        'page_obj': page_obj, 'status': status, 'query': query,
        'status_choices': DiagnosticAssignment.STATUS_CHOICES, 'notifications': notifications,
        'pending_count': DiagnosticAssignment.objects.filter(status='assigned').count(),
    })


@login_required
def diagnostic_result_form(request, pk):
    """Diagnost — tekshiruv natijasini kiritish/ko'rish sahifasi (tayinlagan shifokor ham faqat ko'rish uchun kira oladi)."""
    assignment = get_object_or_404(DiagnosticAssignment.objects.select_related('patient_card', 'assigned_by'), pk=pk)
    is_diagnostician = request.user.role == 'diagnostician' or request.user.is_superuser
    doctor_viewer = _get_doctor_profile(request.user)
    if not is_diagnostician and not (doctor_viewer and _doctor_card_access(doctor_viewer, assignment.patient_card)):
        return redirect('access_denied')

    if is_diagnostician and assignment.status == 'assigned':
        assignment.status = 'in_progress'
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)

    result_log, _created = DiagnosticResultLog.objects.get_or_create(
        assignment=assignment, defaults={'performed_by': request.user}
    )

    return render(request, 'patients/diagnostician/result_form.html', {
        'assignment': assignment,
        'patient': assignment.patient_card,
        'result_log': result_log,
    })


@require_POST
@login_required
def diagnostic_result_save(request, pk):
    """AJAX — diagnost natijani saqlaydi/yakunlaydi."""
    if request.user.role != 'diagnostician' and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': _("Ruxsat yo'q")}, status=403)
    assignment = get_object_or_404(DiagnosticAssignment, pk=pk)

    try:
        payload = json.loads(request.body)
    except ValueError:
        return JsonResponse({'success': False, 'error': _("Noto'g'ri so'rov")}, status=400)

    result_log, _created = DiagnosticResultLog.objects.get_or_create(
        assignment=assignment, defaults={'performed_by': request.user}
    )
    result_log.conclusion = payload.get('conclusion', '').strip()
    result_log.recommendation = payload.get('recommendation', '').strip()
    result_log.comment = payload.get('comment', '').strip()
    result_log.performed_by = request.user
    result_log.save(update_fields=['conclusion', 'recommendation', 'comment', 'performed_by'])

    if payload.get('finish'):
        assignment.status = 'done'
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)
        if assignment.assigned_by:
            _notify_doctor(assignment.assigned_by, assignment.patient_card, f"Diagnostika natijasi kiritildi: {assignment.patient_card.full_name} — {assignment.get_diagnostic_type_display()} ({request.user.get_full_name()})")

    return JsonResponse({'success': True, 'status': assignment.status})


@login_required
def diagnostic_result_print(request, pk):
    """Diagnostika natijasini chop etish uchun toza (shapkali) ko'rinish."""
    assignment = get_object_or_404(DiagnosticAssignment.objects.select_related('patient_card', 'assigned_by'), pk=pk)
    is_diagnostician = request.user.role == 'diagnostician' or request.user.is_superuser
    doctor = _get_doctor_profile(request.user)
    if not is_diagnostician and not (doctor and _doctor_card_access(doctor, assignment.patient_card)):
        return redirect('access_denied')
    result_log = assignment.result_logs.first()

    import base64
    from django.conf import settings as dj_settings
    import os as _os

    header_b64 = ''
    for header_path in [
        _os.path.join(dj_settings.STATIC_ROOT, 'img', 'hospital_header.png'),
        _os.path.join(dj_settings.BASE_DIR, 'static', 'img', 'hospital_header.png'),
    ]:
        if _os.path.exists(header_path):
            with open(header_path, 'rb') as f:
                header_b64 = base64.b64encode(f.read()).decode()
            break

    return render(request, 'patients/diagnostician/result_print.html', {
        'assignment': assignment,
        'patient': assignment.patient_card,
        'result_log': result_log,
        'header_b64': header_b64,
        'print_date': timezone.now(),
    })


@require_POST
@login_required
def diagnostic_update_status(request, pk):
    """Diagnost — tekshiruv holatini o'zgartiradi."""
    if request.user.role != 'diagnostician' and not request.user.is_superuser:
        return redirect('access_denied')
    assignment = get_object_or_404(DiagnosticAssignment, pk=pk)
    new_status = request.POST.get('status')
    if new_status in dict(DiagnosticAssignment.STATUS_CHOICES):
        assignment.status = new_status
        assignment.save(update_fields=['status'])
        _sync_billing_status(assignment)
        if new_status == 'cancelled':
            assignment.schedule_occurrences.filter(status='pending').update(status='cancelled')
        messages.success(request, _("Holat yangilandi: %(status)s") % {'status': assignment.get_status_display()})
    return redirect('diagnostic_queue')


# ==================== KONSULTATSIYALAR ====================

@require_POST
@login_required
def consultation_request_create(request, pk):
    """Shifokor — xizmatlar katalogidagi konsultatsiya xizmatini va unga biriktirilgan mutaxassis(lar)ni tanlab so'rov yuboradi."""
    from apps.services.models import Service

    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    patient = get_object_or_404(PatientCard, pk=pk)
    if not _doctor_card_access(doctor, patient):
        return redirect('access_denied')

    service = Service.objects.filter(
        pk=request.POST.get('service_id'), is_active=True, category__category_type='consultation',
    ).first()
    if not service:
        messages.error(request, _("Konsultatsiya xizmatini tanlang."))
        return redirect('doctor_patient_card', pk=pk)

    consultant_ids = request.POST.getlist('consultants')
    # Faqat shu xizmatga biriktirilgan faol shifokorlar orasidan tanlash mumkin
    consultants = list(service.assigned_doctors.filter(pk__in=consultant_ids, is_active=True))
    if not consultants:
        messages.error(request, _("Kamida bitta mutaxassisni tanlang."))
        return redirect('doctor_patient_card', pk=pk)

    consultation = ConsultationRequest.objects.create(
        patient_card=patient, requested_by=doctor, service=service,
        reason=request.POST.get('reason', '').strip(),
    )
    consultation.consultants.set(consultants)
    for consultant in consultants:
        _notify_doctor(consultant, patient, f"Sizga konsultatsiya so'rovi yuborildi: {patient.full_name} — {consultation.display_label}")
    messages.success(request, _("✅ Konsultatsiya so'rovi yuborildi: %(label)s") % {'label': consultation.display_label})
    return redirect('doctor_patient_card', pk=pk)


@require_POST
@login_required
def consultation_respond(request, pk):
    """Taklif qilingan mutaxassis — konsultatsiya so'roviga xulosa/javob yozadi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    consultation = get_object_or_404(ConsultationRequest, pk=pk)
    if not consultation.consultants.filter(pk=doctor.pk).exists() and not request.user.is_superuser:
        return redirect('access_denied')

    conclusion = request.POST.get('conclusion', '').strip()
    if not conclusion:
        messages.error(request, _("Xulosa matnini kiriting."))
        return redirect('doctor_patient_card', pk=consultation.patient_card_id)

    ConsultationResponse.objects.create(
        request=consultation, responded_by=request.user, conclusion=conclusion,
        comment=request.POST.get('comment', '').strip(),
    )
    if consultation.status == 'assigned':
        consultation.status = 'in_progress'
        consultation.save(update_fields=['status'])
    if consultation.requested_by:
        _notify_doctor(consultation.requested_by, consultation.patient_card, f"Konsultatsiya javobi keldi: {consultation.patient_card.full_name} — {consultation.display_label} ({doctor.full_name})")
    messages.success(request, _("✅ Javob yuborildi."))
    return redirect('doctor_patient_card', pk=consultation.patient_card_id)


@require_POST
@login_required
def consultation_update_status(request, pk):
    """So'rov yuborgan shifokor — konsultatsiyani yakunlaydi yoki bekor qiladi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    consultation = get_object_or_404(ConsultationRequest, pk=pk)
    if consultation.requested_by_id != doctor.pk and not request.user.is_superuser:
        return redirect('access_denied')
    new_status = request.POST.get('status')
    if new_status in dict(ConsultationRequest.STATUS_CHOICES):
        consultation.status = new_status
        consultation.save(update_fields=['status'])
        _sync_billing_status(consultation)
        if new_status == 'cancelled':
            consultation.schedule_occurrences.filter(status='pending').update(status='cancelled')
        messages.success(request, _("Holat yangilandi: %(status)s") % {'status': consultation.get_status_display()})
    return redirect('doctor_patient_card', pk=consultation.patient_card_id)


@require_POST
@login_required
def consultation_start(request, pk):
    """Taklif qilingan mutaxassis — konsultatsiya so'rovini boshlaydi, bemorga
    ko'rik qo'shish oynasiga yo'naltiriladi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    consultation = get_object_or_404(ConsultationRequest, pk=pk)
    if not consultation.consultants.filter(pk=doctor.pk).exists() and not request.user.is_superuser:
        return redirect('access_denied')

    if consultation.status == 'assigned':
        consultation.status = 'in_progress'
        consultation.save(update_fields=['status'])

    return redirect('ambulatory_consultation_form', patient_id=consultation.patient_card_id)


@require_POST
@login_required
def consultation_cancel(request, pk):
    """Taklif qilingan mutaxassis — konsultatsiya so'rovini sabab bilan bekor qiladi."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    consultation = get_object_or_404(ConsultationRequest, pk=pk)
    if not consultation.consultants.filter(pk=doctor.pk).exists() and not request.user.is_superuser:
        return redirect('access_denied')

    reason = request.POST.get('cancel_reason', '').strip()
    if not reason:
        messages.error(request, _("Bekor qilish sababini kiriting."))
        return redirect('doctor_consultation_inbox')

    consultation.status = 'cancelled'
    consultation.cancel_reason = reason
    consultation.save(update_fields=['status', 'cancel_reason'])

    if consultation.requested_by:
        _notify_doctor(
            consultation.requested_by, consultation.patient_card,
            _("Konsultatsiya so'rovi bekor qilindi: %(name)s — %(label)s. Sabab: %(reason)s") % {
                'name': consultation.patient_card.full_name,
                'label': consultation.display_label,
                'reason': reason,
            }
        )
    messages.success(request, _("So'rov bekor qilindi."))
    return redirect('doctor_consultation_inbox')


@login_required
def doctor_consultation_inbox(request):
    """Shifokor — unga taklif sifatida yuborilgan konsultatsiya so'rovlari ro'yxati."""
    doctor = _get_doctor_profile(request.user)
    if not doctor:
        return redirect('access_denied')
    qs = ConsultationRequest.objects.filter(consultants=doctor).select_related(
        'patient_card', 'requested_by'
    ).prefetch_related('responses__responded_by').order_by('-created_at')
    return render(request, 'patients/doctor/consultation_inbox.html', {
        'doctor': doctor,
        'consultations': qs,
        'pending_count': qs.filter(status='assigned').count(),
    })
